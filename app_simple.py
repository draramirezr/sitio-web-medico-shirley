# Versión simplificada del sitio web médico
# Dra. Shirley Ramírez - Ginecóloga y Obstetra

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, make_response, send_file, session
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import os
from datetime import datetime, timedelta
import hashlib
import json
import functools
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from dotenv import load_dotenv
import secrets
import re
from markupsafe import escape
from io import BytesIO
import threading
import mimetypes

# SendGrid para envío de emails (API en lugar de SMTP bloqueado por Railway)
try:
    from sendgrid import SendGridAPIClient
    from sendgrid.helpers.mail import Mail, Email, To, Content, Attachment, FileContent, FileName, FileType, Disposition, TrackingSettings, ClickTracking
    import base64
    SENDGRID_AVAILABLE = True
    print("✅ SendGrid API disponible")
except ImportError:
    SENDGRID_AVAILABLE = False
    print("⚠️ SendGrid no disponible - instalar con: pip install sendgrid")

# Importar MySQL (obligatorio)
import pymysql
pymysql.install_as_MySQLdb()

# Función para parsear MySQL URL
def parse_mysql_url(url):
    """Parsear URL de MySQL en formato: mysql://user:password@host:port/database"""
    import re
    if not url:
        return None
    
    pattern = r'mysql://([^:]+):([^@]+)@([^:]+):(\d+)/(.+)'
    match = re.match(pattern, url)
    
    if match:
        return {
            'user': match.group(1),
            'password': match.group(2),
            'host': match.group(3),
            'port': int(match.group(4)),
            'database': match.group(5),
            'charset': 'utf8mb4'
        }
    return None

# Importar sistema de optimización
try:
    from optimization_system import (
        cache_result, clear_cache, get_cache_stats, 
        optimize_database_connection, compress_response,
        add_security_headers, add_cache_headers, optimize_query,
        get_performance_stats
    )
    OPTIMIZATION_AVAILABLE = True
    print("✅ Sistema de optimización cargado")
except ImportError as e:
    print(f"⚠️ Sistema de optimización no disponible: {e}")
    OPTIMIZATION_AVAILABLE = False
    
    # Funciones dummy para fallback
    def cache_result(expiration_seconds=300):
        def decorator(func):
            return func
        return decorator
    
    def clear_cache(): pass
    def get_cache_stats(): return {}
    def compress_response(response): return response
    def add_security_headers(response): return response
    def add_cache_headers(response, max_age=3600): return response
    def optimize_query(query, params=None): return query
    def get_performance_stats(): return {}

# Importar templates de email
try:
    from email_templates import (
        template_contacto,
        template_cita,
        template_recuperacion,
        template_constancia_pdf,
        template_factura,
        template_confirmacion_cita,
        template_nueva_contrasena
    )
    EMAIL_TEMPLATES_AVAILABLE = True
except ImportError as e:
    print(f"⚠️ Email templates no disponibles: {e}")
    EMAIL_TEMPLATES_AVAILABLE = False
    
    # Funciones dummy para evitar errores
    def template_contacto(*args): return "Email template no disponible"
    def template_cita(*args): return "Email template no disponible"
    def template_recuperacion(*args): return "Email template no disponible"
    def template_constancia_pdf(*args): return "Email template no disponible"
    def template_factura(*args): return "Email template no disponible"
    def template_confirmacion_cita(*args): return "Email template no disponible"
    def template_nueva_contrasena(*args): return "Email template no disponible"

# Importar Flask-Compress de forma opcional
try:
    from flask_compress import Compress
    COMPRESS_AVAILABLE = True
except ImportError:
    COMPRESS_AVAILABLE = False
    print("⚠️  Flask-Compress no disponible. Instalar con: pip install Flask-Compress")

# Rate limiting simple (sin dependencias externas)
from collections import defaultdict
from threading import Lock
from functools import wraps

# Almacén de intentos de login
login_attempts = defaultdict(list)
login_attempts_lock = Lock()

# Importar middleware de seguridad
try:
    from security_middleware import add_security_headers, rate_limit, log_security_event
    SECURITY_MIDDLEWARE_AVAILABLE = True
except ImportError:
    SECURITY_MIDDLEWARE_AVAILABLE = False
    print("⚠️  Security middleware no disponible")

# Importar ReportLab de forma opcional
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("⚠️  ReportLab no disponible. Instalar con: pip install reportlab")

# Cargar variables de entorno
load_dotenv()

# Asegurar tipos MIME correctos para assets modernos
# (Evita que .webp salga como image/jpeg o application/octet-stream en algunos entornos)
mimetypes.add_type('image/webp', '.webp')
mimetypes.add_type('image/avif', '.avif')

app = Flask(__name__)

# ============================================
# CONFIGURACIÓN DE SEGURIDAD CRÍTICA
# ============================================

# Detectar entorno de producción (Railway/hosting)
PRODUCTION = (
    (os.getenv('FLASK_ENV') or '').lower() == 'production'
    or (os.getenv('ENV') or '').lower() == 'production'
    or (os.getenv('RAILWAY_ENVIRONMENT') or '').lower() == 'production'
)

# Clave secreta (CRÍTICO): debe ser estable SIEMPRE (sobre todo con múltiples instancias).
# Si SECRET_KEY cambia entre requests/instancias, el usuario "pierde" la sesión y vuelve al login.
#
# Estrategia:
# - Si existe SECRET_KEY, usarla.
# - Si NO existe, derivar una clave estable desde otros secretos ya configurados (MySQL/SendGrid),
#   para evitar el bug de "re-login" repetido por balanceo.
# - Si no hay nada para derivar, en producción fallar (para no desplegar roto).
def _derive_stable_secret_key() -> str:
    candidates = [
        os.getenv('MYSQL_URL', ''),
        os.getenv('MYSQLPASSWORD', ''),
        os.getenv('MYSQL_ROOT_PASSWORD', ''),
        os.getenv('MYSQL_PASSWORD', ''),
        os.getenv('SENDGRID_API_KEY', ''),
        os.getenv('RECAPTCHA_SECRET_KEY', ''),
    ]
    seed = next((c for c in candidates if c), '')
    if not seed:
        return ''
    # 64 hex chars
    return hashlib.sha256(seed.encode('utf-8')).hexdigest()

_secret_key_env = os.getenv('SECRET_KEY', '').strip()
if not _secret_key_env:
    derived = _derive_stable_secret_key()
    if derived:
        _secret_key_env = derived
        print("⚠️ SECRET_KEY no configurada: usando clave derivada estable (recomendado: configurar SECRET_KEY fija).")
    else:
        if PRODUCTION:
            raise RuntimeError(
                "Falta SECRET_KEY en producción y no hay secretos para derivarla. "
                "Configure SECRET_KEY fija en variables de entorno para evitar que las sesiones se invaliden."
            )
        _secret_key_env = secrets.token_hex(32)

app.secret_key = _secret_key_env

# Google reCAPTCHA v2 - Configuración
RECAPTCHA_SITE_KEY = os.getenv('RECAPTCHA_SITE_KEY', '')
RECAPTCHA_SECRET_KEY = os.getenv('RECAPTCHA_SECRET_KEY', '')

# Configuración de seguridad y sesiones
app.config['SESSION_COOKIE_SECURE'] = PRODUCTION  # True en producción
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_DOMAIN'] = '.draramirez.com'  # Permite cookies en www y sin www
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)
app.config['REMEMBER_COOKIE_DURATION'] = timedelta(days=7)
app.config['REMEMBER_COOKIE_SECURE'] = PRODUCTION
app.config['REMEMBER_COOKIE_HTTPONLY'] = True
app.config['REMEMBER_COOKIE_DOMAIN'] = '.draramirez.com'  # Permite cookies en www y sin www

# Si está detrás de un proxy/reverse proxy (Railway/Nginx/Cloudflare),
# esto hace que Flask vea correctamente IP/proto/host.
# Sin esto, Flask-Login puede invalidar sesión (session_protection) por cambios de IP.
try:
    from werkzeug.middleware.proxy_fix import ProxyFix
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
except Exception as _e:
    print(f"⚠️ ProxyFix no disponible o falló: {_e}")

# ============================================
# EQUIVALENTES FLASK A SETTINGS TIPO DJANGO
# - SECURE_SSL_REDIRECT
# - PREPEND_WWW
# - ALLOWED_HOSTS
# ============================================
def _env_bool(name: str, default: bool = False) -> bool:
    v = (os.getenv(name) or '').strip().lower()
    if v == '':
        return default
    return v in ('1', 'true', 't', 'yes', 'y', 'on')

# Activar solo en producción (por defecto) para evitar loops en local.
# Si necesitas forzarlo en staging/dev, define explícitamente estas variables.
SECURE_SSL_REDIRECT = _env_bool('SECURE_SSL_REDIRECT', default=PRODUCTION)
PREPEND_WWW = _env_bool('PREPEND_WWW', default=PRODUCTION)

# Lista de hosts permitidos (equivalente a Django ALLOWED_HOSTS)
_allowed_hosts_env = (os.getenv('ALLOWED_HOSTS') or '').strip()
if _allowed_hosts_env:
    ALLOWED_HOSTS = [h.strip().lower() for h in _allowed_hosts_env.split(',') if h.strip()]
else:
    # Default seguro para este proyecto
    ALLOWED_HOSTS = [
        'draramirez.com',
        'www.draramirez.com',
        'localhost',
        '127.0.0.1',
    ]

PREFERRED_WWW_HOST = os.getenv('PREFERRED_WWW_HOST', 'www.draramirez.com').strip().lower()

def _canonical_base_url() -> str:
    """
    URL base canónica para SEO (sitemap/robots/canonical).
    - En producción: por defecto https://www.draramirez.com
    - En dev: permite sobreescribir con SITE_URL/CANONICAL_BASE_URL
    """
    env = (os.getenv('CANONICAL_BASE_URL') or os.getenv('SITE_URL') or '').strip()
    if env:
        return env.rstrip('/')
    if PRODUCTION:
        host = (PREFERRED_WWW_HOST or 'www.draramirez.com').strip().lower()
        return f"https://{host}".rstrip('/')
    # Local/dev fallback
    return "http://localhost:5000"

def _canonicalize_path(path: str) -> str:
    if not path:
        return "/"
    return path if path.startswith("/") else f"/{path}"

def _request_host_no_port() -> str:
    # ProxyFix ya ajusta request.host, pero dejamos fallback por headers.
    host = (request.headers.get('X-Forwarded-Host') or request.host or '').split(',')[0].strip()
    host = host.split(':')[0].lower()
    return host

def _request_proto() -> str:
    proto = (request.headers.get('X-Forwarded-Proto') or request.scheme or '').split(',')[0].strip().lower()
    return proto

# Desactivar caché de templates para desarrollo
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

# Headers de seguridad
@app.after_request
def security_and_performance_headers(response):
    """Agregar headers de seguridad y optimización"""
    # Headers de seguridad
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    # CSP: permitir GTM/Analytics + Google Ads (DoubleClick) sin abrir de más.
    # Nota: Google Ads puede cargar scripts desde googleads.g.doubleclick.net y googleadservices.com
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' "
        "https://cdn.jsdelivr.net https://cdnjs.cloudflare.com "
        "https://www.googletagmanager.com https://www.google.com https://www.gstatic.com "
        "https://www.googleadservices.com https://googleads.g.doubleclick.net; "
        "style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://cdnjs.cloudflare.com https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com https://cdnjs.cloudflare.com; "
        "img-src 'self' data: https:; "
        "connect-src 'self' https://cdn.jsdelivr.net https://cdnjs.cloudflare.com "
        "https://www.google-analytics.com https://www.googletagmanager.com https://www.google.com "
        "https://googleads.g.doubleclick.net https://www.googleadservices.com; "
        "frame-src https://www.googletagmanager.com https://www.google.com;"
    )
    
    # Headers de caché
    if response.content_type:
        if 'text/html' in response.content_type:
            # NO cachear HTML en desarrollo para ver cambios inmediatamente
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
        elif any(ext in response.content_type for ext in ['css', 'js', 'png', 'jpg', 'jpeg', 'gif', 'svg']):
            response.headers['Cache-Control'] = 'public, max-age=2592000'  # 30 días
        else:
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    
    # Compresión de respuesta
    if OPTIMIZATION_AVAILABLE:
        response = compress_response(response)
    
    # Headers adicionales de rendimiento
    response.headers['X-Powered-By'] = 'Flask-Optimized'
    response.headers['Server'] = 'Nginx/1.18.0'  # Ocultar información del servidor
    
    return response

# Configuración adicional de sesiones
app.config['SESSION_COOKIE_NAME'] = 'drashirley_session'
app.config['SESSION_REFRESH_EACH_REQUEST'] = True  # Renovar sesión en cada request

# ============================================
# RATE LIMITING Y VALIDACIÓN DE ENTRADA
# ============================================

from collections import defaultdict
from threading import Lock
import time

# Rate limiting simple
request_counts = defaultdict(list)
rate_limit_lock = Lock()

def cleanup_old_rate_limits():
    """Limpiar rate limits antiguos para evitar memory leaks"""
    current_time = time.time()
    with rate_limit_lock:
        # Eliminar entradas completamente vacías después de limpiar
        keys_to_delete = []
        for key in list(request_counts.keys()):
            # Filtrar requests antiguos (más de 10 minutos)
            request_counts[key] = [
                req_time for req_time in request_counts[key]
                if current_time - req_time < 600  # 10 minutos
            ]
            # Si está vacío, marcarlo para eliminación
            if not request_counts[key]:
                keys_to_delete.append(key)
        
        # Eliminar claves vacías
        for key in keys_to_delete:
            del request_counts[key]

def rate_limit(max_requests=10, window=60):
    """Decorador para rate limiting"""
    def decorator(f):
        @functools.wraps(f)
        def wrapper(*args, **kwargs):
            client_ip = request.remote_addr
            current_time = time.time()
            
            with rate_limit_lock:
                # Limpiar requests antiguos
                request_counts[client_ip] = [
                    req_time for req_time in request_counts[client_ip]
                    if current_time - req_time < window
                ]
                
                # Verificar límite
                if len(request_counts[client_ip]) >= max_requests:
                    return jsonify({'error': 'Rate limit exceeded'}), 429
                
                # Agregar request actual
                request_counts[client_ip].append(current_time)
            
            return f(*args, **kwargs)
        return wrapper
    return decorator

def sanitize_input(text):
    """Sanitizar entrada de usuario"""
    if not text:
        return ""
    return escape(str(text).strip())

def validate_email(email):
    """Validar formato de email"""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

# Configurar Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = None  # Desactivar mensaje automático para evitar duplicación
login_manager.login_message_category = 'warning'
# IMPORTANTE: "strong" puede forzar logout si cambia la IP (común detrás de proxies/CDNs).
# Usamos "basic" para reducir logouts inesperados sin perder seguridad de cookie.
login_manager.session_protection = 'basic'
login_manager.refresh_view = 'login'  # Vista para refrescar autenticación si expira

# Configuración de compresión Gzip/Brotli para máxima velocidad
if COMPRESS_AVAILABLE:
    app.config['COMPRESS_MIMETYPES'] = [
        'text/html', 'text/css', 'text/xml', 'application/json',
        'application/javascript', 'application/x-javascript', 'text/javascript'
    ]
    app.config['COMPRESS_LEVEL'] = 6  # Balance entre velocidad y compresión
    app.config['COMPRESS_MIN_SIZE'] = 500  # Comprimir archivos > 500 bytes
    Compress(app)
    print("✅ Flask-Compress activado: Compresión Gzip/Brotli habilitada")
else:
    print("⚠️  Flask-Compress no activado: El sitio funcionará sin compresión automática")

# Headers de seguridad y optimización de cache
@app.after_request
def add_security_and_cache_headers(response):
    """Agregar headers de seguridad HTTP y cache optimizado"""
    # Seguridad
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'geolocation=(), microphone=(), camera=()'
    
    # Cache agresivo para recursos estáticos (máxima velocidad)
    if request.path.startswith('/static/'):
        # Cache de 1 año para CSS, JS, imágenes
        if any(request.path.endswith(ext) for ext in ['.css', '.js', '.jpg', '.jpeg', '.png', '.gif', '.ico', '.svg', '.woff', '.woff2', '.ttf']):
            response.headers['Cache-Control'] = 'public, max-age=31536000, immutable'
            # Usar datetime.now(timezone.utc) en lugar de utcnow() (deprecated)
            from datetime import timezone
            response.headers['Expires'] = (datetime.now(timezone.utc) + timedelta(days=365)).strftime('%a, %d %b %Y %H:%M:%S GMT')
        else:
            response.headers['Cache-Control'] = 'public, max-age=86400'  # 1 día para otros
    else:
        # NO cachear páginas que requieren sesión activa (login, admin, facturación)
        if any(request.path.startswith(path) for path in ['/login', '/admin', '/facturacion']):
            response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
        # Para otras páginas HTML: cache corto con revalidación
        elif response.content_type and 'text/html' in response.content_type:
            response.headers['Cache-Control'] = 'public, max-age=300, must-revalidate'  # 5 minutos
        else:
            response.headers['Cache-Control'] = 'no-cache, must-revalidate'
    
    # ETag para validación eficiente de cache
    if not response.headers.get('ETag') and response.content_type and 'text/html' in response.content_type:
        response.add_etag()
    
    return response

# ============ CONFIGURACIÓN DE EMAIL - SENDGRID API ============
# SendGrid API (Railway bloquea SMTP, usamos API en su lugar)

# API Key de SendGrid (limpiar comillas que Railway agrega automáticamente)
raw_api_key = os.getenv('SENDGRID_API_KEY')
SENDGRID_API_KEY = raw_api_key.strip().strip('"').strip("'") if raw_api_key else None

# Email remitente (debe estar verificado en SendGrid)
raw_email_from = os.getenv('EMAIL_FROM', 'dra.ramirezr@gmail.com')
EMAIL_FROM = raw_email_from.strip().strip('"').strip("'") if raw_email_from else 'dra.ramirezr@gmail.com'

# Email destinatario (donde llegan las notificaciones)
raw_email_dest = os.getenv('EMAIL_DESTINATARIO', 'dra.ramirezr@gmail.com')
EMAIL_DESTINATARIO = raw_email_dest.strip().strip('"').strip("'") if raw_email_dest else 'dra.ramirezr@gmail.com'

# Verificar configuración
EMAIL_CONFIGURED = bool(SENDGRID_API_KEY and SENDGRID_AVAILABLE)

if EMAIL_CONFIGURED:
    print(f"✅ Email configurado con SendGrid API")
    print(f"   📧 From: {EMAIL_FROM}")
    print(f"   📬 Notificaciones a: {EMAIL_DESTINATARIO}")
    print(f"   🔑 API Key: {SENDGRID_API_KEY[:10]}...{SENDGRID_API_KEY[-4:] if SENDGRID_API_KEY else ''}")
else:
    print("⚠️ Email NO configurado - revisa SENDGRID_API_KEY")

# Función para limpiar comillas de variables de entorno
def clean_env_var(var_name, default=''):
    """Limpiar comillas que Railway puede agregar automáticamente a las variables"""
    original = os.getenv(var_name, default)
    value = original
    if value and isinstance(value, str):
        # Eliminar comillas dobles y simples al inicio y final
        value = value.strip().strip('"').strip("'").strip()
    # DEBUG: mostrar transformación
    if original != value:
        print(f"   🔧 {var_name}: '{original}' → '{value}'")
    return value if value else default

# Configuración de la base de datos (SOLO MYSQL)
print("🔍 DEBUG: Configurando conexión MySQL...")

# Intentar primero con MYSQL_URL (local o de Railway)
mysql_url = os.getenv('MYSQL_URL', '')
print(f"   🔍 RAW MYSQL_URL: '{mysql_url[:50] if mysql_url else 'NO DEFINIDA'}...'")

if mysql_url:
    # Opción 1: Usar MYSQL_URL
    parsed_config = parse_mysql_url(mysql_url)
    
    if parsed_config:
        DATABASE_CONFIG = parsed_config
        DATABASE_TYPE = 'mysql'
        print("✅ Configurado para usar MySQL (usando MYSQL_URL)")
        print(f"   🔌 Conectando a: {DATABASE_CONFIG['host']}")
        print(f"   👤 Usuario: {DATABASE_CONFIG['user']}")
        print(f"   📁 Base de datos: {DATABASE_CONFIG['database']}")
    else:
        print("❌ Error: MYSQL_URL inválida")
        raise Exception("MYSQL_URL no pudo ser parseada")
else:
    # Opción 2: Usar variables individuales (MYSQLHOST, MYSQLUSER, etc.)
    mysqlhost = os.getenv('MYSQLHOST', os.getenv('MYSQL_HOST', ''))
    mysqluser = os.getenv('MYSQLUSER', os.getenv('MYSQL_USER', ''))
    mysqlpassword = os.getenv('MYSQLPASSWORD', os.getenv('MYSQL_ROOT_PASSWORD', os.getenv('MYSQL_PASSWORD', '')))
    mysqldatabase = os.getenv('MYSQLDATABASE', os.getenv('MYSQL_DATABASE', ''))
    
    print(f"   🔍 RAW MYSQLHOST: '{mysqlhost if mysqlhost else 'NO DEFINIDA'}'")
    print(f"   🔍 RAW MYSQLUSER: '{mysqluser if mysqluser else 'NO DEFINIDA'}'")
    print(f"   🔍 RAW MYSQLDATABASE: '{mysqldatabase if mysqldatabase else 'NO DEFINIDA'}'")
    
    if not mysqlhost or not mysqluser or not mysqldatabase:
        print("❌ Error: Variables MySQL incompletas")
        raise Exception("Configura MYSQL_URL o las variables MYSQL_HOST, MYSQL_USER, MYSQL_PASSWORD, MYSQL_DATABASE")
    
    # Limpiar comillas si existen
    mysql_host = clean_env_var('MYSQLHOST', clean_env_var('MYSQL_HOST', ''))
    mysql_user = clean_env_var('MYSQLUSER', clean_env_var('MYSQL_USER', ''))
    mysql_password = clean_env_var('MYSQLPASSWORD', clean_env_var('MYSQL_ROOT_PASSWORD', clean_env_var('MYSQL_PASSWORD', '')))
    mysql_database = clean_env_var('MYSQLDATABASE', clean_env_var('MYSQL_DATABASE', ''))
    
    DATABASE_CONFIG = {
        'host': mysql_host,
        'user': mysql_user,
        'password': mysql_password,
        'database': mysql_database,
        'charset': 'utf8mb4'
    }
    DATABASE_TYPE = 'mysql'
    print("✅ Configurado para usar MySQL (usando variables individuales)")
    print(f"   🔌 Conectando a: {DATABASE_CONFIG['host']}")
    print(f"   👤 Usuario: {DATABASE_CONFIG['user']}")
    print(f"   📁 Base de datos: {DATABASE_CONFIG['database']}")

# Funciones de validación y sanitización
def sanitize_input(text, max_length=500):
    """Sanitizar entrada de texto"""
    if not text:
        return ""
    text = str(text).strip()
    text = re.sub(r'<[^>]*>', '', text)  # Remover tags HTML
    return text[:max_length]

# =========================
# VALIDACIONES DE INTEGRIDAD (FACTURACIÓN)
# =========================
# Permitidos:
# - Nombres: letras (incl. acentos), espacios y punto
# - Servicios / ARS / Centro: letras (incl. acentos), espacios, /, -, .
# - NSS: números y guiones
# - Autorización: alfanumérico y guiones

_RE_NSS = re.compile(r'^[0-9\-]+$')
_RE_AUTORIZACION = re.compile(r'^[A-Z0-9\-]+$')
_RE_NOMBRE = re.compile(r'^[A-ZÁÉÍÓÚÜÑ\s\.\-]+$')
_RE_TEXTO_GENERAL = re.compile(r'^[A-ZÁÉÍÓÚÜÑ\s\.\-/]+$')
_RE_RNC = re.compile(r'^[0-9\-]+$')

def _upper_clean_spaces(value: str) -> str:
    value = (value or '').strip().upper()
    value = re.sub(r'\s+', ' ', value).strip()
    return value

def _validate_or_raise(value: str, pattern: re.Pattern, field_name: str, idx=None, max_len=None, required: bool = True) -> str:
    """Valida y retorna valor (ya limpiado). Lanza ValueError si inválido."""
    v = _upper_clean_spaces(value)
    if max_len is not None and len(v) > max_len:
        prefix = f'Línea {idx}: ' if idx else ''
        raise ValueError(f'{prefix}{field_name} supera el máximo de {max_len} caracteres')
    if not v:
        if not required:
            return ''
        prefix = f'Línea {idx}: ' if idx else ''
        raise ValueError(f'{prefix}{field_name} es obligatorio')
    if not pattern.fullmatch(v):
        prefix = f'Línea {idx}: ' if idx else ''
        raise ValueError(f'{prefix}{field_name} contiene caracteres no permitidos')
    return v

def validate_email(email):
    """Validar formato de email"""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None

def validate_phone(phone):
    """Validar formato de teléfono"""
    phone = re.sub(r'[^\d+]', '', phone)
    return len(phone) >= 10 and len(phone) <= 15

def validar_password_segura(password):
    """Validar que la contraseña cumpla con los requisitos de seguridad"""
    errores = []
    
    if len(password) < 8:
        errores.append("Mínimo 8 caracteres")
    
    if not re.search(r'[A-Z]', password):
        errores.append("Al menos una mayúscula")
    
    if not re.search(r'[a-z]', password):
        errores.append("Al menos una minúscula")
    
    if not re.search(r'\d', password):
        errores.append("Al menos un número")
    
    if not re.search(r'[!@#$%^&*(),.?":{}|<>]', password):
        errores.append("Al menos un carácter especial")
    
    return errores

# Filtro personalizado para formatear moneda
@app.template_filter('formato_moneda')
def formato_moneda(valor):
    """Formatear números con coma para miles y punto para decimales
    Ejemplo: 1000.50 -> 1,000.50
    """
    try:
        # Convertir a float si es necesario
        if isinstance(valor, str):
            valor = float(valor)
        # Formatear con 2 decimales, coma para miles y punto para decimales
        return "{:,.2f}".format(float(valor))
    except (ValueError, TypeError):
        return "0.00"

def obtener_fecha_rd():
    """Obtener fecha actual en zona horaria de República Dominicana (UTC-4)"""
    from datetime import datetime, timezone, timedelta
    tz_rd = timezone(timedelta(hours=-4))
    return datetime.now(tz_rd).date()

def formato_fecha_pdf(fecha):
    """Formatear fecha a dd/mm/yyyy para PDFs
    
    Args:
        fecha: String, date o datetime
        
    Returns:
        String en formato dd/mm/yyyy
    """
    if not fecha:
        return ''
    
    try:
        # Si es string
        if isinstance(fecha, str):
            # Intentar parsear yyyy-mm-dd
            try:
                fecha_obj = datetime.strptime(fecha, '%Y-%m-%d')
                return fecha_obj.strftime('%d/%m/%Y')
            except:
                # Intentar otros formatos
                for fmt in ['%Y-%m-%d %H:%M:%S', '%d/%m/%Y']:
                    try:
                        fecha_obj = datetime.strptime(fecha, fmt)
                        return fecha_obj.strftime('%d/%m/%Y')
                    except:
                        continue
                return str(fecha)
        
        # Si es objeto datetime o date
        elif hasattr(fecha, 'strftime'):
            return fecha.strftime('%d/%m/%Y')
        
        return str(fecha)
    except:
        return str(fecha) if fecha else ''

# Filtro personalizado para obtener el nombre del día de la semana en español
@app.template_filter('nombre_dia')
def nombre_dia(fecha_str):
    """Convertir fecha a nombre del día en español
    Ejemplo: '2025-10-30' -> 'Jueves'
    """
    try:
        # Mapeo de días en español
        dias_espanol = {
            0: 'Lunes',
            1: 'Martes',
            2: 'Miércoles',
            3: 'Jueves',
            4: 'Viernes',
            5: 'Sábado',
            6: 'Domingo'
        }
        
        # Convertir string a datetime
        if isinstance(fecha_str, str):
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d')
        else:
            fecha = fecha_str
        
        # Obtener número del día (0=Lunes, 6=Domingo)
        dia_numero = fecha.weekday()
        
        # Retornar nombre en español
        return dias_espanol.get(dia_numero, '')
    except (ValueError, TypeError, AttributeError):
        return ''

# ============================================
# FUNCIONES DE GOOGLE reCAPTCHA
# ============================================

def verificar_recaptcha(response_token):
    """
    Verificar el token de reCAPTCHA v3 con Google
    Retorna True si es válido y el score >= 0.5, False si no
    """
    if not RECAPTCHA_SECRET_KEY or not response_token:
        return False
    
    import requests
    
    try:
        # Hacer petición a Google para verificar el token
        verify_url = 'https://www.google.com/recaptcha/api/siteverify'
        data = {
            'secret': RECAPTCHA_SECRET_KEY,
            'response': response_token
        }
        
        response = requests.post(verify_url, data=data, timeout=5)
        result = response.json()
        
        print(f"🔐 reCAPTCHA v3 Response: {result}")
        
        # Para reCAPTCHA v3, verificar success y score
        if result.get('success'):
            score = result.get('score', 0)
            print(f"   Score: {score} | Action: {result.get('action')}")
            # Score >= 0.5 se considera humano (ajustable según necesidad)
            # 1.0 = muy probablemente humano, 0.0 = muy probablemente bot
            return score >= 0.5
        else:
            print(f"   Error codes: {result.get('error-codes', [])}")
            return False
            
    except Exception as e:
        print(f"❌ Error al verificar reCAPTCHA: {e}")
        return False

@app.context_processor
def inject_recaptcha():
    """Inyectar la Site Key de reCAPTCHA en todos los templates"""
    return {
        'RECAPTCHA_SITE_KEY': RECAPTCHA_SITE_KEY
    }

# Filtro personalizado para formatear fechas a dd/mm/yyyy
@app.template_filter('fecha_es')
def fecha_es_filter(fecha):
    """Convertir fecha a formato dd/mm/yyyy (español)"""
    if not fecha:
        return ''
    
    try:
        # Si es string
        if isinstance(fecha, str):
            # Intentar parsear diferentes formatos
            for fmt in ['%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y']:
                try:
                    fecha_obj = datetime.strptime(fecha, fmt)
                    return fecha_obj.strftime('%d/%m/%Y')
                except:
                    continue
            return fecha  # Si no se puede convertir, devolver original
        
        # Si es objeto datetime o date
        elif hasattr(fecha, 'strftime'):
            return fecha.strftime('%d/%m/%Y')
        
        return str(fecha)
    except:
        return str(fecha) if fecha else ''

# Modelo de Usuario para Flask-Login
class User(UserMixin):
    def __init__(self, id, nombre, email, perfil):
        self.id = id
        self.nombre = nombre
        self.email = email
        self.perfil = perfil

@login_manager.user_loader
def load_user(user_id):
    """Cargar usuario desde la base de datos"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        # Verificar que el usuario existe Y está activo (MySQL)
        cursor.execute('SELECT id, nombre, email, perfil, activo, password_temporal FROM usuarios WHERE id = %s AND activo = 1', (user_id,))
        user_data = cursor.fetchone()
        conn.close()
        
        if user_data:
            # Si tiene password_temporal=1, significa que su contraseña fue cambiada
            # y debe cambiarla en su próxima petición
            user = User(
                id=user_data['id'], 
                nombre=user_data['nombre'], 
                email=user_data['email'], 
                perfil=user_data['perfil']
            )
            # Guardar flag para verificar en cada petición
            user.password_temporal = user_data.get('password_temporal', 0)
            return user
        return None
    except Exception as e:
        print(f"Error en load_user: {e}")
        return None

@app.before_request
def check_password_temporal():
    """
    Verificar si el usuario tiene contraseña temporal y forzar cambio
    """
    # Seguridad/SEO: ALLOWED_HOSTS + redirección HTTPS/WWW (equivalentes tipo Django)
    req_host = _request_host_no_port()
    if req_host and (req_host not in ALLOWED_HOSTS):
        return ("Host no permitido", 400)

    if PRODUCTION:
        proto = _request_proto()
        path = request.path or '/'
        qs = request.query_string.decode('utf-8', errors='ignore') if request.query_string else ''
        path_qs = f"{path}?{qs}" if qs else path

        # Forzar www primero (si aplica)
        if PREPEND_WWW and req_host == 'draramirez.com' and PREFERRED_WWW_HOST:
            code = 301 if request.method in ('GET', 'HEAD') else 308
            scheme = 'https' if (SECURE_SSL_REDIRECT or proto == 'https') else (proto or 'http')
            return redirect(f"{scheme}://{PREFERRED_WWW_HOST}{path_qs}", code=code)

        # Forzar HTTPS (si aplica)
        if SECURE_SSL_REDIRECT and proto and proto != 'https':
            code = 301 if request.method in ('GET', 'HEAD') else 308
            return redirect(f"https://{req_host}{path_qs}", code=code)
    
    # Verificar contraseña temporal (solo para usuarios autenticados)
    if current_user.is_authenticated:
        # Excluir rutas de cambio de contraseña y logout para evitar loops
        if request.endpoint not in ['cambiar_password_obligatorio', 'logout', 'static']:
            # Verificar si tiene password_temporal
            if hasattr(current_user, 'password_temporal') and current_user.password_temporal:
                # Redirigir a cambio de contraseña obligatorio
                from flask import session
                session['cambio_password_usuario_id'] = current_user.id
                session['cambio_password_email'] = current_user.email
                return redirect(url_for('cambiar_password_obligatorio'))

def init_db():
    """Inicializar la base de datos"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    print(f"✅ Base de datos conectada: {DATABASE_TYPE}")
    
    # Tabla de servicios
    cursor.execute(adapt_sql_for_database('''
        CREATE TABLE IF NOT EXISTS services (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            description TEXT NOT NULL,
            icon TEXT,
            price_range TEXT,
            duration TEXT,
            active BOOLEAN DEFAULT 1
        )
    '''))
    
    # Tabla de usuarios para autenticación
    cursor.execute(adapt_sql_for_database('''
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            email VARCHAR(255) UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            perfil TEXT NOT NULL CHECK(perfil IN ('Administrador', 'Nivel 2', 'Registro de Facturas')),
            activo BOOLEAN DEFAULT 1,
            password_temporal BOOLEAN DEFAULT 0,
            token_recuperacion TEXT,
            token_expiracion TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            last_login TIMESTAMP
        )
    '''))
    
    # Tabla de testimonios
    cursor.execute(adapt_sql_for_database('''
        CREATE TABLE IF NOT EXISTS testimonials (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            patient_name TEXT NOT NULL,
            patient_initials TEXT,
            testimonial_text TEXT NOT NULL,
            rating INTEGER DEFAULT 5,
            approved BOOLEAN DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            display_date DATE
        )
    '''))
    
    # Tabla de mensajes de contacto
    cursor.execute(adapt_sql_for_database('''
        CREATE TABLE IF NOT EXISTS contact_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email VARCHAR(255) NOT NULL,
            phone TEXT,
            subject TEXT NOT NULL,
            message TEXT NOT NULL,
            `read` BOOLEAN DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    '''))
    
    # Tabla de lista negra de emails (blacklist)
    cursor.execute(adapt_sql_for_database('''
        CREATE TABLE IF NOT EXISTS email_blacklist (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email VARCHAR(255) UNIQUE NOT NULL,
            blocked_by VARCHAR(255) NOT NULL,
            reason TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    '''))
    
    # Tabla de configuración del sitio (para temas y configuraciones generales)
    cursor.execute(adapt_sql_for_database('''
        CREATE TABLE IF NOT EXISTS configuracion_sitio (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            clave VARCHAR(100) UNIQUE NOT NULL,
            valor VARCHAR(255),
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    '''))
    
    # Insertar configuración por defecto si no existe
    try:
        cursor.execute(adapt_sql_for_database('''
            INSERT INTO configuracion_sitio (clave, valor) 
            VALUES ('tema_principal', 'original')
        '''))
    except:
        pass  # Ya existe, no hacer nada
    
    # Tabla de citas
    cursor.execute(adapt_sql_for_database('''
        CREATE TABLE IF NOT EXISTS appointments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            first_name TEXT NOT NULL,
            last_name TEXT NOT NULL,
            email VARCHAR(255),
            phone TEXT NOT NULL,
            appointment_date TEXT,
            appointment_time TEXT,
            appointment_type TEXT NOT NULL,
            medical_insurance TEXT NOT NULL,
            emergency_datetime TEXT,
            reason TEXT,
            status VARCHAR(50) DEFAULT 'pending',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    '''))
    
    # Verificar si la columna medical_insurance existe, si no, agregarla
    try:
        cursor.execute("SELECT medical_insurance FROM appointments LIMIT 1")
    except:
        cursor.execute("ALTER TABLE appointments ADD COLUMN medical_insurance TEXT NOT NULL DEFAULT 'Otros'")
        print("✅ Columna 'medical_insurance' agregada a la tabla appointments")
    
    # Verificar si la columna appointment_time existe, si no, agregarla
    try:
        cursor.execute("SELECT appointment_time FROM appointments LIMIT 1")
    except:
        cursor.execute("ALTER TABLE appointments ADD COLUMN appointment_time VARCHAR(10)")
        print("✅ Columna 'appointment_time' agregada a la tabla appointments")
    
    # Verificar si la columna emergency_datetime existe, si no, agregarla
    try:
        cursor.execute("SELECT emergency_datetime FROM appointments LIMIT 1")
    except:
        cursor.execute("ALTER TABLE appointments ADD COLUMN emergency_datetime VARCHAR(50)")
        print("✅ Columna 'emergency_datetime' agregada a la tabla appointments")
    
    # Tabla de tratamientos estéticos ginecológicos
    cursor.execute(adapt_sql_for_database('''
        CREATE TABLE IF NOT EXISTS aesthetic_treatments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            description TEXT NOT NULL,
            icon TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    '''))
    
    # Tabla de contador de visitas
    cursor.execute(adapt_sql_for_database('''
        CREATE TABLE IF NOT EXISTS site_visits (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            total_visits INTEGER DEFAULT 0,
            last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    '''))
    
    # Inicializar contador si no existe (MySQL)
    cursor.execute('INSERT IGNORE INTO site_visits (id, total_visits) VALUES (1, 0)')
    
    # ============ TABLAS DE FACTURACIÓN ============
    
    # Tabla de ARS (Administradoras de Riesgos de Salud)
    cursor.execute(adapt_sql_for_database('''
        CREATE TABLE IF NOT EXISTS ars (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre_ars VARCHAR(255) NOT NULL,
            rnc TEXT NOT NULL,
            activo BOOLEAN DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    '''))
    
    # Tabla de Médicos
    cursor.execute(adapt_sql_for_database('''
        CREATE TABLE IF NOT EXISTS medicos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre VARCHAR(255) NOT NULL,
            especialidad TEXT NOT NULL,
            cedula VARCHAR(50) NOT NULL UNIQUE,
            email VARCHAR(255),
            activo BOOLEAN DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    '''))
    
    # Agregar columna email si no existe
    try:
        cursor.execute("SELECT email FROM medicos LIMIT 1")
    except:
        cursor.execute("ALTER TABLE medicos ADD COLUMN email VARCHAR(255)")
        print("✅ Columna 'email' agregada a la tabla medicos")
    
    # Agregar columna exequatur si no existe
    try:
        cursor.execute("SELECT exequatur FROM medicos LIMIT 1")
    except:
        cursor.execute("ALTER TABLE medicos ADD COLUMN exequatur TEXT")
        print("✅ Columna 'exequatur' agregada a la tabla medicos")
    
    # Agregar columna factura si no existe (indica si está habilitado para facturar)
    try:
        cursor.execute("SELECT factura FROM medicos LIMIT 1")
    except:
        cursor.execute("ALTER TABLE medicos ADD COLUMN factura BOOLEAN DEFAULT 0")
        print("✅ Columna 'factura' agregada a la tabla medicos")
    
    # Tabla de Centros Médicos
    cursor.execute(adapt_sql_for_database('''
        CREATE TABLE IF NOT EXISTS centros_medicos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre VARCHAR(255) NOT NULL,
            direccion TEXT NOT NULL,
            rnc VARCHAR(50) NOT NULL,
            telefono VARCHAR(50),
            activo BOOLEAN DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    '''))
    
    # Tabla de Relación Médico-Centro (muchos a muchos)
    cursor.execute(adapt_sql_for_database('''
        CREATE TABLE IF NOT EXISTS medico_centro (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            medico_id INTEGER NOT NULL,
            centro_id INTEGER NOT NULL,
            activo BOOLEAN DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (medico_id) REFERENCES medicos(id),
            FOREIGN KEY (centro_id) REFERENCES centros_medicos(id),
            UNIQUE(medico_id, centro_id)
        )
    '''))
    
    # Agregar columnas ncf_id y ncf_numero a facturas si no existen
    try:
        cursor.execute("SELECT ncf_id FROM facturas LIMIT 1")
    except:
        cursor.execute("ALTER TABLE facturas ADD COLUMN ncf_id INTEGER")
        print("✅ Columna 'ncf_id' agregada a la tabla facturas")
    
    try:
        cursor.execute("SELECT ncf_numero FROM facturas LIMIT 1")
    except:
        cursor.execute("ALTER TABLE facturas ADD COLUMN ncf_numero TEXT")
        print("✅ Columna 'ncf_numero' agregada a la tabla facturas")
    
    # Agregar columna fecha_fin a NCF si no existe
    try:
        cursor.execute("SELECT fecha_fin FROM ncf LIMIT 1")
    except:
        cursor.execute("ALTER TABLE ncf ADD COLUMN fecha_fin DATE")
        print("✅ Columna 'fecha_fin' agregada a la tabla ncf")
    
    # Agregar columna password_temporal a usuarios si no existe
    try:
        cursor.execute("SELECT password_temporal FROM usuarios LIMIT 1")
    except:
        cursor.execute("ALTER TABLE usuarios ADD COLUMN password_temporal BOOLEAN DEFAULT 0")
        print("✅ Columna 'password_temporal' agregada a la tabla usuarios")
    
    # Crear índices para optimizar consultas
    # NOTA: Los índices ya se crean en el script SQL y en create_database_indexes()
    # MySQL no soporta CREATE INDEX IF NOT EXISTS
    # cursor.execute("CREATE INDEX IF NOT EXISTS idx_facturas_detalle_estado ON facturas_detalle(estado)")
    # cursor.execute("CREATE INDEX IF NOT EXISTS idx_facturas_detalle_factura_id ON facturas_detalle(factura_id)")
    # cursor.execute("CREATE INDEX IF NOT EXISTS idx_facturas_detalle_medico_id ON facturas_detalle(medico_id)")
    # cursor.execute("CREATE INDEX IF NOT EXISTS idx_facturas_detalle_ars_id ON facturas_detalle(ars_id)")
    # cursor.execute("CREATE INDEX IF NOT EXISTS idx_facturas_fecha ON facturas(fecha_factura)")
    # cursor.execute("CREATE INDEX IF NOT EXISTS idx_appointments_date ON appointments(appointment_date)")
    # cursor.execute("CREATE INDEX IF NOT EXISTS idx_contact_messages_read ON contact_messages(read)")
    print("✅ Índices de base de datos creados/verificados")
    
    # Tabla de Código ARS (relación médico-ars con su código)
    cursor.execute(adapt_sql_for_database('''
        CREATE TABLE IF NOT EXISTS codigo_ars (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            medico_id INTEGER NOT NULL,
            ars_id INTEGER NOT NULL,
            codigo_ars TEXT NOT NULL,
            activo BOOLEAN DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (medico_id) REFERENCES medicos(id),
            FOREIGN KEY (ars_id) REFERENCES ars(id),
            UNIQUE(medico_id, ars_id)
        )
    '''))
    
    # Tabla de Tipos de Servicios
    cursor.execute(adapt_sql_for_database('''
        CREATE TABLE IF NOT EXISTS tipos_servicios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            descripcion TEXT NOT NULL,
            precio_base REAL DEFAULT 0,
            activo BOOLEAN DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    '''))
    
    # Tabla de NCF (Números de Comprobante Fiscal)
    cursor.execute(adapt_sql_for_database('''
        CREATE TABLE IF NOT EXISTS ncf (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipo TEXT NOT NULL,
            prefijo TEXT NOT NULL,
            tamaño INTEGER NOT NULL,
            ultimo_numero INTEGER DEFAULT 0,
            activo BOOLEAN DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    '''))
    
    # Tabla de Pacientes (Maestra)
    cursor.execute(adapt_sql_for_database('''
        CREATE TABLE IF NOT EXISTS pacientes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nss VARCHAR(50) NOT NULL,
            nombre VARCHAR(255) NOT NULL,
            ars_id INTEGER,
            activo BOOLEAN DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (ars_id) REFERENCES ars(id),
            UNIQUE(nss, ars_id)
        )
    '''))
    
    # Tabla de Facturas (Encabezado)
    try:
        cursor.execute(adapt_sql_for_database('''
            CREATE TABLE IF NOT EXISTS facturas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                numero_factura TEXT,
                ncf TEXT,
                fecha_factura DATE NOT NULL,
                medico_id INTEGER NOT NULL,
                ars_id INTEGER NOT NULL,
                total REAL DEFAULT 0,
                observaciones TEXT,
                activo BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (medico_id) REFERENCES medicos(id),
                FOREIGN KEY (ars_id) REFERENCES ars(id)
            )
        '''))
        print("✅ Tabla 'facturas' creada/verificada")
    except Exception as e:
        print(f"❌ Error creando tabla facturas: {e}")
        raise
    
    # Tabla de Detalle de Facturas (Líneas)
    cursor.execute(adapt_sql_for_database('''
        CREATE TABLE IF NOT EXISTS facturas_detalle (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            factura_id INTEGER,
            paciente_id INTEGER NOT NULL,
            nss VARCHAR(50) NOT NULL,
            nombre_paciente TEXT NOT NULL,
            fecha_servicio DATE NOT NULL,
            autorizacion TEXT,
            servicio_id INTEGER NOT NULL,
            descripcion_servicio TEXT NOT NULL,
            monto REAL NOT NULL,
            medico_id INTEGER,
            medico_consulta INTEGER NOT NULL,
            ars_id INTEGER NOT NULL,
            estado VARCHAR(50) DEFAULT 'pendiente',
            activo BOOLEAN DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (factura_id) REFERENCES facturas(id),
            FOREIGN KEY (paciente_id) REFERENCES pacientes(id),
            FOREIGN KEY (servicio_id) REFERENCES tipos_servicios(id),
            FOREIGN KEY (medico_id) REFERENCES medicos(id),
            FOREIGN KEY (medico_consulta) REFERENCES medicos(id),
            FOREIGN KEY (ars_id) REFERENCES ars(id)
        )
    '''))
    
    # Crear usuario por defecto si no existe
    cursor.execute('SELECT COUNT(*) as count FROM usuarios')
    result = cursor.fetchone()
    count = result['count'] if isinstance(result, dict) else result[0]
    
    if count == 0:
        # Usuario por defecto: ing.fpaula@gmail.com - Francisco Paula
        # Normalizar email a lowercase para evitar problemas de case-sensitivity
        admin_email = 'ing.fpaula@gmail.com'.lower()
        password_hash = generate_password_hash('2416Xpos@')
        cursor.execute('''
            INSERT INTO usuarios (nombre, email, password_hash, perfil, activo, password_temporal)
            VALUES (%s, %s, %s, %s, 1, 0)
        ''', ('Francisco Paula', admin_email, password_hash, 'Administrador'))
        print(f"✅ Usuario por defecto creado: {admin_email}")
    
    conn.commit()
    conn.close()
    print("✅ Base de datos inicializada correctamente")
    
    # Crear índices críticos para optimización
    create_database_indexes()

def adapt_sql_for_database(sql):
    """Adaptar consultas SQL para MySQL"""
    # Cambios específicos para MySQL
    sql = sql.replace('INTEGER PRIMARY KEY AUTOINCREMENT', 'INT AUTO_INCREMENT PRIMARY KEY')
    sql = sql.replace('BOOLEAN', 'TINYINT(1)')
    sql = sql.replace('TIMESTAMP DEFAULT CURRENT_TIMESTAMP', 'TIMESTAMP DEFAULT CURRENT_TIMESTAMP')
    sql = sql.replace('CURRENT_TIMESTAMP', 'NOW()')
    sql = sql.replace('TEXT', 'TEXT')
    sql = sql.replace('REAL', 'DECIMAL(10,2)')
    sql = sql.replace('BLOB', 'LONGBLOB')
    sql = sql.replace('INTEGER', 'INT')
    return sql

class MySQLConnectionWrapper:
    """Wrapper para hacer que PyMySQL se comporte como SQLite con conn.execute()"""
    def __init__(self, connection):
        self._conn = connection
        self._cursor = None
    
    def execute(self, query, params=None):
        """Ejecutar query y devolver cursor - ESCAPA % en parámetros"""
        self._cursor = self._conn.cursor()
        if params:
            # ✅ SOLUCIÓN: Escapar % en los parámetros para evitar interpolación de PyMySQL
            escaped_params = []
            for param in params:
                if isinstance(param, str) and '%' in param:
                    # Duplicar % para escaparlo: % -> %%
                    escaped_params.append(param.replace('%', '%%'))
                else:
                    escaped_params.append(param)
            self._cursor.execute(query, tuple(escaped_params))
        else:
            self._cursor.execute(query)
        return self._cursor
    
    def commit(self):
        return self._conn.commit()
    
    def rollback(self):
        return self._conn.rollback()
    
    def close(self):
        if self._cursor:
            self._cursor.close()
        return self._conn.close()
    
    def cursor(self):
        return self._conn.cursor()

def get_db_connection():
    """Obtener conexión a MySQL con optimizaciones - RETORNA WRAPPER"""
    try:
        # Conectar a MySQL con pool settings para evitar rate limit
        config = DATABASE_CONFIG.copy()
        config['cursorclass'] = pymysql.cursors.DictCursor
        config['connect_timeout'] = 60  # Aumentado para Railway
        config['read_timeout'] = 60     # Aumentado para Railway
        config['write_timeout'] = 60    # Aumentado para Railway
        # Reutilizar conexiones
        config['autocommit'] = True  # Evita mantener transacciones abiertas
        
        conn = pymysql.connect(**config)
        
        # ✅ RETORNAR WRAPPER que permite conn.execute()
        return MySQLConnectionWrapper(conn)
    except Exception as e:
        print(f"❌ Error al conectar a MySQL: {e}")
        print(f"   Host: {DATABASE_CONFIG.get('host', 'N/A')}")
        print(f"   Database: {DATABASE_CONFIG.get('database', 'N/A')}")
        
        # Si falla por rate limit, esperar y reintentar
        if "rate limit" in str(e).lower() or "too many connections" in str(e).lower():
            print("⚠️ Rate limit detectado - esperando 2 segundos...")
            import time
            time.sleep(2)
            # Reintentar UNA vez
            try:
                conn = pymysql.connect(**config)
                return MySQLConnectionWrapper(conn)
            except:
                pass
        
        raise e

def increment_visit_counter():
    """Incrementar el contador de visitas del sitio"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE site_visits 
            SET total_visits = total_visits + 1, 
                last_updated = CURRENT_TIMESTAMP 
            WHERE id = 1
        ''')
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Error al incrementar contador de visitas: {e}")

def obtener_configuracion(clave, default='original'):
    """Obtener valor de configuración del sitio"""
    try:
        conn = get_db_connection()
        config = conn.execute(
            'SELECT valor FROM configuracion_sitio WHERE clave = %s',
            (clave,)
        ).fetchone()
        conn.close()
        
        if config and config['valor']:
            return config['valor']
        return default
    except Exception as e:
        print(f"Error al obtener configuración {clave}: {e}")
        return default

def actualizar_configuracion(clave, valor):
    """Actualizar o crear configuración del sitio"""
    try:
        conn = get_db_connection()
        
        # Verificar si existe
        existe = conn.execute(
            'SELECT id FROM configuracion_sitio WHERE clave = %s',
            (clave,)
        ).fetchone()
        
        if existe:
            # Actualizar
            conn.execute(
                'UPDATE configuracion_sitio SET valor = %s, updated_at = CURRENT_TIMESTAMP WHERE clave = %s',
                (valor, clave)
            )
        else:
            # Insertar
            conn.execute(
                'INSERT INTO configuracion_sitio (clave, valor) VALUES (%s, %s)',
                (clave, valor)
            )
        
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Error al actualizar configuración {clave}: {e}")
        return False

def create_database_indexes():
    """Verificar y crear índices críticos si no existen - MySQL optimized"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Obtener lista de índices existentes en la base de datos
        cursor.execute("""
            SELECT DISTINCT INDEX_NAME 
            FROM INFORMATION_SCHEMA.STATISTICS 
            WHERE TABLE_SCHEMA = DATABASE()
        """)
        existing_indexes = {row['INDEX_NAME'] for row in cursor.fetchall()}
        
        print(f"🔧 Verificando índices de base de datos... ({len(existing_indexes)} existentes)")
        
        # Índices críticos para mejorar velocidad
        indexes = {
            # Facturas Detalle (tabla más consultada)
            "idx_facturas_detalle_estado_activo": "CREATE INDEX idx_facturas_detalle_estado_activo ON facturas_detalle(estado, activo)",
            "idx_facturas_detalle_ars_estado": "CREATE INDEX idx_facturas_detalle_ars_estado ON facturas_detalle(ars_id, estado, activo)",
            "idx_facturas_detalle_medico_estado": "CREATE INDEX idx_facturas_detalle_medico_estado ON facturas_detalle(medico_id, estado, activo)",
            "idx_facturas_detalle_fecha_servicio": "CREATE INDEX idx_facturas_detalle_fecha_servicio ON facturas_detalle(fecha_servicio)",
            "idx_facturas_detalle_nss": "CREATE INDEX idx_facturas_detalle_nss ON facturas_detalle(nss)",
            "idx_facturas_detalle_factura_id": "CREATE INDEX idx_facturas_detalle_factura_id ON facturas_detalle(factura_id, activo)",
            
            # Pacientes
            "idx_pacientes_nss_ars_activo": "CREATE INDEX idx_pacientes_nss_ars_activo ON pacientes(nss, ars_id, activo)",
            "idx_pacientes_nombre_activo": "CREATE INDEX idx_pacientes_nombre_activo ON pacientes(nombre, activo)",
            "idx_pacientes_activo": "CREATE INDEX idx_pacientes_activo ON pacientes(activo)",
            
            # Facturas
            "idx_facturas_fecha_factura": "CREATE INDEX idx_facturas_fecha_factura ON facturas(fecha_factura)",
            "idx_facturas_medico_activo": "CREATE INDEX idx_facturas_medico_activo ON facturas(medico_id, activo)",
            "idx_facturas_ars_activo": "CREATE INDEX idx_facturas_ars_activo ON facturas(ars_id, activo)",
            "idx_facturas_activo": "CREATE INDEX idx_facturas_activo ON facturas(activo)",
            
            # Appointments
            "idx_appointments_status_created": "CREATE INDEX idx_appointments_status_created ON appointments(status, created_at)",
            "idx_appointments_date": "CREATE INDEX idx_appointments_date ON appointments(appointment_date)",
            "idx_appointments_email": "CREATE INDEX idx_appointments_email ON appointments(email)",
            
            # Contact Messages
            "idx_messages_read_created": "CREATE INDEX idx_messages_read_created ON contact_messages(`read`, created_at)",
            "idx_messages_email": "CREATE INDEX idx_messages_email ON contact_messages(email)",
            
            # Usuarios
            "idx_usuarios_email_activo": "CREATE INDEX idx_usuarios_email_activo ON usuarios(email, activo)",
            "idx_usuarios_activo": "CREATE INDEX idx_usuarios_activo ON usuarios(activo)",
            "idx_usuarios_perfil_activo": "CREATE INDEX idx_usuarios_perfil_activo ON usuarios(perfil, activo)",
            
            # Médicos
            "idx_medicos_activo_factura": "CREATE INDEX idx_medicos_activo_factura ON medicos(activo, factura)",
            "idx_medicos_email": "CREATE INDEX idx_medicos_email ON medicos(email)",
            "idx_medicos_nombre": "CREATE INDEX idx_medicos_nombre ON medicos(nombre)",
            
            # ARS
            "idx_ars_activo": "CREATE INDEX idx_ars_activo ON ars(activo)",
            "idx_ars_nombre": "CREATE INDEX idx_ars_nombre ON ars(nombre_ars)",
            
            # NCF
            "idx_ncf_activo": "CREATE INDEX idx_ncf_activo ON ncf(activo)",
            "idx_ncf_tipo_activo": "CREATE INDEX idx_ncf_tipo_activo ON ncf(tipo, activo)",
            
            # Código ARS
            "idx_codigo_ars_medico_ars": "CREATE INDEX idx_codigo_ars_medico_ars ON codigo_ars(medico_id, ars_id, activo)",
            "idx_codigo_ars_codigo": "CREATE INDEX idx_codigo_ars_codigo ON codigo_ars(codigo_ars)"
        }
        
        created_count = 0
        
        # Solo crear índices que NO existen
        for index_name, index_sql in indexes.items():
            if index_name not in existing_indexes:
                try:
                    cursor.execute(index_sql)
                    created_count += 1
                except Exception as e:
                    print(f"⚠️ Error creando {index_name}: {e}")
        
        conn.commit()
        conn.close()
        
        if created_count > 0:
            print(f"✅ Nuevos índices creados: {created_count}")
        print("✅ Índices de base de datos verificados")
        
    except Exception as e:
        print(f"❌ Error verificando índices: {e}")

def get_visit_count():
    """Obtener el número total de visitas del sitio"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT total_visits FROM site_visits WHERE id = 1')
        result = cursor.fetchone()
        conn.close()
        if result:
            return result['total_visits'] if isinstance(result, dict) else result[0]
        return 0
    except Exception as e:
        print(f"Error al obtener contador de visitas: {e}")
        return 0

def create_sample_data():
    """Crear datos de ejemplo"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Verificar si ya existen datos
    cursor.execute('SELECT COUNT(*) as count FROM services')
    result = cursor.fetchone()
    count = result['count'] if isinstance(result, dict) else result[0]
    
    if count > 0:
        conn.close()
        return
    
    # Servicios de ejemplo
    services = [
        ('Consulta Ginecológica', 'Consulta ginecológica\nPapanicolau\nColposcopia/biopsias\nCirugías ginecológicas\nIrregularidad menstrual\nGenotipificación y manejo del virus de papiloma humano\nManejo del síndrome de ovario poliquístico\nPlanificación familiar\nMenopausia y climaterio', 'fas fa-female', 'Consultar', '45 min'),
        ('Consulta Obstétrica', 'Consulta de obstetricia\nConsulta preconcepcional\nControl prenatal\nSeguimiento de embarazo de alto riesgo\nPartos y cesáreas\nAsesoría en lactancia materna', 'fas fa-person-pregnant', 'Consultar', '60 min'),
        ('Ecografías', 'Estudios de imagen para diagnóstico y seguimiento del embarazo y condiciones ginecológicas.', 'fas fa-heartbeat', 'Consultar', '30 min'),
        ('Cirugía Ginecológica', 'Procedimientos quirúrgicos especializados en ginecología con técnicas avanzadas.', 'fas fa-cut', 'Consultar', 'Variable'),
        ('Planificación Familiar', 'Asesoría sobre métodos anticonceptivos y planificación reproductiva personalizada.', 'fas fa-calendar-check', 'Consultar', '30 min'),
        ('Tratamientos Estéticos Ginecológicos', 'Tecnología láser de última generación para rejuvenecimiento vaginal, blanqueamiento genital, corrección de cicatrices y más. Click para ver todos los tratamientos disponibles.', 'fas fa-wand-magic-sparkles', 'Ver Tratamientos', 'Variable')
    ]
    
    cursor.executemany('INSERT INTO services (name, description, icon, price_range, duration) VALUES (%s, %s, %s, %s, %s)', services)
    
    # Tratamientos Estéticos Ginecológicos
    aesthetic_treatments = [
        ('Rejuvenecimiento vaginal', 'Mejora la elasticidad y tonicidad del canal vaginal.\nReduce la sequedad y mejora la lubricación natural.\nFavorece la producción de colágeno y elastina.', 'fas fa-magic'),
        ('Blanqueamiento genital', 'Aclara la pigmentación de la vulva, periné o zona anal.\nUnifica el tono de la piel íntima.', 'fas fa-sun'),
        ('Tensado o lifting vulvar', 'Mejora la apariencia externa de los labios mayores y menores.\nCorrige leve flacidez o laxitud de la vulva.', 'fas fa-compress-arrows-alt'),
        ('Corrección de cicatrices postparto o episiotomía', 'Suaviza la textura, color y relieve de cicatrices.\nDisminuye molestias o retracciones cicatriciales.', 'fas fa-hand-holding-medical'),
        ('Atrofia vulvovaginal (Síndrome Genitourinario de la Menopausia)', 'Alivia sequedad, ardor, picazón y dispareunia.\nEstimula la regeneración del epitelio vaginal.', 'fas fa-leaf'),
        ('Incontinencia urinaria leve a moderada', 'Fortalece las paredes vaginales y la uretra.\nMejora el soporte del suelo pélvico.\nDisminuye las pérdidas de orina al toser o hacer esfuerzo.', 'fas fa-shield-alt'),
        ('Laxitud vaginal postparto o por envejecimiento', 'Reafirma los tejidos del canal vaginal.\nMejora la sensibilidad y la satisfacción sexual.', 'fas fa-heart'),
        ('Vulvodinia y vestibulodinia leves', 'Reduce el dolor crónico vulvar mediante la bioestimulación tisular.', 'fas fa-spa'),
        ('Condilomas vulvares o vaginales (verrugas por VPH)', 'Eliminación precisa de lesiones con mínima lesión tisular.', 'fas fa-exclamation-triangle'),
        ('Lesiones cervicales leves (ej. NIC I)', 'Vaporización controlada de lesiones benignas o displásicas superficiales.', 'fas fa-stethoscope'),
        ('Quistes o pólipos pequeños del cuello uterino o vulvares', 'Resección con mínimo sangrado y rápida cicatrización.', 'fas fa-cut'),
        ('Liquen escleroso vulvar (casos seleccionados)', 'Mejora la textura y síntomas, reduciendo picazón y ardor.\nEstimula la regeneración epitelial.', 'fas fa-microscope'),
        ('Bartolinitis crónica o recidivante (marsupialización asistida con láser)', 'Facilita drenaje y recuperación más rápida.', 'fas fa-bolt'),
    ]
    
    cursor.executemany('INSERT INTO aesthetic_treatments (name, description, icon) VALUES (%s, %s, %s)', aesthetic_treatments)
    
    # Pool de 50 testimonios variados con nombres diferentes
    testimonials = [
        ('María González', 'M.G.', 'La Dra. Shirley es una excelente profesional. Su atención es muy cuidadosa y me hizo sentir cómoda durante toda la consulta. Recomiendo sus servicios sin dudarlo.', 5),
        ('Ana Rodríguez', 'A.R.', 'Durante mi embarazo recibí un excelente cuidado prenatal. La doctora siempre estuvo disponible para resolver mis dudas y me acompañó en todo el proceso.', 5),
        ('Carmen López', 'C.L.', 'Profesional, empática y muy conocedora de su especialidad. Me ayudó mucho durante mi tratamiento y siempre me explicó todo claramente.', 5),
        ('Patricia Martínez', 'P.M.', 'Quedé muy satisfecha con la atención recibida. La Dra. Shirley es muy profesional y dedicada. Mi experiencia fue excelente.', 5),
        ('Laura Fernández', 'L.F.', 'Excelente doctora, muy humana y profesional. Me explicó todo con paciencia y me sentí muy bien atendida.', 5),
        ('Isabel Sánchez', 'I.S.', 'La mejor ginecóloga que he visitado. Su trato es excepcional y siempre está al tanto de los últimos avances médicos.', 5),
        ('Rosa Ramírez', 'R.R.', 'Muy agradecida por el cuidado durante mi embarazo. La Dra. Shirley hizo que todo fuera más fácil y tranquilo.', 5),
        ('Sofía Torres', 'S.T.', 'Profesional y cariñosa. Me sentí en muy buenas manos desde el primer momento. Totalmente recomendada.', 5),
        ('Gabriela Díaz', 'G.D.', 'Su conocimiento y experiencia son evidentes. Me ayudó muchísimo con mi tratamiento. Muy agradecida.', 5),
        ('Daniela Ruiz', 'D.R.', 'La Dra. Shirley es increíble. Siempre tan atenta y profesional. Mi familia y yo estamos muy contentas.', 5),
        ('Valeria Morales', 'V.M.', 'Excelente atención, instalaciones muy limpias y personal amable. La doctora es muy profesional.', 5),
        ('Natalia Jiménez', 'N.J.', 'Me encantó su forma de explicar todo. Es muy clara y paciente. Definitivamente volveré.', 5),
        ('Andrea Castro', 'A.C.', 'Gran profesional con mucha experiencia. Me sentí muy segura durante todo mi tratamiento.', 5),
        ('Carolina Vargas', 'C.V.', 'La recomiendo al 100%. Es una doctora excepcional que realmente se preocupa por sus pacientes.', 5),
        ('Lucía Herrera', 'L.H.', 'Trato excelente y muy profesional. Respondió todas mis preguntas con mucha paciencia.', 5),
        ('Elena Mendoza', 'E.M.', 'Muy satisfecha con la consulta. La Dra. Shirley es muy detallista y cuidadosa.', 5),
        ('Paula Ortiz', 'P.O.', 'Excelente doctora, muy preparada y con un trato humano excepcional. La mejor decisión.', 5),
        ('Mónica Silva', 'M.S.', 'Me acompañó durante todo mi embarazo con mucho profesionalismo. Muy agradecida.', 5),
        ('Victoria Reyes', 'V.R.', 'Súper recomendada. Es una doctora muy capacitada y con excelente trato humano.', 5),
        ('Diana Flores', 'D.F.', 'Muy profesional y atenta. Me explicó todo detalladamente y me sentí muy cómoda.', 5),
        ('Alejandra Cruz', 'A.C.', 'La mejor ginecóloga que he conocido. Su atención es personalizada y muy profesional.', 5),
        ('Mariana Guzmán', 'M.G.', 'Excelente experiencia. La doctora es muy profesional y el personal muy amable.', 5),
        ('Beatriz Rojas', 'B.R.', 'Muy contenta con mi consulta. La Dra. Shirley es excepcional en todos los aspectos.', 5),
        ('Cristina Medina', 'C.M.', 'Gran profesional con mucho conocimiento. Me sentí muy bien atendida y segura.', 5),
        ('Adriana Vega', 'A.V.', 'La recomiendo completamente. Es una doctora muy preparada y con excelente trato.', 5),
        ('Rocío Campos', 'R.C.', 'Muy satisfecha con el servicio. La doctora es muy profesional y dedicada.', 5),
        ('Teresa Santos', 'T.S.', 'Excelente atención durante todo mi embarazo. La Dra. Shirley es maravillosa.', 5),
        ('Silvia Peña', 'S.P.', 'Muy profesional y empática. Me ayudó muchísimo con mi tratamiento. Gracias.', 5),
        ('Julia Cortés', 'J.C.', 'La mejor experiencia que he tenido con una ginecóloga. Súper recomendada.', 5),
        ('Claudia Aguilar', 'C.A.', 'Excelente doctora, muy humana y profesional. Me sentí muy bien cuidada.', 5),
        ('Mercedes Luna', 'M.L.', 'Muy agradecida por su atención. Es una doctora excepcional con gran vocación.', 5),
        ('Lorena Chávez', 'L.C.', 'Profesional, cariñosa y muy preparada. La recomiendo sin dudarlo.', 5),
        ('Raquel Domínguez', 'R.D.', 'Excelente servicio y atención. La Dra. Shirley es muy profesional y dedicada.', 5),
        ('Susana Ramos', 'S.R.', 'Me encantó su trato y profesionalismo. Definitivamente la mejor ginecóloga.', 5),
        ('Verónica Gil', 'V.G.', 'Muy satisfecha con todo. La doctora es excelente y el personal muy amable.', 5),
        ('Pilar Núñez', 'P.N.', 'Gran experiencia. La Dra. Shirley es muy profesional y atenta con sus pacientes.', 5),
        ('Gloria Molina', 'G.M.', 'La mejor doctora. Su trato es excepcional y siempre te hace sentir cómoda.', 5),
        ('Cecilia Paredes', 'C.P.', 'Muy recomendada. Es una profesional de primera con excelente trato humano.', 5),
        ('Irene Fuentes', 'I.F.', 'Excelente atención y seguimiento. La Dra. Shirley es muy dedicada y profesional.', 5),
        ('Alicia Márquez', 'A.M.', 'Muy contenta con mi consulta. La doctora es muy profesional y empática.', 5),
        ('Rebeca Ibáñez', 'R.I.', 'La recomiendo al 100%. Es una doctora excepcional en todos los sentidos.', 5),
        ('Norma Guerrero', 'N.G.', 'Excelente profesional. Me ayudó muchísimo durante mi embarazo. Muy agradecida.', 5),
        ('Leticia Soto', 'L.S.', 'Muy profesional y atenta. Su trato es excepcional y su conocimiento impecable.', 5),
        ('Yolanda Ríos', 'Y.R.', 'Gran doctora con mucha experiencia. Me sentí muy segura bajo su cuidado.', 5),
        ('Amparo Navarro', 'A.N.', 'Excelente en todos los aspectos. La Dra. Shirley es una gran profesional.', 5),
        ('Dolores Romero', 'D.R.', 'Muy satisfecha con la atención recibida. La doctora es muy dedicada y profesional.', 5),
        ('Francisca León', 'F.L.', 'La mejor ginecóloga. Su trato es excepcional y siempre está disponible para ayudar.', 5),
        ('Antonia Blanco', 'A.B.', 'Excelente doctora y mejor persona. Me acompañó en todo mi proceso con mucho cariño.', 5),
        ('Josefina Pascual', 'J.P.', 'Muy profesional y humana. La Dra. Shirley es simplemente excepcional.', 5),
        ('Encarnación Cano', 'E.C.', 'Gran experiencia desde el primer día. La recomiendo completamente. Gracias doctora.', 5)
    ]
    
    cursor.executemany('INSERT INTO testimonials (patient_name, patient_initials, testimonial_text, rating, approved) VALUES (%s, %s, %s, %s, 1)', testimonials)
    
    conn.commit()
    conn.close()

# Rutas principales
# ============ CACHE SIMPLE PARA HOMEPAGE ============
homepage_cache = {
    'services': None,
    'testimonials': None,
    'timestamp': None
}

CACHE_DURATION = 300  # 5 minutos

def get_cached_homepage_data():
    """Obtener datos de homepage con cache de 5 minutos"""
    from datetime import datetime
    
    now = datetime.now()
    
    # Si no hay cache o expiró, recargar
    if (homepage_cache['timestamp'] is None or 
        (now - homepage_cache['timestamp']).total_seconds() > CACHE_DURATION):
        
        conn = get_db_connection()
        homepage_cache['services'] = conn.execute('SELECT * FROM services WHERE active = 1 LIMIT 6').fetchall()
        homepage_cache['testimonials'] = conn.execute('SELECT * FROM testimonials WHERE approved = 1').fetchall()
        homepage_cache['timestamp'] = now
        conn.close()
        
        print(f"🔄 Cache actualizado a las {now.strftime('%H:%M:%S')}")
    
    return homepage_cache['services'], homepage_cache['testimonials']

@app.route('/')
def index():
    """Página principal con testimonios rotativos"""
    from datetime import datetime, timedelta, timezone
    import random
    
    # Limpiar rate limits antiguos periódicamente (cada visita a inicio)
    try:
        cleanup_old_rate_limits()
    except:
        pass  # No detener la página si falla la limpieza
    
    # Incrementar contador de visitas (solo UNA VEZ por sesión, NO usuarios logueados)
    if not current_user.is_authenticated and not session.get('visit_counted'):
        increment_visit_counter()
        session['visit_counted'] = True  # Marcar que ya se contó esta sesión
    
    # Usar cache para datos (reduce consultas DB)
    services, all_testimonials = get_cached_homepage_data()
    
    # Rotación diaria de testimonios (zona horaria RD: UTC-4)
    tz_rd = timezone(timedelta(hours=-4))
    today = datetime.now(tz_rd)
    seed = today.timetuple().tm_yday
    random.seed(seed)
    
    testimonials_list = list(all_testimonials)
    random.shuffle(testimonials_list)
    selected_testimonials = testimonials_list[:3]  # Mostrar 3 en inicio
    
    # Asignar fechas dinámicas
    testimonials_with_dates = []
    for i, testimonial in enumerate(selected_testimonials):
        random.seed(seed + i + 1000)  # Semilla diferente al de página testimonios
        days_ago = random.randint(1, 30)
        display_date = today - timedelta(days=days_ago)
        
        if days_ago == 1:
            time_ago = "Hace 1 día"
        elif days_ago < 7:
            time_ago = f"Hace {days_ago} días"
        elif days_ago < 14:
            time_ago = "Hace 1 semana"
        elif days_ago < 30:
            weeks = days_ago // 7
            time_ago = f"Hace {weeks} semanas"
        else:
            time_ago = "Hace 1 mes"
        
        testimonial_dict = dict(testimonial)
        testimonial_dict['time_ago'] = time_ago
        testimonials_with_dates.append(testimonial_dict)
    
    # Obtener tema configurado con manejo de errores
    try:
        tema = obtener_configuracion('tema_principal', 'original')
    except Exception as e:
        print(f"Error al obtener tema: {e}")
        tema = 'original'
    
    # Lógica inteligente de temas según el mes y día (con manejo de errores)
    try:
        mes_actual = today.month
        dia_actual = today.day
        
        # Auto-activación especial por fecha
        if mes_actual == 2 and dia_actual == 14:
            # 14 de Febrero: San Valentín (reemplaza Mes de la Patria ese día)
            if tema != 'san_valentin':
                if actualizar_configuracion('tema_principal', 'san_valentin'):
                    tema = 'san_valentin'
        elif mes_actual == 2 and tema == 'original':
            # Resto de Febrero: activar Mes de la Patria
            if actualizar_configuracion('tema_principal', 'mes_patria'):
                tema = 'mes_patria'
        elif mes_actual == 3 and tema == 'original':
            # Marzo: activar Mes de la Mujer
            if actualizar_configuracion('tema_principal', 'mes_mujer'):
                tema = 'mes_mujer'
        elif mes_actual == 10 and tema == 'original':
            # Octubre: activar Cáncer de Mama
            if actualizar_configuracion('tema_principal', 'cancer_mama'):
                tema = 'cancer_mama'
        
        # Auto-desactivación al terminar el mes o día especial
        elif mes_actual == 2 and dia_actual == 15 and tema == 'san_valentin':
            # 15 de Febrero: volver a Mes de la Patria
            if actualizar_configuracion('tema_principal', 'mes_patria'):
                tema = 'mes_patria'
        elif mes_actual == 3 and tema == 'mes_patria':
            # Marzo: desactivar Mes de la Patria y activar Mes de la Mujer
            if actualizar_configuracion('tema_principal', 'mes_mujer'):
                tema = 'mes_mujer'
        elif mes_actual == 4 and tema == 'mes_mujer':
            # Abril: desactivar Mes de la Mujer
            if actualizar_configuracion('tema_principal', 'original'):
                tema = 'original'
        elif mes_actual == 11 and tema == 'cancer_mama':
            # Noviembre: desactivar Cáncer de Mama
            if actualizar_configuracion('tema_principal', 'original'):
                tema = 'original'
    except Exception as e:
        print(f"Error en lógica de temas: {e}")
        # Si hay error, usar tema por defecto
        tema = 'original'
    
    return render_template('index.html', 
                         services=services, 
                         testimonials=testimonials_with_dates,
                         tema=tema,
                         ano_actual=today.year)

@app.route('/index-v2')
def index_v2():
    """Página principal - Versión refrescante"""
    from datetime import datetime, timedelta
    import random
    
    conn = get_db_connection()
    services = conn.execute('SELECT * FROM services WHERE active = 1 LIMIT 6').fetchall()
    all_testimonials = conn.execute('SELECT * FROM testimonials WHERE approved = 1').fetchall()
    conn.close()
    
    # Rotación diaria de testimonios
    today = datetime.now()
    seed = today.timetuple().tm_yday
    random.seed(seed)
    
    testimonials_list = list(all_testimonials)
    random.shuffle(testimonials_list)
    selected_testimonials = testimonials_list[:3]  # Mostrar 3 en inicio
    
    # Asignar fechas dinámicas
    testimonials_with_dates = []
    for i, testimonial in enumerate(selected_testimonials):
        random.seed(seed + i + 1000)  # Semilla diferente
        days_ago = random.randint(1, 30)
        display_date = today - timedelta(days=days_ago)
        
        if days_ago == 1:
            time_ago = "Hace 1 día"
        elif days_ago < 7:
            time_ago = f"Hace {days_ago} días"
        elif days_ago < 14:
            time_ago = "Hace 1 semana"
        elif days_ago < 30:
            weeks = days_ago // 7
            time_ago = f"Hace {weeks} semanas"
        else:
            time_ago = "Hace 1 mes"
        
        testimonial_dict = dict(testimonial)
        testimonial_dict['time_ago'] = time_ago
        testimonials_with_dates.append(testimonial_dict)
    
    return render_template('index_v2.html', 
                         services=services, 
                         testimonials=testimonials_with_dates)

@app.route('/old-index')
def old_index():
    """Página principal original - Backup"""
    from datetime import datetime, timedelta
    import random
    
    conn = get_db_connection()
    services = conn.execute('SELECT * FROM services WHERE active = 1 LIMIT 6').fetchall()
    all_testimonials = conn.execute('SELECT * FROM testimonials WHERE approved = 1').fetchall()
    conn.close()
    
    # Rotación diaria de testimonios
    today = datetime.now()
    seed = today.timetuple().tm_yday
    random.seed(seed)
    
    testimonials_list = list(all_testimonials)
    random.shuffle(testimonials_list)
    selected_testimonials = testimonials_list[:3]  # Mostrar 3 en inicio
    
    # Asignar fechas dinámicas
    testimonials_with_dates = []
    for i, testimonial in enumerate(selected_testimonials):
        random.seed(seed + i + 1000)  # Semilla diferente al de página testimonios
        days_ago = random.randint(1, 30)
        display_date = today - timedelta(days=days_ago)
        
        if days_ago == 1:
            time_ago = "Hace 1 día"
        elif days_ago < 7:
            time_ago = f"Hace {days_ago} días"
        elif days_ago < 14:
            time_ago = "Hace 1 semana"
        elif days_ago < 30:
            weeks = days_ago // 7
            time_ago = f"Hace {weeks} semanas"
        else:
            time_ago = "Hace 1 mes"
        
        testimonial_dict = dict(testimonial)
        testimonial_dict['time_ago'] = time_ago
        testimonial_dict['formatted_date'] = display_date.strftime('%d/%m/%Y')
        testimonials_with_dates.append(testimonial_dict)
    
    return render_template('index.html', services=services, testimonials=testimonials_with_dates)

@app.route('/servicios')
def services():
    """Página de servicios"""
    conn = get_db_connection()
    services = conn.execute('SELECT * FROM services WHERE active = 1').fetchall()
    conn.close()
    
    return render_template('services.html', services=services)

@app.route('/test-icons')
def test_icons():
    """Test de iconos Font Awesome"""
    return render_template('test_icons.html')

@app.route('/test-simple')
def test_simple():
    """Test sin base de datos"""
    return "<h1>✅ SERVIDOR FUNCIONANDO</h1><p>Si ves esto, el servidor Flask está OK</p>"

@app.route('/tratamientos-esteticos')
def aesthetic_treatments():
    """Página de Tratamientos Estéticos Ginecológicos (no en menú)"""
    conn = get_db_connection()
    treatments = conn.execute('SELECT * FROM aesthetic_treatments ORDER BY id').fetchall()
    conn.close()
    
    return render_template('tratamientos-esteticos.html', treatments=treatments)

@app.route('/sobre-mi')
def about():
    """Página sobre la doctora"""
    return render_template('about.html')

@app.route('/ginecologa-santo-domingo')
def seo_ginecologa_santo_domingo():
    """Landing SEO para búsqueda amplia en Santo Domingo"""
    return render_template('seo_ginecologa_santo_domingo.html')

@app.route('/testimonios')
def testimonials():
    """Página de testimonios con rotación diaria y fechas dinámicas"""
    from datetime import datetime, timedelta
    import random
    
    conn = get_db_connection()
    
    # Obtener TODOS los testimonios
    all_testimonials = conn.execute('SELECT * FROM testimonials WHERE approved = 1').fetchall()
    conn.close()
    
    # Usar el día del año como semilla para rotación consistente por día
    today = datetime.now()
    seed = today.timetuple().tm_yday  # Día del año (1-365)
    random.seed(seed)
    
    # Mezclar los testimonios de forma consistente para el día
    testimonials_list = list(all_testimonials)
    random.shuffle(testimonials_list)
    
    # Tomar 9 testimonios para mostrar (cambian cada día)
    selected_testimonials = testimonials_list[:9]
    
    # Asignar fechas aleatorias de los últimos 30 días a cada testimonio
    testimonials_with_dates = []
    for i, testimonial in enumerate(selected_testimonials):
        # Generar fecha aleatoria entre 1 y 30 días atrás (consistente para el día)
        random.seed(seed + i)  # Semilla única para cada testimonio
        days_ago = random.randint(1, 30)
        display_date = today - timedelta(days=days_ago)
        
        # Calcular texto "hace X días/semanas"
        if days_ago == 1:
            time_ago = "Hace 1 día"
        elif days_ago < 7:
            time_ago = f"Hace {days_ago} días"
        elif days_ago < 14:
            time_ago = "Hace 1 semana"
        elif days_ago < 30:
            weeks = days_ago // 7
            time_ago = f"Hace {weeks} semanas"
        else:
            time_ago = "Hace 1 mes"
        
        # Convertir Row a dict y agregar información de fecha
        testimonial_dict = dict(testimonial)
        testimonial_dict['display_date'] = display_date.strftime('%Y-%m-%d')
        testimonial_dict['time_ago'] = time_ago
        testimonial_dict['formatted_date'] = display_date.strftime('%d/%m/%Y')
        
        testimonials_with_dates.append(testimonial_dict)
    
    # Ordenar por fecha más reciente primero
    testimonials_with_dates.sort(key=lambda x: x['display_date'], reverse=True)
    
    return render_template('testimonials.html', testimonials=testimonials_with_dates)

# ============================================================================
# FUNCIONES DE ENVÍO DE EMAIL CON SENDGRID API
# ============================================================================

def send_email_sendgrid(to_email, subject, html_content, attachment_data=None, attachment_filename=None):
    """
    Enviar email usando SendGrid API (funciona en Railway donde SMTP está bloqueado)
    
    Args:
        to_email: Destinatario
        subject: Asunto del email
        html_content: Contenido HTML del email
        attachment_data: Datos del adjunto (BytesIO o bytes) - opcional
        attachment_filename: Nombre del archivo adjunto - opcional
    
    Returns:
        bool: True si se envió exitosamente, False si hubo error
    """
    try:
        # Verificaciones previas ANTES de intentar enviar
        if not SENDGRID_AVAILABLE:
            print("\n⚠️ SendGrid no está instalado. Emails deshabilitados.")
            print("   Instala con: pip install sendgrid")
            return False
            
        if not SENDGRID_API_KEY or SENDGRID_API_KEY == '':
            print("\n⚠️ SENDGRID_API_KEY no está configurada. Emails deshabilitados.")
            print("   Configura SENDGRID_API_KEY en Railway → Variables")
            print("   O establece EMAIL_CONFIGURED=false para deshabilitar emails")
            return False
        
        if len(SENDGRID_API_KEY) < 30 or not SENDGRID_API_KEY.startswith('SG.'):
            print("\n⚠️ SENDGRID_API_KEY parece inválida (debe empezar con 'SG.')")
            print("   Verifica la API Key en SendGrid: https://app.sendgrid.com/settings/api_keys")
            return False
        
        # Crear mensaje
        message = Mail(
            from_email=Email(EMAIL_FROM, "Dra. Shirley Ramírez - Ginecóloga"),
            to_emails=To(to_email),
            subject=subject,
            html_content=Content("text/html", html_content)
        )
        
        # Configurar Reply-To para mejorar deliverability
        message.reply_to = Email(EMAIL_FROM, "Dra. Shirley Ramírez")
        
        # Desactivar click tracking de SendGrid (evita URLs reescritas)
        message.tracking_settings = TrackingSettings()
        message.tracking_settings.click_tracking = ClickTracking(False, False)
        
        # Agregar adjunto si existe
        if attachment_data and attachment_filename:
            # Convertir BytesIO a bytes si es necesario
            if isinstance(attachment_data, BytesIO):
                attachment_data.seek(0)
                file_data = attachment_data.read()
            else:
                file_data = attachment_data
            
            # Codificar en base64
            encoded_file = base64.b64encode(file_data).decode()
            
            # Agregar adjunto
            attached_file = Attachment(
                FileContent(encoded_file),
                FileName(attachment_filename),
                FileType('application/pdf'),
                Disposition('attachment')
            )
            message.attachment = attached_file
        
        # Enviar
        sg = SendGridAPIClient(SENDGRID_API_KEY)
        response = sg.send(message)
        
        print(f"✅ Email enviado exitosamente (Status: {response.status_code})")
        print(f"   📧 To: {to_email}")
        print(f"   📝 Subject: {subject}")
        
        return True
        
    except Exception as e:
        error_str = str(e).lower()
        
        # Manejar errores específicos con mensajes claros
        if '401' in error_str or 'unauthorized' in error_str:
            print("\n" + "="*60)
            print("❌ ERROR: SENDGRID_API_KEY INVÁLIDA O EXPIRADA")
            print("="*60)
            print("La API Key de SendGrid no es válida.")
            print("\nSOLUCIÓN:")
            print("1. Ir a: https://app.sendgrid.com/settings/api_keys")
            print("2. Crear nueva API Key con Full Access")
            print("3. Copiar la key (empieza con SG.)")
            print("4. En Railway → Variables → SENDGRID_API_KEY → Pegar")
            print("\nO DESHABILITAR EMAILS:")
            print("Railway → Variables → Agregar: EMAIL_CONFIGURED = false")
            print("="*60 + "\n")
        elif '403' in error_str or 'forbidden' in error_str:
            print("\n⚠️ Error 403: API Key sin permisos suficientes")
            print("   Crea una nueva con 'Full Access'")
        else:
            print(f"\n❌ Error al enviar email: {e}")
            import traceback
            traceback.print_exc()
        
        return False

def enviar_email_pdf_pacientes(medico_email, medico_nombre, pdf_buffer, num_pacientes, total):
    """Enviar email con PDF adjunto de pacientes agregados usando SendGrid API
    También envía copia a todos los médicos facturadores
    """
    try:
        if not EMAIL_CONFIGURED:
            print("\n⚠️  Email no configurado. El PDF no se envió.")
            return False
        
        # Usar template estandarizado
        html = template_constancia_pdf(medico_nombre, num_pacientes, total)
        
        # Nombre del archivo
        filename = f'constancia_pacientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        
        # Enviar al médico principal
        success = send_email_sendgrid(
            to_email=medico_email,
            subject=f'📋 Constancia - {num_pacientes} Paciente(s) Pendiente(s) de Facturación',
            html_content=html,
            attachment_data=pdf_buffer,
            attachment_filename=filename
        )
        
        if success:
            print("\n" + "=" * 60)
            print("✅ EMAIL CON PDF ENVIADO EXITOSAMENTE")
            print("=" * 60)
            print(f"📧 Destinatario: {medico_email}")
            print(f"👨‍⚕️ Médico: {medico_nombre}")
            print(f"📋 Pacientes: {num_pacientes}")
            print(f"💰 Total: {total:,.2f}")
            print("=" * 60 + "\n")
            
            # NUEVO: Enviar copia a médicos facturadores
            try:
                conn = get_db_connection()
                medicos_facturadores = conn.execute('''
                    SELECT DISTINCT email, nombre 
                    FROM medicos 
                    WHERE activo = 1 AND factura = 1 AND email IS NOT NULL AND email != %s
                    ORDER BY nombre
                ''', (medico_email,)).fetchall()
                conn.close()
                
                # Enviar copia a cada médico facturador
                for med_fac in medicos_facturadores:
                    if med_fac['email']:
                        # Crear nueva copia del buffer para cada email
                        pdf_copia = BytesIO(pdf_buffer.getvalue())
                        
                        # Enviar copia (sin esperar resultado)
                        threading.Thread(target=send_email_sendgrid, args=(
                            med_fac['email'],
                            f'📋 [Copia] Constancia - {num_pacientes} Paciente(s) - {medico_nombre}',
                            html,
                            pdf_copia,
                            filename
                        )).start()
                        
                        print(f"📤 Copia enviada a: {med_fac['nombre']} ({med_fac['email']})")
            except Exception as e:
                print(f"⚠️ Error al enviar copias a médicos: {e}")
                # No fallar si las copias no se envían
        
        return success
        
    except Exception as e:
        print("\n" + "=" * 60)
        print("❌ ERROR AL ENVIAR EMAIL CON PDF")
        print("=" * 60)
        print(f"Error: {e}")
        print("\nEl PDF se descargó correctamente, pero no se envió el email.")
        print("=" * 60 + "\n")
        return False

def enviar_email_notificacion(name, email, phone, subject, message):
    """Enviar email de notificación a la doctora usando SendGrid API"""
    try:
        # Verificar configuración de email
        if not EMAIL_CONFIGURED:
            print("\n⚠️  CONFIGURACIÓN DE EMAIL NECESARIA")
            print("=" * 60)
            print("Para recibir emails, configura SENDGRID_API_KEY")
            print("=" * 60)
            print("\nPor ahora, el mensaje se guardó en la base de datos.")
            print("=" * 60 + "\n")
            return False
        
        # Usar template estandarizado
        html = template_contacto(name, email, phone, subject, message)
        
        # Enviar usando SendGrid API
        success = send_email_sendgrid(
            to_email=EMAIL_DESTINATARIO,
            subject=f'🔔 Nuevo mensaje: {subject}',
            html_content=html
        )
        
        if success:
            print(f"✅ Notificación enviada a {EMAIL_DESTINATARIO}")
        
        return success
        
    except Exception as e:
        print(f"❌ ERROR AL ENVIAR EMAIL: {e}")
        print("\nEl mensaje se guardó en la base de datos.")
        import traceback
        traceback.print_exc()
        return False

def enviar_email_recuperacion(email, nombre, link_recuperacion):
    """Enviar email de recuperación de contraseña usando SendGrid API"""
    try:
        if not EMAIL_CONFIGURED:
            print("⚠️  No se puede enviar email de recuperación: Email no configurado")
            return False
        
        # Usar template estandarizado
        html = template_recuperacion(nombre, link_recuperacion)
        
        # Enviar con SendGrid API
        success = send_email_sendgrid(
            to_email=email,
            subject='🔐 Recuperación de Contraseña - Panel Administrativo',
            html_content=html
        )
        
        if success:
            print(f"✅ Email de recuperación enviado a {email}")
        
        return success
        
    except Exception as e:
        print(f"❌ Error al enviar email de recuperación: {e}")
        return False

def enviar_email_cita(first_name, last_name, email, phone, appointment_date, appointment_time, appointment_type, medical_insurance, emergency_datetime, reason):
    """Enviar email de notificación de cita a la doctora usando SendGrid API"""
    try:
        # Verificar configuración
        if not EMAIL_CONFIGURED:
            print("\n⚠️  Email no configurado. La cita se guardó en la base de datos.")
            return False
        
        # Preparar datos para el template
        fecha = emergency_datetime if appointment_type == "emergencia" else appointment_date
        hora = appointment_time if appointment_type != "emergencia" else None
        tipo_cita = f"{appointment_type} {'(EMERGENCIA)' if appointment_type == 'emergencia' else ''}".strip()
        seguro = medical_insurance if medical_insurance else "No especificado"
        motivo = reason if reason else "No especificado"
        
        # Usar template estandarizado
        html = template_cita(
            first_name, last_name, email if email else "No proporcionado", 
            phone, fecha, hora if hora else "URGENTE", tipo_cita, 
            seguro, emergency_datetime if appointment_type == "emergencia" else None, motivo
        )
        
        # Enviar usando SendGrid API
        success = send_email_sendgrid(
            to_email=EMAIL_DESTINATARIO,
            subject=f'📅 Nueva Solicitud de Cita - {first_name} {last_name}',
            html_content=html
        )
        
        if success:
            print(f"✅ Notificación de cita enviada a {EMAIL_DESTINATARIO}")
        
        return success
        
    except Exception as e:
        print(f"❌ ERROR AL ENVIAR EMAIL DE CITA: {e}")
        print("\nLa cita se guardó en la base de datos.")
        import traceback
        traceback.print_exc()
        return False

def enviar_email_confirmacion_cita(paciente_email, nombre, apellido, fecha, hora, tipo, estatus, motivo=None):
    """Enviar email de confirmación de cambio de estatus de cita al paciente"""
    try:
        # Verificar si hay contraseña configurada
        if not EMAIL_PASSWORD or EMAIL_PASSWORD == "tu_password_aqui":
            print("\n⚠️ CONFIGURACIÓN DE EMAIL PENDIENTE")
            print("Por favor, configura EMAIL_PASSWORD en el archivo .env")
            return False
        
        # Verificar que el paciente tenga email
        if not paciente_email:
            print("\n⚠️ El paciente no tiene email registrado")
            return False
        
        print("\n" + "=" * 60)
        print("📧 ENVIANDO EMAIL DE CONFIRMACIÓN DE CITA AL PACIENTE")
        print("=" * 60)
        
        # Crear mensaje
        msg = MIMEMultipart('alternative')
        
        # Asunto según el estatus
        asuntos = {
            'pending': '⏳ Tu Cita está Pendiente de Confirmación',
            'confirmed': '✅ ¡Tu Cita ha sido Confirmada!',
            'cancelled': '❌ Tu Cita ha sido Cancelada',
            'completed': '✔️ Tu Cita ha sido Completada'
        }
        
        msg['Subject'] = asuntos.get(estatus, '📅 Actualización de tu Cita')
        msg['From'] = EMAIL_USERNAME
        msg['To'] = paciente_email
        
        # Generar HTML usando el template
        html = template_confirmacion_cita(nombre, apellido, fecha, hora, tipo, estatus, motivo)
        
        # Adjuntar HTML
        part = MIMEText(html, 'html')
        msg.attach(part)
        
        # Conectar y enviar con configuración desde variables de entorno
        server = smtplib.SMTP(EMAIL_HOST, EMAIL_PORT, timeout=30)
        if EMAIL_USE_TLS:
            server.starttls()
        server.login(EMAIL_USERNAME, EMAIL_PASSWORD)
        server.send_message(msg)
        server.quit()
        
        print("\n✅ EMAIL DE CONFIRMACIÓN ENVIADO EXITOSAMENTE")
        print("=" * 60)
        print(f"📧 Destinatario: {paciente_email}")
        print(f"👤 Paciente: {nombre} {apellido}")
        print(f"🏥 Estatus: {estatus}")
        print(f"📅 Fecha: {fecha if fecha else 'Por confirmar'}")
        print(f"🕐 Hora: {hora if hora else 'Por confirmar'}")
        print("=" * 60 + "\n")
        
        return True
        
    except Exception as e:
        print("\n" + "=" * 60)
        print("❌ ERROR AL ENVIAR EMAIL DE CONFIRMACIÓN")
        print("=" * 60)
        print(f"Error: {e}")
        print("\nEl cambio de estatus se guardó en la base de datos.")
        print("Pero el email no pudo ser enviado al paciente.")
        print("=" * 60 + "\n")
        return False

def send_email(destinatario, asunto, cuerpo):
    """Función genérica para enviar emails HTML usando SendGrid API"""
    try:
        if not EMAIL_CONFIGURED:
            print("\n⚠️ CONFIGURACIÓN DE EMAIL PENDIENTE")
            print("Por favor, configura SENDGRID_API_KEY")
            return False
        
        # Verificar que el destinatario tenga email
        if not destinatario:
            print("\n⚠️ No se proporcionó un destinatario")
            return False
        
        print("\n" + "=" * 60)
        print("📧 ENVIANDO EMAIL")
        print("=" * 60)
        print(f"📧 Destinatario: {destinatario}")
        print(f"📝 Asunto: {asunto}")
        
        # Enviar con SendGrid API
        success = send_email_sendgrid(
            to_email=destinatario,
            subject=asunto,
            html_content=cuerpo
        )
        
        if success:
            print("\n✅ EMAIL ENVIADO EXITOSAMENTE")
            print("=" * 60 + "\n")
        
        return success
        
    except Exception as e:
        print("\n" + "=" * 60)
        print("❌ ERROR AL ENVIAR EMAIL")
        print("=" * 60)
        print(f"Error: {e}")
        print("=" * 60 + "\n")
        return False

@app.route('/contacto', methods=['GET', 'POST'])
def contact():
    """Página de contacto"""
    if request.method == 'POST':
        try:
            # Rate limit solo para POST (envío de formularios)
            client_ip = request.remote_addr
            current_time = time.time()
            
            with rate_limit_lock:
                # Limpiar requests antiguos (últimos 5 minutos)
                request_counts[f'{client_ip}_contact'] = [
                    req_time for req_time in request_counts.get(f'{client_ip}_contact', [])
                    if current_time - req_time < 300  # 5 minutos
                ]
                
                # Verificar límite (5 envíos por 5 minutos)
                if len(request_counts.get(f'{client_ip}_contact', [])) >= 5:
                    flash('⚠️ Has enviado demasiados mensajes. Por favor espera 5 minutos.', 'warning')
                    return redirect(url_for('contact'))
                
                # Agregar request actual
                if f'{client_ip}_contact' not in request_counts:
                    request_counts[f'{client_ip}_contact'] = []
                request_counts[f'{client_ip}_contact'].append(current_time)
            
            name = request.form['name']
            email = request.form['email'].strip().lower()  # Normalizar email
            phone = request.form.get('phone', '')
            subject = request.form['subject']
            message = request.form['message']
            
            # Validar Google reCAPTCHA
            recaptcha_response = request.form.get('g-recaptcha-response')
            print(f"🔐 reCAPTCHA Token recibido: {recaptcha_response[:50] if recaptcha_response else 'NINGUNO'}")
            print(f"🔑 RECAPTCHA_SITE_KEY configurada: {'Sí' if RECAPTCHA_SITE_KEY else 'No'}")
            print(f"🔑 RECAPTCHA_SECRET_KEY configurada: {'Sí' if RECAPTCHA_SECRET_KEY else 'No'}")
            
            # Solo validar reCAPTCHA si está configurada
            if RECAPTCHA_SITE_KEY and RECAPTCHA_SECRET_KEY:
                if not verificar_recaptcha(recaptcha_response):
                    flash('⚠️ Por favor, completa la verificación de seguridad (reCAPTCHA).', 'warning')
                    return redirect(url_for('contact'))
            else:
                print("⚠️ reCAPTCHA no configurada, saltando validación")
            
            # Validar que todos los campos estén completos
            if not all([name, email, phone, subject, message]):
                flash('Por favor, completa todos los campos obligatorios.', 'danger')
                return redirect(url_for('contact'))
            
            # Verificar si el email está en la lista negra (SILENCIOSAMENTE)
            conn = get_db_connection()
            blacklist_check = conn.execute(
                'SELECT id FROM email_blacklist WHERE email = %s', 
                (email,)
            ).fetchone()
            
            if blacklist_check:
                # Email bloqueado - mostrar mensaje de éxito pero NO guardar ni enviar
                conn.close()
                print(f"🚫 Email bloqueado rechazado silenciosamente: {email}")
                flash('¡Mensaje enviado correctamente! Te contactaremos pronto.', 'success')
                return redirect(url_for('contact'))
            
            # Email válido - procesar normalmente
            # Guardar en base de datos
            conn.execute('''
                INSERT INTO contact_messages (name, email, phone, subject, message)
                VALUES (%s, %s, %s, %s, %s)
            ''', (name, email, phone, subject, message))
            conn.commit()
            conn.close()
            
            # Enviar email de notificación de manera asíncrona (no bloquea la respuesta)
            try:
                threading.Thread(target=enviar_email_notificacion, args=(name, email, phone, subject, message)).start()
            except Exception as email_error:
                # Si falla el email, solo lo registramos pero no detenemos el proceso
                print(f"⚠️ Error al enviar email (no crítico): {email_error}")
            
            flash('¡Mensaje enviado correctamente! Te contactaremos pronto.', 'success')
            return redirect(url_for('contact'))
            
        except Exception as e:
            # Capturar cualquier error y mostrar un mensaje amigable al usuario
            print(f"❌ ERROR EN CONTACTO: {e}")
            import traceback
            traceback.print_exc()
            flash('⚠️ Ocurrió un error al enviar tu mensaje. Por favor intenta nuevamente.', 'danger')
            return redirect(url_for('contact'))
    
    return render_template('contact.html')

@app.route('/api/horarios-disponibles', methods=['GET'])
def horarios_disponibles():
    """API para obtener horarios disponibles según la fecha"""
    fecha = request.args.get('fecha')
    
    if not fecha:
        return jsonify({'error': 'Fecha requerida'}), 400
    
    try:
        # Parsear la fecha para obtener el día de la semana
        from datetime import datetime
        fecha_obj = datetime.strptime(fecha, '%Y-%m-%d')
        dia_semana = fecha_obj.weekday()  # 0=Lunes, 1=Martes, 3=Jueves
        
        # Definir horarios según el día
        if dia_semana == 1:  # Martes
            horarios_base = [
                '13:00', '13:30', '14:00', '14:30', '15:00', 
                '15:30', '16:00', '16:30', '17:00', '17:30', '18:00'
            ]
        elif dia_semana == 3:  # Jueves
            horarios_base = [
                '08:00', '08:30', '09:00', '09:30', '10:00',
                '10:30', '11:00', '11:30', '12:00', '12:30', '13:00'
            ]
        else:
            return jsonify({'horarios': []})
        
        # Obtener citas ya agendadas para esa fecha (que no estén canceladas)
        conn = get_db_connection()
        citas_ocupadas = conn.execute(
            'SELECT appointment_time FROM appointments WHERE appointment_date = %s AND status != "cancelled"',
            (fecha,)
        ).fetchall()
        conn.close()
        
        # Crear lista de horarios ocupados
        horarios_ocupados = [cita[0] for cita in citas_ocupadas if cita[0]]
        
        # Filtrar horarios disponibles
        horarios_disponibles = [h for h in horarios_base if h not in horarios_ocupados]
        
        return jsonify({'horarios': horarios_disponibles})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/solicitar-cita', methods=['GET', 'POST'])
def request_appointment():
    """Solicitar cita"""
    if request.method == 'POST':
        try:
            # Rate limit solo para POST (envío de formularios)
            client_ip = request.remote_addr
            current_time = time.time()
            
            with rate_limit_lock:
                # Limpiar requests antiguos (últimos 5 minutos)
                request_counts[f'{client_ip}_appointment'] = [
                    req_time for req_time in request_counts.get(f'{client_ip}_appointment', [])
                    if current_time - req_time < 300  # 5 minutos
                ]
                
                # Verificar límite (3 envíos por 5 minutos)
                if len(request_counts.get(f'{client_ip}_appointment', [])) >= 3:
                    flash('⚠️ Has enviado demasiadas solicitudes. Por favor espera 5 minutos.', 'warning')
                    return redirect(url_for('request_appointment'))
                
                # Agregar request actual
                if f'{client_ip}_appointment' not in request_counts:
                    request_counts[f'{client_ip}_appointment'] = []
                request_counts[f'{client_ip}_appointment'].append(current_time)
            
            # Validar y sanitizar entrada
            first_name = sanitize_input(request.form.get('first_name', ''))
            last_name = sanitize_input(request.form.get('last_name', ''))
            email = sanitize_input(request.form.get('email', ''))
            phone = sanitize_input(request.form.get('phone', ''))
            appointment_date = sanitize_input(request.form.get('appointment_date', ''))
            appointment_time = sanitize_input(request.form.get('appointment_time', ''))
            appointment_type = sanitize_input(request.form.get('appointment_type', ''))
            medical_insurance = sanitize_input(request.form.get('medical_insurance', ''))
            
            # Validar Google reCAPTCHA
            recaptcha_response = request.form.get('g-recaptcha-response')
            print(f"🔐 reCAPTCHA Token recibido (citas): {recaptcha_response[:50] if recaptcha_response else 'NINGUNO'}")
            print(f"🔑 RECAPTCHA_SITE_KEY configurada: {'Sí' if RECAPTCHA_SITE_KEY else 'No'}")
            print(f"🔑 RECAPTCHA_SECRET_KEY configurada: {'Sí' if RECAPTCHA_SECRET_KEY else 'No'}")
            
            # Solo validar reCAPTCHA si está configurada
            if RECAPTCHA_SITE_KEY and RECAPTCHA_SECRET_KEY:
                if not verificar_recaptcha(recaptcha_response):
                    flash('⚠️ Por favor, completa la verificación de seguridad (reCAPTCHA).', 'warning')
                    return redirect(url_for('request_appointment'))
            else:
                print("⚠️ reCAPTCHA no configurada, saltando validación")
            
            # Validaciones críticas
            if not all([first_name, last_name, phone, appointment_type, medical_insurance]):
                flash('Por favor, completa todos los campos obligatorios.', 'danger')
                return redirect(url_for('request_appointment'))
            
            if email and not validate_email(email):
                flash('Por favor, ingresa un email válido.', 'danger')
                return redirect(url_for('request_appointment'))
            
            if len(first_name) > 50 or len(last_name) > 50:
                flash('Los nombres no pueden exceder 50 caracteres.', 'danger')
                return redirect(url_for('request_appointment'))
            
            emergency_datetime = sanitize_input(request.form.get('emergency_datetime', ''))
            reason = sanitize_input(request.form.get('reason', ''))
            
            # Verificar disponibilidad de horario (solo para citas normales con fecha y hora)
            if appointment_type in ["consulta", "estetico"] and appointment_date and appointment_time:
                conn = get_db_connection()
                cita_existente = conn.execute(
                    'SELECT id FROM appointments WHERE appointment_date = %s AND appointment_time = %s AND status != "cancelled"',
                    (appointment_date, appointment_time)
                ).fetchone()
                
                if cita_existente:
                    conn.close()
                    flash('⚠️ Lo sentimos, ese horario ya está ocupado. Por favor selecciona otro horario.', 'warning')
                    return redirect(url_for('request_appointment'))
                
                conn.close()
            
            # Verificar disponibilidad de horario de emergencia
            if appointment_type == "emergencia" and emergency_datetime:
                # Extraer fecha y hora del datetime
                try:
                    from datetime import datetime
                    emergency_dt = datetime.strptime(emergency_datetime, '%Y-%m-%dT%H:%M')
                    emergency_date = emergency_dt.strftime('%Y-%m-%d')  # Mantener para SQL
                    emergency_time = emergency_dt.strftime('%H:%M')
                    
                    conn = get_db_connection()
                    cita_existente = conn.execute(
                        'SELECT id FROM appointments WHERE emergency_datetime LIKE %s AND status != "cancelled"',
                        (f'{emergency_date}%',)
                    ).fetchall()
                    
                    # Verificar si hay conflicto de horario (dentro de 30 minutos)
                    if cita_existente and len(cita_existente) > 0:
                        conn.close()
                        flash('⚠️ Lo sentimos, ya hay una cita de emergencia cerca de ese horario. Por favor selecciona otro horario.', 'warning')
                        return redirect(url_for('request_appointment'))
                    
                    conn.close()
                except:
                    pass
            
            conn = get_db_connection()
            
            # Convertir campos vacíos a None (NULL en SQL)
            appointment_date_val = appointment_date if appointment_date else None
            appointment_time_val = appointment_time if appointment_time else None
            emergency_datetime_val = emergency_datetime if emergency_datetime else None
            reason_val = reason if reason else None
            email_val = email if email else None
            
            conn.execute('''
                INSERT INTO appointments (first_name, last_name, email, phone, appointment_date, appointment_time, appointment_type, medical_insurance, emergency_datetime, reason)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ''', (first_name, last_name, email_val, phone, appointment_date_val, appointment_time_val, appointment_type, medical_insurance, emergency_datetime_val, reason_val))
            conn.commit()
            conn.close()
            
            # Enviar email de notificación a la doctora de manera asíncrona (no bloquea la respuesta)
            try:
                threading.Thread(target=enviar_email_cita, args=(first_name, last_name, email, phone, appointment_date, appointment_time, appointment_type, medical_insurance, emergency_datetime, reason)).start()
            except Exception as email_error:
                # Si falla el email, solo lo registramos pero no detenemos el proceso
                print(f"⚠️ Error al enviar email (no crítico): {email_error}")
            
            flash('¡Cita solicitada correctamente! Te contactaremos para confirmar.', 'success')
            return redirect(url_for('request_appointment'))
            
        except Exception as e:
            # Capturar cualquier error y mostrar un mensaje amigable al usuario
            print(f"❌ ERROR EN SOLICITAR CITA: {e}")
            import traceback
            traceback.print_exc()
            flash('⚠️ Ocurrió un error al procesar tu solicitud. Por favor intenta nuevamente o contáctanos directamente.', 'danger')
            return redirect(url_for('request_appointment'))
    
    return render_template('request_appointment.html')

# ============================================================================
# AUTENTICACIÓN Y GESTIÓN DE USUARIOS
# ============================================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Inicio de sesión para el panel de administración"""
    if current_user.is_authenticated:
        return redirect(url_for('admin'))
    
    if request.method == 'POST':
        # Rate limit manual para POST (5 intentos por 5 minutos)
        client_ip = request.remote_addr
        current_time = time.time()
        
        with rate_limit_lock:
            # Limpiar requests antiguos
            request_counts[f'{client_ip}_login'] = [
                req_time for req_time in request_counts.get(f'{client_ip}_login', [])
                if current_time - req_time < 300  # 5 minutos
            ]
            
            # Verificar límite
            if len(request_counts.get(f'{client_ip}_login', [])) >= 5:
                flash('⚠️ Demasiados intentos de inicio de sesión. Por favor espera 5 minutos.', 'error')
                return redirect(url_for('login'))
            
            # Agregar request actual
            if f'{client_ip}_login' not in request_counts:
                request_counts[f'{client_ip}_login'] = []
            request_counts[f'{client_ip}_login'].append(current_time)
        
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        remember = request.form.get('remember', False)
        
        if not email or not password:
            flash('Por favor ingresa email y contraseña', 'error')
            return redirect(url_for('login'))
        
        # Buscar usuario en la base de datos
        conn = get_db_connection()
        user_data = conn.execute(
            'SELECT id, nombre, email, password_hash, perfil, activo, password_temporal FROM usuarios WHERE email = %s',
            (email,)
        ).fetchone()
        conn.close()
        
        # Verificar usuario y contraseña
        if user_data and user_data['activo']:
            if check_password_hash(user_data['password_hash'], password):
                # Verificar si tiene contraseña temporal
                if user_data['password_temporal']:
                    # Guardar datos en sesión y redirigir a cambio de contraseña
                    session['cambio_password_usuario_id'] = user_data['id']
                    session['cambio_password_email'] = user_data['email']
                    flash('Debes cambiar tu contraseña temporal antes de continuar', 'warning')
                    return redirect(url_for('cambiar_password_obligatorio'))
                
                # Crear objeto usuario y hacer login
                user = User(
                    id=user_data['id'],
                    nombre=user_data['nombre'],
                    email=user_data['email'],
                    perfil=user_data['perfil']
                )
                
                # Marcar sesión como permanente para que persista
                session.permanent = True
                login_user(user, remember=True)
                
                # Actualizar last_login
                conn = get_db_connection()
                conn.execute('UPDATE usuarios SET last_login = %s WHERE id = %s',
                           (datetime.now(), user_data['id']))
                conn.commit()
                conn.close()
                
                # Redirigir a la página solicitada o al admin
                next_page = request.args.get('next')
                if next_page and next_page.startswith('/'):
                    return redirect(next_page)
                return redirect(url_for('admin'))
            else:
                flash('Contraseña incorrecta', 'error')
        else:
            flash('Usuario no encontrado o inactivo', 'error')
        
        return redirect(url_for('login'))
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    """Cerrar sesión"""
    logout_user()
    flash('Sesión cerrada correctamente', 'success')
    return redirect(url_for('index'))

@app.route('/cambiar-password-obligatorio', methods=['GET', 'POST'])
def cambiar_password_obligatorio():
    """Cambiar contraseña temporal obligatoriamente"""
    # Verificar que hay datos en sesión
    usuario_id = session.get('cambio_password_usuario_id')
    email = session.get('cambio_password_email')
    
    if not usuario_id or not email:
        flash('Sesión inválida. Por favor inicia sesión nuevamente.', 'error')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        password_temporal = request.form.get('password_temporal', '')
        nueva_password = request.form.get('nueva_password', '')
        confirmar_password = request.form.get('confirmar_password', '')
        
        if not password_temporal or not nueva_password or not confirmar_password:
            flash('Todos los campos son obligatorios', 'error')
            return render_template('cambiar_password_obligatorio.html', email=email)
        
        if nueva_password != confirmar_password:
            flash('Las contraseñas nuevas no coinciden', 'error')
            return render_template('cambiar_password_obligatorio.html', email=email)
        
        # Validar fortaleza de la nueva contraseña
        errores = validar_password_segura(nueva_password)
        if errores:
            flash('La contraseña no cumple con los requisitos: ' + ', '.join(errores), 'error')
            return render_template('cambiar_password_obligatorio.html', email=email)
        
        # Verificar contraseña temporal
        conn = get_db_connection()
        user_data = conn.execute(
            'SELECT password_hash FROM usuarios WHERE id = %s',
            (usuario_id,)
        ).fetchone()
        
        if not user_data or not check_password_hash(user_data['password_hash'], password_temporal):
            conn.close()
            flash('La contraseña temporal es incorrecta', 'error')
            return render_template('cambiar_password_obligatorio.html', email=email)
        
        # Actualizar contraseña
        nueva_password_hash = generate_password_hash(nueva_password)
        conn.execute(
            'UPDATE usuarios SET password_hash = %s, password_temporal = 0 WHERE id = %s',
            (nueva_password_hash, usuario_id)
        )
        conn.commit()
        conn.close()
        
        # Limpiar sesión
        session.pop('cambio_password_usuario_id', None)
        session.pop('cambio_password_email', None)
        
        flash('Contraseña cambiada exitosamente. Por favor inicia sesión con tu nueva contraseña.', 'success')
        return redirect(url_for('login'))
    
    return render_template('cambiar_password_obligatorio.html', email=email)

@app.route('/solicitar-recuperacion', methods=['GET', 'POST'])
def solicitar_recuperacion():
    """Solicitar recuperación de contraseña"""
    if current_user.is_authenticated:
        return redirect(url_for('admin'))
    
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        
        if not email:
            flash('Por favor ingresa tu email', 'error')
            return redirect(url_for('solicitar_recuperacion'))
        
        # Buscar usuario
        conn = get_db_connection()
        user = conn.execute('SELECT id, nombre FROM usuarios WHERE email = %s AND activo = 1', (email,)).fetchone()
        
        if user:
            # Generar token de recuperación
            token = secrets.token_urlsafe(32)
            expiracion = datetime.now() + timedelta(hours=1)
            
            conn.execute(
                'UPDATE usuarios SET token_recuperacion = %s, token_expiracion = %s WHERE id = %s',
                (token, expiracion, user['id'])
            )
            conn.commit()
            
            # Enviar email con el link de recuperación
            link_recuperacion = url_for('recuperar_contrasena', token=token, _external=True)
            enviar_email_recuperacion(email, user['nombre'], link_recuperacion)
            
        conn.close()
        
        # Siempre mostrar el mismo mensaje (seguridad)
        flash('Si el email existe, recibirás un enlace de recuperación', 'info')
        return redirect(url_for('login'))
    
    return render_template('solicitar_recuperacion.html')

@app.route('/recuperar-contrasena/<token>', methods=['GET', 'POST'])
def recuperar_contrasena(token):
    """Cambiar contraseña con token de recuperación"""
    if current_user.is_authenticated:
        return redirect(url_for('admin'))
    
    # Verificar token
    conn = get_db_connection()
    user = conn.execute(
        'SELECT id, nombre, email FROM usuarios WHERE token_recuperacion = %s AND token_expiracion > %s AND activo = 1',
        (token, datetime.now())
    ).fetchone()
    
    if not user:
        conn.close()
        flash('Token inválido o expirado', 'error')
        return redirect(url_for('solicitar_recuperacion'))
    
    if request.method == 'POST':
        password = request.form.get('password', '')
        password_confirm = request.form.get('password_confirm', '')
        
        if not password or len(password) < 8:
            flash('La contraseña debe tener al menos 8 caracteres', 'error')
            return redirect(url_for('recuperar_contrasena', token=token))
        
        if password != password_confirm:
            flash('Las contraseñas no coinciden', 'error')
            return redirect(url_for('recuperar_contrasena', token=token))
        
        # Actualizar contraseña
        password_hash = generate_password_hash(password)
        conn.execute(
            'UPDATE usuarios SET password_hash = %s, token_recuperacion = NULL, token_expiracion = NULL WHERE id = %s',
            (password_hash, user['id'])
        )
        conn.commit()
        conn.close()
        
        flash('Contraseña actualizada correctamente. Ya puedes iniciar sesión', 'success')
        return redirect(url_for('login'))
    
    conn.close()
    return render_template('recuperar_contrasena.html', token=token, user=user)

@app.route('/admin')
@login_required
@cache_result(expiration_seconds=60)  # Cachear estadísticas por 1 minuto
def admin():
    """Panel de administración simple con caché"""
    conn = get_db_connection()
    
    # Estadísticas (con caché)
    def get_count(query):
        result = conn.execute(query).fetchone()
        if isinstance(result, dict):
            return list(result.values())[0]
        return result[0]
    
    stats = {
        'total_appointments': get_count('SELECT COUNT(*) FROM appointments WHERE status = "confirmed"'),
        'pending_appointments': get_count('SELECT COUNT(*) FROM appointments WHERE status = "pending"'),
        'unread_messages': get_count('SELECT COUNT(*) FROM contact_messages WHERE `read` = 0'),
        'total_testimonials': get_count('SELECT COUNT(*) FROM testimonials'),
        'total_visits': get_visit_count()
    }
    
    # Datos recientes (citas pendientes, citas confirmadas, y mensajes sin leer)
    recent_appointments = conn.execute('SELECT * FROM appointments WHERE status = "pending" ORDER BY created_at DESC LIMIT 5').fetchall()
    confirmed_appointments = conn.execute('SELECT * FROM appointments WHERE status = "confirmed" ORDER BY appointment_date DESC, appointment_time DESC LIMIT 5').fetchall()
    recent_messages = conn.execute('SELECT * FROM contact_messages WHERE `read` = 0 ORDER BY created_at DESC LIMIT 5').fetchall()
    
    conn.close()
    
    return render_template('admin.html', stats=stats, recent_appointments=recent_appointments, confirmed_appointments=confirmed_appointments, recent_messages=recent_messages)

@app.route('/admin/appointments')
@login_required
def admin_appointments():
    """Gestión de citas"""
    conn = get_db_connection()
    appointments = conn.execute('SELECT * FROM appointments ORDER BY created_at DESC').fetchall()
    conn.close()
    
    return render_template('admin_appointments.html', appointments=appointments)

@app.route('/admin/messages')
@login_required
def admin_messages():
    """Gestión de mensajes"""
    conn = get_db_connection()
    messages = conn.execute('SELECT * FROM contact_messages ORDER BY created_at DESC').fetchall()
    conn.close()
    
    return render_template('admin_messages.html', messages=messages)

@app.route('/admin/visor-pagina')
@login_required
def admin_visor_pagina():
    """Página para seleccionar el tema de la página principal"""
    tema_actual = obtener_configuracion('tema_principal', 'original')
    return render_template('admin_visor_pagina.html', tema_actual=tema_actual)

@app.route('/admin/visor-pagina/guardar', methods=['POST'])
@login_required
def guardar_tema_pagina():
    """Guardar el tema seleccionado"""
    nuevo_tema = request.form.get('theme', 'original')
    
    # Validar que sea un tema válido
    if nuevo_tema not in ['original', 'mes_patria', 'san_valentin', 'mes_mujer', 'cancer_mama']:
        flash('Tema inválido', 'error')
        return redirect(url_for('admin_visor_pagina'))
    
    # Guardar en base de datos
    if actualizar_configuracion('tema_principal', nuevo_tema):
        # Nombres de temas
        nombres_temas = {
            'original': 'Diseño Original',
            'mes_patria': '🇩🇴 Mes de la Patria',
            'san_valentin': '💕 San Valentín',
            'mes_mujer': '♀ Mes de la Mujer',
            'cancer_mama': '🎗️ Cáncer de Mama'
        }
        nombre_tema = nombres_temas.get(nuevo_tema, 'Diseño Original')
        flash(f'✅ Tema actualizado a: {nombre_tema}', 'success')
    else:
        flash('Error al guardar la configuración', 'error')
    
    return redirect(url_for('admin_visor_pagina'))

@app.route('/admin/appointments/<int:appointment_id>/update', methods=['POST'])
@login_required
def update_appointment_status(appointment_id):
    """Actualizar estado de cita y enviar notificación al paciente"""
    new_status = request.form['status']
    
    conn = get_db_connection()
    
    # Obtener datos de la cita antes de actualizar
    appointment = conn.execute(
        'SELECT * FROM appointments WHERE id = %s',
        (appointment_id,)
    ).fetchone()
    
    if not appointment:
        flash('Cita no encontrada', 'error')
        conn.close()
        return redirect(url_for('admin_appointments'))
    
    # Actualizar estado
    conn.execute('UPDATE appointments SET status = %s WHERE id = %s', (new_status, appointment_id))
    conn.commit()
    conn.close()
    
    # Enviar email de confirmación al paciente si tiene email
    if appointment['email']:
        # Preparar los datos
        nombre = appointment['first_name']
        apellido = appointment['last_name']
        email_paciente = appointment['email']
        fecha = appointment['appointment_date'] if appointment['appointment_date'] else appointment['emergency_datetime']
        hora = appointment['appointment_time']
        tipo = appointment['appointment_type']
        motivo = appointment['reason'] if appointment['reason'] else None
        
        # Enviar email de manera asíncrona (no bloquea la respuesta)
        threading.Thread(target=enviar_email_confirmacion_cita, args=(
            email_paciente,
            nombre,
            apellido,
            fecha,
            hora,
            tipo,
            new_status,
            motivo
        )).start()
        
        flash(f'Estado de la cita actualizado y notificación enviada a {nombre} {apellido}', 'success')
    else:
        flash('Estado de la cita actualizado (paciente sin email registrado)', 'warning')
    
    return redirect(url_for('admin_appointments'))

@app.route('/admin/appointments/<int:appointment_id>/status', methods=['POST'])
@login_required
def update_appointment_status_ajax(appointment_id):
    """Actualizar estado de cita via AJAX (JSON)"""
    try:
        data = request.get_json()
        new_status = data.get('status')
        
        # Validar el estado
        valid_statuses = ['pending', 'confirmed', 'completed', 'cancelled']
        if new_status not in valid_statuses:
            return jsonify({'success': False, 'error': 'Estado inválido'}), 400
        
        conn = get_db_connection()
        
        # Obtener datos de la cita antes de actualizar
        appointment = conn.execute(
            'SELECT * FROM appointments WHERE id = %s',
            (appointment_id,)
        ).fetchone()
        
        if not appointment:
            conn.close()
            return jsonify({'success': False, 'error': 'Cita no encontrada'}), 404
        
        # Actualizar estado
        conn.execute('UPDATE appointments SET status = %s WHERE id = %s', (new_status, appointment_id))
        conn.commit()
        conn.close()
        
        # Enviar email de confirmación al paciente si tiene email (asíncrono)
        if appointment['email']:
            nombre = appointment['first_name']
            apellido = appointment['last_name']
            email_paciente = appointment['email']
            fecha = appointment['appointment_date'] if appointment['appointment_date'] else appointment['emergency_datetime']
            hora = appointment['appointment_time']
            tipo = appointment['appointment_type']
            motivo = appointment['reason'] if appointment['reason'] else None
            
            threading.Thread(target=enviar_email_confirmacion_cita, args=(
                email_paciente,
                nombre,
                apellido,
                fecha,
                hora,
                tipo,
                new_status,
                motivo
            )).start()
        
        return jsonify({
            'success': True,
            'message': f'Estado actualizado a: {new_status}',
            'new_status': new_status
        })
        
    except Exception as e:
        print(f"Error al actualizar estado de cita: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/appointments/<int:appointment_id>/delete', methods=['POST'])
@login_required
def delete_appointment(appointment_id):
    """Eliminar una cita permanentemente"""
    try:
        conn = get_db_connection()
        
        # Obtener datos de la cita antes de eliminar (para el log)
        appointment = conn.execute(
            'SELECT * FROM appointments WHERE id = %s',
            (appointment_id,)
        ).fetchone()
        
        if not appointment:
            conn.close()
            flash('Cita no encontrada', 'error')
            return redirect(url_for('admin_appointments'))
        
        # Eliminar la cita
        conn.execute('DELETE FROM appointments WHERE id = %s', (appointment_id,))
        conn.commit()
        conn.close()
        
        print(f"✅ Cita eliminada: ID={appointment_id}, Paciente: {appointment['first_name']} {appointment['last_name']}")
        flash(f'Cita de {appointment["first_name"]} {appointment["last_name"]} eliminada correctamente', 'success')
        
        return redirect(url_for('admin_appointments'))
        
    except Exception as e:
        print(f"❌ Error al eliminar cita: {e}")
        flash(f'Error al eliminar la cita: {str(e)}', 'error')
        return redirect(url_for('admin_appointments'))

@app.route('/admin/messages/<int:message_id>/mark-read', methods=['POST'])
@login_required
def mark_message_read(message_id):
    """Marcar mensaje como leído"""
    try:
        conn = get_db_connection()
        # Usar backticks para escapar la palabra reservada 'read' en MySQL
        conn.execute('UPDATE contact_messages SET `read` = 1 WHERE id = %s', (message_id,))
        conn.commit()
        conn.close()
        
        print(f"✅ Mensaje {message_id} marcado como leído")
        return jsonify({'success': True})
    except Exception as e:
        print(f"❌ Error al marcar mensaje como leído: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/messages/<int:message_id>/delete', methods=['POST'])
@login_required
def delete_message(message_id):
    """Eliminar un mensaje permanentemente"""
    try:
        conn = get_db_connection()
        
        # Obtener datos del mensaje antes de eliminar (para el log)
        message = conn.execute(
            'SELECT * FROM contact_messages WHERE id = %s',
            (message_id,)
        ).fetchone()
        
        if not message:
            conn.close()
            flash('Mensaje no encontrado', 'error')
            return redirect(url_for('admin_messages'))
        
        # Eliminar el mensaje
        conn.execute('DELETE FROM contact_messages WHERE id = %s', (message_id,))
        conn.commit()
        conn.close()
        
        print(f"✅ Mensaje eliminado: ID={message_id}, De: {message['name']} ({message['email']})")
        flash(f'Mensaje de {message["name"]} eliminado correctamente', 'success')
        
        return redirect(url_for('admin_messages'))
        
    except Exception as e:
        print(f"❌ Error al eliminar mensaje: {e}")
        flash(f'Error al eliminar el mensaje: {str(e)}', 'error')
        return redirect(url_for('admin_messages'))

@app.route('/admin/messages/<int:message_id>/block-sender', methods=['POST'])
@login_required
def block_sender(message_id):
    """Bloquear remitente de un mensaje (agregar a lista negra)"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Crear tabla si no existe (por si acaso)
        try:
            cursor.execute(adapt_sql_for_database('''
                CREATE TABLE IF NOT EXISTS email_blacklist (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    email VARCHAR(255) UNIQUE NOT NULL,
                    blocked_by VARCHAR(255) NOT NULL,
                    reason TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            '''))
            conn.commit()
        except Exception as table_error:
            print(f"⚠️ Advertencia al crear tabla: {table_error}")
        
        # Obtener el email del mensaje
        message = conn.execute(
            'SELECT email, name FROM contact_messages WHERE id = %s',
            (message_id,)
        ).fetchone()
        
        if not message:
            conn.close()
            return jsonify({'success': False, 'error': 'Mensaje no encontrado'}), 404
        
        email = message['email'].strip().lower()
        name = message['name']
        
        # Verificar si ya está bloqueado
        existing = conn.execute(
            'SELECT id FROM email_blacklist WHERE email = %s',
            (email,)
        ).fetchone()
        
        if existing:
            conn.close()
            return jsonify({'success': False, 'error': 'Este email ya está bloqueado'}), 400
        
        # Agregar a lista negra
        conn.execute('''
            INSERT INTO email_blacklist (email, blocked_by, reason)
            VALUES (%s, %s, %s)
        ''', (email, current_user.email, f'Bloqueado desde mensaje de {name}'))
        
        conn.commit()
        conn.close()
        
        print(f"🚫 Email bloqueado: {email} por {current_user.email}")
        return jsonify({'success': True, 'email': email})
        
    except Exception as e:
        print(f"❌ Error al bloquear remitente: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/blacklist')
@login_required
def admin_blacklist():
    """Ver lista negra de emails"""
    conn = get_db_connection()
    blacklist = conn.execute('''
        SELECT * FROM email_blacklist 
        ORDER BY created_at DESC
    ''').fetchall()
    conn.close()
    
    return render_template('admin_blacklist.html', blacklist=blacklist)

@app.route('/admin/blacklist/<int:blacklist_id>/unblock', methods=['POST'])
@login_required
def unblock_email(blacklist_id):
    """Desbloquear un email de la lista negra"""
    try:
        conn = get_db_connection()
        
        # Obtener el email antes de eliminarlo
        entry = conn.execute(
            'SELECT email FROM email_blacklist WHERE id = %s',
            (blacklist_id,)
        ).fetchone()
        
        if not entry:
            conn.close()
            return jsonify({'success': False, 'error': 'Entrada no encontrada'}), 404
        
        email = entry['email']
        
        # Eliminar de la lista negra
        conn.execute('DELETE FROM email_blacklist WHERE id = %s', (blacklist_id,))
        conn.commit()
        conn.close()
        
        print(f"✅ Email desbloqueado: {email}")
        return jsonify({'success': True, 'email': email})
        
    except Exception as e:
        print(f"❌ Error al desbloquear email: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/messages/<int:message_id>/reply', methods=['POST'])
@login_required
def reply_to_message(message_id):
    """Enviar respuesta a un mensaje desde el dashboard"""
    try:
        data = request.get_json()
        to_email = data.get('to_email')
        subject = data.get('subject')
        message_text = data.get('message')
        
        if not all([to_email, subject, message_text]):
            return jsonify({'success': False, 'error': 'Faltan campos requeridos'}), 400
        
        # Verificar configuración de email
        if not EMAIL_CONFIGURED:
            return jsonify({'success': False, 'error': 'Email no configurado en el servidor'}), 500
        
        # Crear HTML profesional para el email
        html_content = f"""
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
        </head>
        <body style="margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #F8F4F5;">
            <table width="100%" cellpadding="0" cellspacing="0" style="background-color: #F8F4F5; padding: 30px 15px;">
                <tr>
                    <td align="center">
                        <table width="600" cellpadding="0" cellspacing="0" style="max-width: 600px; width: 100%; background-color: white; border-radius: 15px; box-shadow: 0 4px 20px rgba(0,0,0,0.08);">
                            <tr>
                                <td>
                                    <!-- Header -->
                                    <div style="background: linear-gradient(135deg, #CEB0B7 0%, #B89CA3 100%); padding: 30px; text-align: center; border-radius: 15px 15px 0 0;">
                                        <h1 style="color: white; margin: 0; font-size: 28px; font-weight: 700;">
                                            Dra. Shirley Ramírez
                                        </h1>
                                        <p style="color: rgba(255,255,255,0.95); margin: 8px 0 0 0; font-size: 14px;">
                                            Ginecóloga • Obstetra • Salud Femenina
                                        </p>
                                    </div>
                                    
                                    <!-- Content -->
                                    <div style="padding: 35px 30px;">
                                        <div style="color: #282828; line-height: 1.8; font-size: 15px; white-space: pre-wrap;">
                                            {message_text}
                                        </div>
                                    </div>
                                    
                                    <!-- Footer -->
                                    <div style="background-color: #F2E2E6; padding: 25px; text-align: center; border-radius: 0 0 15px 15px; margin-top: 20px;">
                                        <div style="border-top: 2px solid #CEB0B7; padding-top: 20px; margin-bottom: 15px;">
                                            <p style="color: #ACACAD; font-size: 14px; margin: 8px 0; font-weight: 600;">
                                                📞 829-740-5073 | 📧 dra.ramirezr@gmail.com
                                            </p>
                                            <p style="color: #ACACAD; font-size: 13px; margin: 8px 0;">
                                                📍 Santo Domingo | República Dominicana
                                            </p>
                                        </div>
                                        <div style="margin-top: 15px;">
                                            <a href="https://www.linkedin.com/in/shirley-ramirez-montero-a10964168/" 
                                               style="display: inline-block; margin: 0 8px; color: #CEB0B7; text-decoration: none; font-size: 20px;">
                                                🔗 LinkedIn
                                            </a>
                                            <a href="https://www.instagram.com/dra.ramirezr/" 
                                               style="display: inline-block; margin: 0 8px; color: #CEB0B7; text-decoration: none; font-size: 20px;">
                                                📷 Instagram
                                            </a>
                                        </div>
                                        <p style="color: #999; font-size: 11px; margin: 15px 0 0 0;">
                                            &copy; 2025 Dra. Shirley Ramírez • Todos los derechos reservados
                                        </p>
                                    </div>
                                </td>
                            </tr>
                        </table>
                    </td>
                </tr>
            </table>
        </body>
        </html>
        """
        
        # Enviar email usando SendGrid
        success = send_email_sendgrid(
            to_email=to_email,
            subject=subject,
            html_content=html_content
        )
        
        if success:
            print(f"✅ Respuesta enviada a {to_email} desde dashboard por {current_user.email}")
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Error al enviar el email'}), 500
        
    except Exception as e:
        print(f"❌ Error al responder mensaje: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== FACTURACIÓN ROUTES ====================
@app.route('/facturacion')
@login_required
def facturacion_menu():
    """Menú principal de facturación"""
    return render_template('facturacion/menu.html')

# ========== MAESTRA DE ARS ==========
@app.route('/facturacion/ars')
@login_required
def facturacion_ars():
    """Lista de ARS"""
    search = request.args.get('search', '')
    conn = get_db_connection()
    
    if search:
        search_pattern = f"%{search}%"
        ars_list = conn.execute(
            "SELECT * FROM ars WHERE (nombre_ars LIKE %s OR rnc LIKE %s) AND activo = 1 ORDER BY nombre_ars",
            (search_pattern, search_pattern)
        ).fetchall()
    else:
        ars_list = conn.execute('SELECT * FROM ars WHERE activo = 1 ORDER BY nombre_ars').fetchall()
    
    conn.close()
    return render_template('facturacion/ars.html', ars_list=ars_list, search=search)

@app.route('/facturacion/ars/nuevo', methods=['GET', 'POST'])
@login_required
def facturacion_ars_nuevo():
    """Crear nuevo ARS"""
    if request.method == 'POST':
        nombre_ars = sanitize_input(request.form['nombre_ars'], 200)
        rnc = sanitize_input(request.form['rnc'], 50)

        # Validación de caracteres (integridad)
        try:
            nombre_ars = _validate_or_raise(nombre_ars, _RE_TEXTO_GENERAL, 'Nombre del ARS', max_len=200)
            rnc = _validate_or_raise(rnc, _RE_RNC, 'RNC', max_len=50)
        except ValueError as ve:
            flash(f'❌ {str(ve)}', 'error')
            return redirect(url_for('facturacion_ars_nuevo'))
        
        if not nombre_ars or not rnc:
            flash('Todos los campos son obligatorios', 'error')
            return redirect(url_for('facturacion_ars_nuevo'))
        
        conn = get_db_connection()
        conn.execute('INSERT INTO ars (nombre_ars, rnc) VALUES (%s, %s)', (nombre_ars, rnc))
        conn.commit()
        conn.close()
        
        flash('ARS creado exitosamente', 'success')
        return redirect(url_for('facturacion_ars'))
    
    return render_template('facturacion/ars_form.html', ars=None)

@app.route('/facturacion/ars/<int:ars_id>/editar', methods=['GET', 'POST'])
@login_required
def facturacion_ars_editar(ars_id):
    """Editar ARS"""
    conn = get_db_connection()
    
    if request.method == 'POST':
        nombre_ars = sanitize_input(request.form['nombre_ars'], 200)
        rnc = sanitize_input(request.form['rnc'], 50)

        # Validación de caracteres (integridad)
        try:
            nombre_ars = _validate_or_raise(nombre_ars, _RE_TEXTO_GENERAL, 'Nombre del ARS', max_len=200)
            rnc = _validate_or_raise(rnc, _RE_RNC, 'RNC', max_len=50)
        except ValueError as ve:
            flash(f'❌ {str(ve)}', 'error')
            return redirect(url_for('facturacion_ars_editar', ars_id=ars_id))
        
        if not nombre_ars or not rnc:
            flash('Todos los campos son obligatorios', 'error')
            return redirect(url_for('facturacion_ars_editar', ars_id=ars_id))
        
        conn.execute('UPDATE ars SET nombre_ars = %s, rnc = %s WHERE id = %s', (nombre_ars, rnc, ars_id))
        conn.commit()
        conn.close()
        
        flash('ARS actualizado exitosamente', 'success')
        return redirect(url_for('facturacion_ars'))
    
    ars = conn.execute('SELECT * FROM ars WHERE id = %s', (ars_id,)).fetchone()
    conn.close()
    
    return render_template('facturacion/ars_form.html', ars=ars)

@app.route('/facturacion/ars/<int:ars_id>/eliminar', methods=['POST'])
@login_required
def facturacion_ars_eliminar(ars_id):
    """Eliminar ARS (soft delete)"""
    conn = get_db_connection()
    # Verificar si tiene códigos ARS relacionados
    result = conn.execute('SELECT COUNT(*) FROM codigo_ars WHERE ars_id = %s AND activo = 1', (ars_id,)).fetchone()
    relacionados = result[list(result.keys())[0]] if isinstance(result, dict) else result[0]
    
    if relacionados > 0:
        conn.close()
        flash(f'No se puede eliminar. Hay {relacionados} código(s) ARS asociados.', 'error')
        return redirect(url_for('facturacion_ars'))
    
    conn.execute('UPDATE ars SET activo = 0 WHERE id = %s', (ars_id,))
    conn.commit()
    conn.close()
    
    flash('ARS eliminado exitosamente', 'success')
    return redirect(url_for('facturacion_ars'))

# ========== MAESTRA DE MÉDICOS ==========
@app.route('/facturacion/medicos')
@login_required
def facturacion_medicos():
    """Lista de Médicos"""
    search = request.args.get('search', '')
    conn = get_db_connection()
    
    if search:
        search_pattern = f"%{search}%"
        medicos_list = conn.execute(
            "SELECT * FROM medicos WHERE (nombre LIKE %s OR cedula LIKE %s OR especialidad LIKE %s) AND activo = 1 ORDER BY nombre",
            (search_pattern, search_pattern, search_pattern)
        ).fetchall()
    else:
        medicos_list = conn.execute('SELECT * FROM medicos WHERE activo = 1 ORDER BY nombre').fetchall()
    
    conn.close()
    return render_template('facturacion/medicos.html', medicos_list=medicos_list, search=search)

@app.route('/facturacion/medicos/nuevo', methods=['GET', 'POST'])
@login_required
def facturacion_medicos_nuevo():
    """Crear nuevo Médico"""
    if request.method == 'POST':
        nombre = sanitize_input(request.form['nombre'], 200)
        especialidad = sanitize_input(request.form['especialidad'], 100)
        cedula = sanitize_input(request.form['cedula'], 50)
        exequatur = sanitize_input(request.form.get('exequatur', ''), 50)
        email = sanitize_input(request.form.get('email', ''), 100)
        factura = 1 if request.form.get('factura') == '1' else 0
        
        if not nombre or not especialidad or not cedula:
            flash('Nombre, Especialidad y Cédula son obligatorios', 'error')
            return redirect(url_for('facturacion_medicos_nuevo'))
        
        # Validar email si se proporciona
        if email and not validate_email(email):
            flash('Formato de email inválido', 'error')
            return redirect(url_for('facturacion_medicos_nuevo'))
        
        conn = get_db_connection()
        
        # Verificar si la cédula ya existe
        existe = conn.execute('SELECT id FROM medicos WHERE cedula = %s', (cedula,)).fetchone()
        if existe:
            conn.close()
            flash('Ya existe un médico con esa cédula', 'error')
            return redirect(url_for('facturacion_medicos_nuevo'))
        
        # Verificar si el email ya existe (si se proporciona)
        if email:
            existe_email = conn.execute('SELECT id FROM medicos WHERE email = %s', (email,)).fetchone()
            if existe_email:
                conn.close()
                flash('Ya existe un médico con ese correo electrónico', 'error')
                return redirect(url_for('facturacion_medicos_nuevo'))
        
        # Verificar si el exequatur ya existe (si se proporciona)
        if exequatur:
            existe_exequatur = conn.execute('SELECT id FROM medicos WHERE exequatur = %s', (exequatur,)).fetchone()
            if existe_exequatur:
                conn.close()
                flash('Ya existe un médico con ese exequátur', 'error')
                return redirect(url_for('facturacion_medicos_nuevo'))
        
        conn.execute('INSERT INTO medicos (nombre, especialidad, cedula, exequatur, email, factura) VALUES (%s, %s, %s, %s, %s, %s)', 
                    (nombre, especialidad, cedula, exequatur, email, factura))
        conn.commit()
        conn.close()
        
        flash('Médico creado exitosamente', 'success')
        return redirect(url_for('facturacion_medicos'))
    
    return render_template('facturacion/medicos_form.html', medico=None)

@app.route('/facturacion/medicos/<int:medico_id>/editar', methods=['GET', 'POST'])
@login_required
def facturacion_medicos_editar(medico_id):
    """Editar Médico"""
    conn = get_db_connection()
    
    if request.method == 'POST':
        nombre = sanitize_input(request.form['nombre'], 200)
        especialidad = sanitize_input(request.form['especialidad'], 100)
        cedula = sanitize_input(request.form['cedula'], 50)
        exequatur = sanitize_input(request.form.get('exequatur', ''), 50)
        email = sanitize_input(request.form.get('email', ''), 100)
        factura = 1 if request.form.get('factura') == '1' else 0
        
        if not nombre or not especialidad or not cedula:
            flash('Nombre, Especialidad y Cédula son obligatorios', 'error')
            return redirect(url_for('facturacion_medicos_editar', medico_id=medico_id))
        
        # Validar email si se proporciona
        if email and not validate_email(email):
            flash('Formato de email inválido', 'error')
            return redirect(url_for('facturacion_medicos_editar', medico_id=medico_id))
        
        # Verificar si la cédula ya existe en otro médico
        existe = conn.execute('SELECT id FROM medicos WHERE cedula = %s AND id != %s', (cedula, medico_id)).fetchone()
        if existe:
            conn.close()
            flash('Ya existe otro médico con esa cédula', 'error')
            return redirect(url_for('facturacion_medicos_editar', medico_id=medico_id))
        
        # Verificar si el email ya existe en otro médico (si se proporciona)
        if email:
            existe_email = conn.execute('SELECT id FROM medicos WHERE email = %s AND id != %s', (email, medico_id)).fetchone()
            if existe_email:
                conn.close()
                flash('Ya existe otro médico con ese correo electrónico', 'error')
                return redirect(url_for('facturacion_medicos_editar', medico_id=medico_id))
        
        # Verificar si el exequatur ya existe en otro médico (si se proporciona)
        if exequatur:
            existe_exequatur = conn.execute('SELECT id FROM medicos WHERE exequatur = %s AND id != %s', (exequatur, medico_id)).fetchone()
            if existe_exequatur:
                conn.close()
                flash('Ya existe otro médico con ese exequátur', 'error')
                return redirect(url_for('facturacion_medicos_editar', medico_id=medico_id))
        
        conn.execute('UPDATE medicos SET nombre = %s, especialidad = %s, cedula = %s, exequatur = %s, email = %s, factura = %s WHERE id = %s', 
                    (nombre, especialidad, cedula, exequatur, email, factura, medico_id))
        conn.commit()
        conn.close()
        
        flash('Médico actualizado exitosamente', 'success')
        return redirect(url_for('facturacion_medicos'))
    
    medico = conn.execute('SELECT * FROM medicos WHERE id = %s', (medico_id,)).fetchone()
    conn.close()
    
    return render_template('facturacion/medicos_form.html', medico=medico)

@app.route('/facturacion/medicos/<int:medico_id>/eliminar', methods=['POST'])
@login_required
def facturacion_medicos_eliminar(medico_id):
    """Eliminar Médico (soft delete)"""
    conn = get_db_connection()
    # Verificar si tiene códigos ARS relacionados
    result = conn.execute('SELECT COUNT(*) FROM codigo_ars WHERE medico_id = %s AND activo = 1', (medico_id,)).fetchone()
    relacionados = result[list(result.keys())[0]] if isinstance(result, dict) else result[0]
    
    if relacionados > 0:
        conn.close()
        flash(f'No se puede eliminar. Hay {relacionados} código(s) ARS asociados.', 'error')
        return redirect(url_for('facturacion_medicos'))
    
    conn.execute('UPDATE medicos SET activo = 0 WHERE id = %s', (medico_id,))
    conn.commit()
    conn.close()
    
    flash('Médico eliminado exitosamente', 'success')
    return redirect(url_for('facturacion_medicos'))

# ========== MAESTRA DE CÓDIGO ARS ==========
@app.route('/facturacion/codigo-ars')
@login_required
def facturacion_codigo_ars():
    """Lista de Códigos ARS"""
    search = request.args.get('search', '')
    conn = get_db_connection()
    
    if search:
        search_pattern = f"%{search}%"
        codigos_list = conn.execute("""
            SELECT ca.*, m.nombre as medico_nombre, a.nombre_ars 
            FROM codigo_ars ca
            JOIN medicos m ON ca.medico_id = m.id
            JOIN ars a ON ca.ars_id = a.id
            WHERE (m.nombre LIKE %s OR a.nombre_ars LIKE %s OR ca.codigo_ars LIKE %s) AND ca.activo = 1
            ORDER BY m.nombre, a.nombre_ars
        """, (search_pattern, search_pattern, search_pattern)).fetchall()
    else:
        codigos_list = conn.execute('''
            SELECT ca.*, m.nombre as medico_nombre, a.nombre_ars 
            FROM codigo_ars ca
            JOIN medicos m ON ca.medico_id = m.id
            JOIN ars a ON ca.ars_id = a.id
            WHERE ca.activo = 1
            ORDER BY m.nombre, a.nombre_ars
        ''').fetchall()
    
    conn.close()
    return render_template('facturacion/codigo_ars.html', codigos_list=codigos_list, search=search)

@app.route('/facturacion/codigo-ars/nuevo', methods=['GET', 'POST'])
@login_required
def facturacion_codigo_ars_nuevo():
    """Crear nuevo Código ARS"""
    conn = get_db_connection()
    
    if request.method == 'POST':
        medico_id = request.form.get('medico_id')
        ars_id = request.form.get('ars_id')
        codigo_ars = sanitize_input(request.form['codigo_ars'], 50)
        
        if not medico_id or not ars_id or not codigo_ars:
            flash('Todos los campos son obligatorios', 'error')
            return redirect(url_for('facturacion_codigo_ars_nuevo'))
        
        # Verificar si ya existe la combinación médico-ars
        existe = conn.execute('SELECT id FROM codigo_ars WHERE medico_id = %s AND ars_id = %s AND activo = 1', 
                             (medico_id, ars_id)).fetchone()
        if existe:
            conn.close()
            flash('Ya existe un código para esta combinación de médico y ARS', 'error')
            return redirect(url_for('facturacion_codigo_ars_nuevo'))
        
        # Verificar si ya existe la combinación codigo_ars + ars_id
        existe_codigo = conn.execute('SELECT id FROM codigo_ars WHERE codigo_ars = %s AND ars_id = %s AND activo = 1', 
                                     (codigo_ars, ars_id)).fetchone()
        if existe_codigo:
            conn.close()
            flash('Ya existe ese código para la ARS seleccionada', 'error')
            return redirect(url_for('facturacion_codigo_ars_nuevo'))
        
        conn.execute('INSERT INTO codigo_ars (medico_id, ars_id, codigo_ars) VALUES (%s, %s, %s)', 
                    (medico_id, ars_id, codigo_ars))
        conn.commit()
        conn.close()
        
        flash('Código ARS creado exitosamente', 'success')
        return redirect(url_for('facturacion_codigo_ars'))
    
    # Obtener listas para los select
    medicos = conn.execute('SELECT id, nombre, especialidad FROM medicos WHERE activo = 1 ORDER BY nombre').fetchall()
    ars_list = conn.execute('SELECT id, nombre_ars FROM ars WHERE activo = 1 ORDER BY nombre_ars').fetchall()
    conn.close()
    
    return render_template('facturacion/codigo_ars_form.html', codigo=None, medicos=medicos, ars_list=ars_list)

@app.route('/facturacion/codigo-ars/<int:codigo_id>/editar', methods=['GET', 'POST'])
@login_required
def facturacion_codigo_ars_editar(codigo_id):
    """Editar Código ARS"""
    conn = get_db_connection()
    
    if request.method == 'POST':
        medico_id = request.form.get('medico_id')
        ars_id = request.form.get('ars_id')
        codigo_ars = sanitize_input(request.form['codigo_ars'], 50)
        
        if not medico_id or not ars_id or not codigo_ars:
            flash('Todos los campos son obligatorios', 'error')
            return redirect(url_for('facturacion_codigo_ars_editar', codigo_id=codigo_id))
        
        # Verificar si ya existe la combinación en otro registro
        existe = conn.execute('SELECT id FROM codigo_ars WHERE medico_id = %s AND ars_id = %s AND id != %s AND activo = 1', 
                             (medico_id, ars_id, codigo_id)).fetchone()
        if existe:
            conn.close()
            flash('Ya existe un código para esta combinación de médico y ARS', 'error')
            return redirect(url_for('facturacion_codigo_ars_editar', codigo_id=codigo_id))
        
        # Verificar si ya existe la combinación codigo_ars + ars_id en otro registro
        existe_codigo = conn.execute('SELECT id FROM codigo_ars WHERE codigo_ars = %s AND ars_id = %s AND id != %s AND activo = 1', 
                                     (codigo_ars, ars_id, codigo_id)).fetchone()
        if existe_codigo:
            conn.close()
            flash('Ya existe ese código para la ARS seleccionada', 'error')
            return redirect(url_for('facturacion_codigo_ars_editar', codigo_id=codigo_id))
        
        conn.execute('UPDATE codigo_ars SET medico_id = %s, ars_id = %s, codigo_ars = %s WHERE id = %s', 
                    (medico_id, ars_id, codigo_ars, codigo_id))
        conn.commit()
        conn.close()
        
        flash('Código ARS actualizado exitosamente', 'success')
        return redirect(url_for('facturacion_codigo_ars'))
    
    codigo = conn.execute('SELECT * FROM codigo_ars WHERE id = %s', (codigo_id,)).fetchone()
    medicos = conn.execute('SELECT id, nombre, especialidad FROM medicos WHERE activo = 1 ORDER BY nombre').fetchall()
    ars_list = conn.execute('SELECT id, nombre_ars FROM ars WHERE activo = 1 ORDER BY nombre_ars').fetchall()
    conn.close()
    
    return render_template('facturacion/codigo_ars_form.html', codigo=codigo, medicos=medicos, ars_list=ars_list)

@app.route('/facturacion/codigo-ars/<int:codigo_id>/eliminar', methods=['POST'])
@login_required
def facturacion_codigo_ars_eliminar(codigo_id):
    """Eliminar Código ARS (soft delete)"""
    conn = get_db_connection()
    conn.execute('UPDATE codigo_ars SET activo = 0 WHERE id = %s', (codigo_id,))
    conn.commit()
    conn.close()
    
    flash('Código ARS eliminado exitosamente', 'success')
    return redirect(url_for('facturacion_codigo_ars'))

# ========== MAESTRA DE CENTROS MÉDICOS ==========
@app.route('/facturacion/centros-medicos')
@login_required
def facturacion_centros_medicos():
    """Lista de Centros Médicos"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Crear tablas si no existen (por si acaso)
    try:
        cursor.execute(adapt_sql_for_database('''
            CREATE TABLE IF NOT EXISTS centros_medicos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre VARCHAR(255) NOT NULL,
                direccion TEXT NOT NULL,
                rnc VARCHAR(50) NOT NULL,
                telefono VARCHAR(50),
                activo BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        '''))
        cursor.execute(adapt_sql_for_database('''
            CREATE TABLE IF NOT EXISTS medico_centro (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                medico_id INTEGER NOT NULL,
                centro_id INTEGER NOT NULL,
                activo BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (medico_id) REFERENCES medicos(id),
                FOREIGN KEY (centro_id) REFERENCES centros_medicos(id),
                UNIQUE(medico_id, centro_id)
            )
        '''))
        conn.commit()
    except Exception as table_error:
        print(f"⚠️ Advertencia al crear tablas: {table_error}")
    
    search = request.args.get('search', '')
    
    if search:
        search_pattern = f"%{search}%"
        centros_list = conn.execute("""
            SELECT * FROM centros_medicos 
            WHERE (nombre LIKE %s OR rnc LIKE %s) AND activo = 1
            ORDER BY nombre
        """, (search_pattern, search_pattern)).fetchall()
    else:
        centros_list = conn.execute('''
            SELECT * FROM centros_medicos 
            WHERE activo = 1
            ORDER BY nombre
        ''').fetchall()
    
    conn.close()
    return render_template('facturacion/centros_medicos.html', centros_list=centros_list, search=search)

@app.route('/facturacion/centros-medicos/nuevo', methods=['GET', 'POST'])
@login_required
def facturacion_centros_medicos_nuevo():
    """Crear Centro Médico"""
    if request.method == 'POST':
        nombre = sanitize_input(request.form['nombre'], 255)
        direccion = sanitize_input(request.form['direccion'], 500)
        rnc = sanitize_input(request.form['rnc'], 50)
        telefono = sanitize_input(request.form.get('telefono', ''), 50)

        # Validación de caracteres (integridad)
        try:
            nombre = _validate_or_raise(nombre, _RE_TEXTO_GENERAL, 'Nombre del Centro Médico', max_len=255)
            rnc = _validate_or_raise(rnc, _RE_RNC, 'RNC', max_len=50)
        except ValueError as ve:
            flash(f'❌ {str(ve)}', 'error')
            return redirect(url_for('facturacion_centros_medicos_nuevo'))
        
        if not nombre or not direccion or not rnc:
            flash('Nombre, dirección y RNC son obligatorios', 'error')
            return redirect(url_for('facturacion_centros_medicos_nuevo'))
        
        conn = get_db_connection()
        
        # Verificar si ya existe el RNC
        existe = conn.execute('SELECT id FROM centros_medicos WHERE rnc = %s AND activo = 1', (rnc,)).fetchone()
        if existe:
            conn.close()
            flash('Ya existe un centro médico con ese RNC', 'error')
            return redirect(url_for('facturacion_centros_medicos_nuevo'))
        
        conn.execute('INSERT INTO centros_medicos (nombre, direccion, rnc, telefono) VALUES (%s, %s, %s, %s)', 
                    (nombre, direccion, rnc, telefono))
        conn.commit()
        conn.close()
        
        flash('Centro Médico creado exitosamente', 'success')
        return redirect(url_for('facturacion_centros_medicos'))
    
    return render_template('facturacion/centro_medico_form.html', centro=None)

@app.route('/facturacion/centros-medicos/<int:centro_id>/editar', methods=['GET', 'POST'])
@login_required
def facturacion_centros_medicos_editar(centro_id):
    """Editar Centro Médico"""
    conn = get_db_connection()
    
    if request.method == 'POST':
        nombre = sanitize_input(request.form['nombre'], 255)
        direccion = sanitize_input(request.form['direccion'], 500)
        rnc = sanitize_input(request.form['rnc'], 50)
        telefono = sanitize_input(request.form.get('telefono', ''), 50)

        # Validación de caracteres (integridad)
        try:
            nombre = _validate_or_raise(nombre, _RE_TEXTO_GENERAL, 'Nombre del Centro Médico', max_len=255)
            rnc = _validate_or_raise(rnc, _RE_RNC, 'RNC', max_len=50)
        except ValueError as ve:
            flash(f'❌ {str(ve)}', 'error')
            return redirect(url_for('facturacion_centros_medicos_editar', centro_id=centro_id))
        
        if not nombre or not direccion or not rnc:
            flash('Nombre, dirección y RNC son obligatorios', 'error')
            return redirect(url_for('facturacion_centros_medicos_editar', centro_id=centro_id))
        
        # Verificar si ya existe el RNC en otro registro
        existe = conn.execute('SELECT id FROM centros_medicos WHERE rnc = %s AND id != %s AND activo = 1', 
                             (rnc, centro_id)).fetchone()
        if existe:
            conn.close()
            flash('Ya existe un centro médico con ese RNC', 'error')
            return redirect(url_for('facturacion_centros_medicos_editar', centro_id=centro_id))
        
        conn.execute('UPDATE centros_medicos SET nombre = %s, direccion = %s, rnc = %s, telefono = %s WHERE id = %s', 
                    (nombre, direccion, rnc, telefono, centro_id))
        conn.commit()
        conn.close()
        
        flash('Centro Médico actualizado exitosamente', 'success')
        return redirect(url_for('facturacion_centros_medicos'))
    
    centro = conn.execute('SELECT * FROM centros_medicos WHERE id = %s', (centro_id,)).fetchone()
    conn.close()
    
    return render_template('facturacion/centro_medico_form.html', centro=centro)

@app.route('/facturacion/centros-medicos/<int:centro_id>/eliminar', methods=['POST'])
@login_required
def facturacion_centros_medicos_eliminar(centro_id):
    """Eliminar Centro Médico (soft delete)"""
    conn = get_db_connection()
    conn.execute('UPDATE centros_medicos SET activo = 0 WHERE id = %s', (centro_id,))
    conn.commit()
    conn.close()
    
    flash('Centro Médico eliminado exitosamente', 'success')
    return redirect(url_for('facturacion_centros_medicos'))

# ========== MAESTRA DE CENTRO x MÉDICO ==========
@app.route('/facturacion/medico-centro')
@login_required
def facturacion_medico_centro():
    """Lista de Relación Médico-Centro"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Crear tablas si no existen (por si acaso)
    try:
        cursor.execute(adapt_sql_for_database('''
            CREATE TABLE IF NOT EXISTS centros_medicos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre VARCHAR(255) NOT NULL,
                direccion TEXT NOT NULL,
                rnc VARCHAR(50) NOT NULL,
                telefono VARCHAR(50),
                activo BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        '''))
        cursor.execute(adapt_sql_for_database('''
            CREATE TABLE IF NOT EXISTS medico_centro (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                medico_id INTEGER NOT NULL,
                centro_id INTEGER NOT NULL,
                activo BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (medico_id) REFERENCES medicos(id),
                FOREIGN KEY (centro_id) REFERENCES centros_medicos(id),
                UNIQUE(medico_id, centro_id)
            )
        '''))
        conn.commit()
    except Exception as table_error:
        print(f"⚠️ Advertencia al crear tablas: {table_error}")
    
    search = request.args.get('search', '')
    
    if search:
        search_pattern = f"%{search}%"
        relaciones_list = conn.execute("""
            SELECT mc.*, m.nombre as medico_nombre, m.especialidad, c.nombre as centro_nombre 
            FROM medico_centro mc
            JOIN medicos m ON mc.medico_id = m.id
            JOIN centros_medicos c ON mc.centro_id = c.id
            WHERE (m.nombre LIKE %s OR c.nombre LIKE %s) AND mc.activo = 1
            ORDER BY mc.es_defecto DESC, m.nombre, c.nombre
        """, (search_pattern, search_pattern)).fetchall()
    else:
        relaciones_list = conn.execute('''
            SELECT mc.*, m.nombre as medico_nombre, m.especialidad, c.nombre as centro_nombre 
            FROM medico_centro mc
            JOIN medicos m ON mc.medico_id = m.id
            JOIN centros_medicos c ON mc.centro_id = c.id
            WHERE mc.activo = 1
            ORDER BY mc.es_defecto DESC, m.nombre, c.nombre
        ''').fetchall()
    
    conn.close()
    return render_template('facturacion/medico_centro.html', relaciones_list=relaciones_list, search=search)

@app.route('/facturacion/medico-centro/nuevo', methods=['GET', 'POST'])
@login_required
def facturacion_medico_centro_nuevo():
    """Crear Relación Médico-Centro"""
    conn = get_db_connection()
    
    if request.method == 'POST':
        medico_id = request.form.get('medico_id')
        centro_id = request.form.get('centro_id')
        es_defecto = 1 if request.form.get('es_defecto') == '1' else 0
        
        if not medico_id or not centro_id:
            flash('Debe seleccionar médico y centro médico', 'error')
            return redirect(url_for('facturacion_medico_centro_nuevo'))
        
        # Verificar si ya existe la relación
        existe = conn.execute('SELECT id FROM medico_centro WHERE medico_id = %s AND centro_id = %s AND activo = 1', 
                             (medico_id, centro_id)).fetchone()
        if existe:
            conn.close()
            flash('Ya existe esa relación entre el médico y el centro médico', 'error')
            return redirect(url_for('facturacion_medico_centro_nuevo'))
        
        # Si se marca como defecto, desmarcar otros centros de ese médico
        if es_defecto == 1:
            conn.execute('UPDATE medico_centro SET es_defecto = 0 WHERE medico_id = %s', (medico_id,))
        
        conn.execute('INSERT INTO medico_centro (medico_id, centro_id, es_defecto) VALUES (%s, %s, %s)', 
                    (medico_id, centro_id, es_defecto))
        conn.commit()
        conn.close()
        
        mensaje = 'Relación Médico-Centro creada exitosamente'
        if es_defecto == 1:
            mensaje += ' (marcada como centro por defecto)'
        flash(mensaje, 'success')
        return redirect(url_for('facturacion_medico_centro'))
    
    # Obtener listas para los select
    medicos = conn.execute('SELECT id, nombre, especialidad FROM medicos WHERE activo = 1 ORDER BY nombre').fetchall()
    centros = conn.execute('SELECT id, nombre FROM centros_medicos WHERE activo = 1 ORDER BY nombre').fetchall()
    conn.close()
    
    return render_template('facturacion/medico_centro_form.html', relacion=None, medicos=medicos, centros=centros)

@app.route('/facturacion/medico-centro/<int:relacion_id>/eliminar', methods=['POST'])
@login_required
def facturacion_medico_centro_eliminar(relacion_id):
    """Eliminar Relación Médico-Centro (soft delete)"""
    conn = get_db_connection()
    conn.execute('UPDATE medico_centro SET activo = 0 WHERE id = %s', (relacion_id,))
    conn.commit()
    conn.close()
    
    flash('Relación eliminada exitosamente', 'success')
    return redirect(url_for('facturacion_medico_centro'))

# ========== MAESTRA DE TIPOS DE SERVICIOS ==========
@app.route('/facturacion/servicios')
@login_required
def facturacion_servicios():
    """Lista de Tipos de Servicios"""
    search = request.args.get('search', '')
    conn = get_db_connection()
    
    if search:
        servicios_list = conn.execute(
            'SELECT * FROM tipos_servicios WHERE descripcion LIKE %s AND activo = 1 ORDER BY descripcion',
            (f'%{search}%',)
        ).fetchall()
    else:
        servicios_list = conn.execute('SELECT * FROM tipos_servicios WHERE activo = 1 ORDER BY descripcion').fetchall()
    
    conn.close()
    return render_template('facturacion/servicios.html', servicios_list=servicios_list, search=search)

@app.route('/facturacion/servicios/nuevo', methods=['GET', 'POST'])
@login_required
def facturacion_servicios_nuevo():
    """Crear nuevo Tipo de Servicio"""
    if request.method == 'POST':
        descripcion = sanitize_input(request.form['descripcion'], 200)
        precio_base = request.form.get('precio_base', 0)
        
        if not descripcion:
            flash('La descripción es obligatoria', 'error')
            return redirect(url_for('facturacion_servicios_nuevo'))
        
        try:
            precio_base = float(precio_base) if precio_base else 0
        except ValueError:
            precio_base = 0
        
        conn = get_db_connection()
        conn.execute('INSERT INTO tipos_servicios (descripcion, precio_base) VALUES (%s, %s)', 
                    (descripcion, precio_base))
        conn.commit()
        conn.close()
        
        flash('Tipo de servicio creado exitosamente', 'success')
        return redirect(url_for('facturacion_servicios'))
    
    return render_template('facturacion/servicios_form.html', servicio=None)

@app.route('/facturacion/servicios/<int:servicio_id>/editar', methods=['GET', 'POST'])
@login_required
def facturacion_servicios_editar(servicio_id):
    """Editar Tipo de Servicio"""
    conn = get_db_connection()
    
    if request.method == 'POST':
        descripcion = sanitize_input(request.form['descripcion'], 200)
        precio_base = request.form.get('precio_base', 0)
        
        if not descripcion:
            flash('La descripción es obligatoria', 'error')
            return redirect(url_for('facturacion_servicios_editar', servicio_id=servicio_id))
        
        try:
            precio_base = float(precio_base) if precio_base else 0
        except ValueError:
            precio_base = 0
        
        conn.execute('UPDATE tipos_servicios SET descripcion = %s, precio_base = %s WHERE id = %s', 
                    (descripcion, precio_base, servicio_id))
        conn.commit()
        conn.close()
        
        flash('Tipo de servicio actualizado exitosamente', 'success')
        return redirect(url_for('facturacion_servicios'))
    
    servicio = conn.execute('SELECT * FROM tipos_servicios WHERE id = %s', (servicio_id,)).fetchone()
    conn.close()
    
    return render_template('facturacion/servicios_form.html', servicio=servicio)

@app.route('/facturacion/servicios/<int:servicio_id>/eliminar', methods=['POST'])
@login_required
def facturacion_servicios_eliminar(servicio_id):
    """Eliminar Tipo de Servicio (soft delete)"""
    conn = get_db_connection()
    conn.execute('UPDATE tipos_servicios SET activo = 0 WHERE id = %s', (servicio_id,))
    conn.commit()
    conn.close()
    
    flash('Tipo de servicio eliminado exitosamente', 'success')
    return redirect(url_for('facturacion_servicios'))

# ========== MAESTRA DE PACIENTES ==========
@app.route('/facturacion/pacientes')
@login_required
def facturacion_pacientes():
    """Lista de Pacientes Únicos"""
    search = request.args.get('search', '')
    similitud = request.args.get('similitud') in ('1', 'true', 'True', 'on', 'yes', 'si')
    umbral_similitud = 80
    conn = get_db_connection()
    
    if search:
        pacientes_list = conn.execute('''
            SELECT p.*, a.nombre_ars,
                   (SELECT COUNT(*) FROM facturas_detalle fd WHERE fd.paciente_id = p.id) as total_registros
            FROM pacientes p
            LEFT JOIN ars a ON p.ars_id = a.id
            WHERE (p.nss LIKE %s OR p.nombre LIKE %s) AND p.activo = 1
            ORDER BY p.nombre
        ''', (f'%{search}%', f'%{search}%')).fetchall()
    else:
        pacientes_list = conn.execute('''
            SELECT p.*, a.nombre_ars,
                   (SELECT COUNT(*) FROM facturas_detalle fd WHERE fd.paciente_id = p.id) as total_registros
            FROM pacientes p
            LEFT JOIN ars a ON p.ars_id = a.id
            WHERE p.activo = 1
            ORDER BY p.nombre
        ''').fetchall()
    
    conn.close()

    # Filtro opcional: pacientes con similitud >= 80% (NSS + NOMBRE)
    # Nota: se aplica sobre el resultado actual (respetando la búsqueda).
    pacientes_out = [dict(p) for p in pacientes_list] if pacientes_list else []
    total_base = len(pacientes_out)
    total_similares = None

    if similitud and pacientes_out:
        from difflib import SequenceMatcher
        import re
        import unicodedata

        def _strip_accents(s: str) -> str:
            return ''.join(
                c for c in unicodedata.normalize('NFKD', s)
                if not unicodedata.combining(c)
            )

        def _norm_nss(v) -> str:
            s = '' if v is None else str(v)
            return re.sub(r'[^0-9]', '', s)

        def _norm_nombre(v) -> str:
            s = '' if v is None else str(v)
            s = _strip_accents(s).upper()
            # mantener letras/espacios
            s = re.sub(r'[^A-Z\s]', ' ', s)
            s = re.sub(r'\s+', ' ', s).strip()
            return s

        def _combo(p) -> str:
            return f"{p['_nss_norm']} {p['_nombre_norm']}".strip()

        # Pre-normalización
        for p in pacientes_out:
            p['_nss_norm'] = _norm_nss(p.get('nss'))
            p['_nombre_norm'] = _norm_nombre(p.get('nombre'))
            p['_combo'] = _combo(p)

        # Buckets para evitar O(n^2) full
        buckets = {}
        for idx, p in enumerate(pacientes_out):
            nss = p['_nss_norm']
            nom = p['_nombre_norm']
            keys = set()
            if nss:
                keys.add(f"nss6:{nss[:6]}")
                keys.add(f"nss4:{nss[-4:]}")
            if nom:
                keys.add(f"nom3:{nom[:3]}")
                # primer token
                first = nom.split(' ', 1)[0]
                if first:
                    keys.add(f"tok:{first[:5]}")
            for k in keys:
                buckets.setdefault(k, []).append(idx)

        compared = set()
        best = {}  # idx -> (score, other_idx)

        def _score(a: str, b: str) -> int:
            if not a or not b:
                return 0
            return int(round(SequenceMatcher(None, a, b).ratio() * 100))

        # Comparar dentro de buckets (limitando buckets enormes)
        for _, idxs in buckets.items():
            if len(idxs) < 2:
                continue
            # Si el bucket es demasiado grande, saltar para evitar degradación
            if len(idxs) > 400:
                continue
            for i_pos in range(len(idxs) - 1):
                i = idxs[i_pos]
                for j_pos in range(i_pos + 1, len(idxs)):
                    j = idxs[j_pos]
                    a, b = (i, j) if i < j else (j, i)
                    pair = (a, b)
                    if pair in compared:
                        continue
                    compared.add(pair)
                    s = _score(pacientes_out[i]['_combo'], pacientes_out[j]['_combo'])
                    if s >= umbral_similitud:
                        if (i not in best) or (s > best[i][0]):
                            best[i] = (s, j)
                        if (j not in best) or (s > best[j][0]):
                            best[j] = (s, i)

        # Anotar y filtrar salida
        filtrados = []
        for idx, p in enumerate(pacientes_out):
            if idx in best:
                s, other_idx = best[idx]
                other = pacientes_out[other_idx]
                p['similitud_score'] = s
                p['similitud_con_id'] = other.get('id')
                p['similitud_con_nss'] = other.get('nss')
                p['similitud_con_nombre'] = other.get('nombre')
                p['similitud_con_ars'] = other.get('nombre_ars')
                filtrados.append(p)
        pacientes_out = sorted(filtrados, key=lambda x: (-int(x.get('similitud_score', 0)), str(x.get('nombre', ''))))
        total_similares = len(pacientes_out)

    return render_template(
        'facturacion/pacientes.html',
        pacientes_list=pacientes_out,
        search=search,
        similitud=similitud,
        umbral_similitud=umbral_similitud,
        total_base=total_base,
        total_similares=total_similares
    )

@app.route('/facturacion/pacientes/excel')
@login_required
def facturacion_pacientes_excel():
    """Descargar lista de pacientes en Excel"""
    from datetime import datetime
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        search = request.args.get('search', '')
        conn = get_db_connection()
        
        # Obtener pacientes con los mismos filtros que la vista
        if search:
            pacientes_list = conn.execute('''
                SELECT p.*, a.nombre_ars,
                       (SELECT COUNT(*) FROM facturas_detalle fd WHERE fd.paciente_id = p.id) as total_registros
                FROM pacientes p
                LEFT JOIN ars a ON p.ars_id = a.id
                WHERE (p.nss LIKE %s OR p.nombre LIKE %s) AND p.activo = 1
                ORDER BY p.nombre
            ''', (f'%{search}%', f'%{search}%')).fetchall()
        else:
            pacientes_list = conn.execute('''
                SELECT p.*, a.nombre_ars,
                       (SELECT COUNT(*) FROM facturas_detalle fd WHERE fd.paciente_id = p.id) as total_registros
                FROM pacientes p
                LEFT JOIN ars a ON p.ars_id = a.id
                WHERE p.activo = 1
                ORDER BY p.nombre
            ''').fetchall()
        
        conn.close()
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pacientes"
        
        # Estilos
        header_fill = PatternFill(start_color="AB9B9F", end_color="AB9B9F", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Encabezados
        headers = ['ID', 'NSS', 'Nombre', 'ARS', 'Total Registros']
        ws.append(headers)
        
        # Aplicar estilos a encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Agregar datos
        for paciente in pacientes_list:
            ws.append([
                paciente['id'],
                paciente['nss'],
                paciente['nombre'],
                paciente['nombre_ars'] if paciente['nombre_ars'] else 'Sin ARS',
                paciente['total_registros']
            ])
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 18
        
        # Guardar en BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Nombre del archivo
        filename = f"pacientes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except ImportError:
        flash('⚠️ openpyxl no está instalado. Instala con: pip install openpyxl', 'error')
        return redirect(url_for('facturacion_pacientes'))
    except Exception as e:
        print(f"Error al generar Excel: {e}")
        flash(f'Error al generar Excel: {str(e)}', 'error')
        return redirect(url_for('facturacion_pacientes'))

# ========== MAESTRA DE NCF ==========
@app.route('/facturacion/ncf')
@login_required
def facturacion_ncf():
    """Lista de NCF"""
    search = request.args.get('search', '')
    conn = get_db_connection()
    
    if search:
        ncf_list = conn.execute(
            'SELECT * FROM ncf WHERE (tipo LIKE %s OR prefijo LIKE %s) AND activo = 1 ORDER BY tipo',
            (f'%{search}%', f'%{search}%')
        ).fetchall()
    else:
        ncf_list = conn.execute('SELECT * FROM ncf WHERE activo = 1 ORDER BY tipo').fetchall()
    
    # Mapa de uso en facturas (para bloquear eliminación en UI)
    usados_rows = conn.execute('SELECT ncf_id, COUNT(*) as total FROM facturas GROUP BY ncf_id').fetchall()
    ncf_usados = {row['ncf_id']: row['total'] for row in usados_rows} if usados_rows else {}

    # Calcular el próximo número para cada registro
    ncf_with_proximo = []
    for ncf in ncf_list:
        ncf_dict = dict(ncf)
        ncf_dict['proximo_numero'] = ncf_dict['ultimo_numero'] + 1
        ncf_dict['facturas_asignadas'] = int(ncf_usados.get(ncf_dict['id'], 0) or 0)
        ncf_with_proximo.append(ncf_dict)
    
    conn.close()
    return render_template('facturacion/ncf.html', ncf_list=ncf_with_proximo, search=search)

@app.route('/facturacion/ncf/nuevo', methods=['GET', 'POST'])
@login_required
def facturacion_ncf_nuevo():
    """Crear nuevo NCF"""
    if request.method == 'POST':
        tipo = sanitize_input(request.form['tipo'], 100)
        prefijo = sanitize_input(request.form['prefijo'], 10)
        tamaño = request.form.get('tamaño', 0)
        ultimo_numero = request.form.get('ultimo_numero', 0)
        fecha_fin = request.form.get('fecha_fin', '').strip()
        
        # Convertir fecha_fin de dd/mm/yyyy a yyyy-mm-dd si tiene valor
        fecha_fin_db = None
        if fecha_fin:
            try:
                # Intentar parsear dd/mm/yyyy
                from datetime import datetime
                fecha_obj = datetime.strptime(fecha_fin, '%d/%m/%Y')
                fecha_fin_db = fecha_obj.strftime('%Y-%m-%d')
            except:
                # Si falla, intentar yyyy-mm-dd (por si viene de formato antiguo)
                try:
                    fecha_obj = datetime.strptime(fecha_fin, '%Y-%m-%d')
                    fecha_fin_db = fecha_fin
                except:
                    flash('Formato de fecha inválido. Use dd/mm/yyyy (ejemplo: 31/12/2026)', 'error')
                    return redirect(url_for('facturacion_ncf_nuevo'))
        
        if not tipo or not prefijo or not tamaño:
            flash('Los campos Tipo, Prefijo y Tamaño son obligatorios', 'error')
            return redirect(url_for('facturacion_ncf_nuevo'))
        
        try:
            tamaño = int(tamaño)
            ultimo_numero = int(ultimo_numero) if ultimo_numero else 0
        except ValueError:
            flash('Tamaño y Último Número deben ser números enteros', 'error')
            return redirect(url_for('facturacion_ncf_nuevo'))
        
        conn = get_db_connection()
        conn.execute('INSERT INTO ncf (tipo, prefijo, tamaño, ultimo_numero, fecha_fin) VALUES (%s, %s, %s, %s, %s)', 
                    (tipo, prefijo, tamaño, ultimo_numero, fecha_fin_db))
        conn.commit()
        conn.close()
        
        flash('NCF creado exitosamente', 'success')
        return redirect(url_for('facturacion_ncf'))
    
    return render_template('facturacion/ncf_form.html', ncf=None)

@app.route('/facturacion/ncf/<int:ncf_id>/editar', methods=['GET', 'POST'])
@login_required
def facturacion_ncf_editar(ncf_id):
    """Editar NCF"""
    conn = get_db_connection()
    
    if request.method == 'POST':
        tipo = sanitize_input(request.form['tipo'], 100)
        prefijo = sanitize_input(request.form['prefijo'], 10)
        tamaño = request.form.get('tamaño', 0)
        ultimo_numero = request.form.get('ultimo_numero', 0)
        fecha_fin = request.form.get('fecha_fin', '').strip()
        
        # Convertir fecha_fin de dd/mm/yyyy a yyyy-mm-dd si tiene valor
        fecha_fin_db = None
        if fecha_fin:
            try:
                # Intentar parsear dd/mm/yyyy
                from datetime import datetime
                fecha_obj = datetime.strptime(fecha_fin, '%d/%m/%Y')
                fecha_fin_db = fecha_obj.strftime('%Y-%m-%d')
            except:
                # Si falla, intentar yyyy-mm-dd (por si viene de formato antiguo)
                try:
                    fecha_obj = datetime.strptime(fecha_fin, '%Y-%m-%d')
                    fecha_fin_db = fecha_fin
                except:
                    flash('Formato de fecha inválido. Use dd/mm/yyyy (ejemplo: 31/12/2026)', 'error')
                    return redirect(url_for('facturacion_ncf_editar', ncf_id=ncf_id))
        
        if not tipo or not prefijo or not tamaño:
            flash('Los campos Tipo, Prefijo y Tamaño son obligatorios', 'error')
            return redirect(url_for('facturacion_ncf_editar', ncf_id=ncf_id))
        
        try:
            tamaño = int(tamaño)
            ultimo_numero = int(ultimo_numero) if ultimo_numero else 0
        except ValueError:
            flash('Tamaño y Último Número deben ser números enteros', 'error')
            return redirect(url_for('facturacion_ncf_editar', ncf_id=ncf_id))
        
        conn.execute('UPDATE ncf SET tipo = %s, prefijo = %s, tamaño = %s, ultimo_numero = %s, fecha_fin = %s WHERE id = %s', 
                    (tipo, prefijo, tamaño, ultimo_numero, fecha_fin_db, ncf_id))
        conn.commit()
        conn.close()
        
        flash('NCF actualizado exitosamente', 'success')
        return redirect(url_for('facturacion_ncf'))
    
    ncf = conn.execute('SELECT * FROM ncf WHERE id = %s', (ncf_id,)).fetchone()
    conn.close()
    
    return render_template('facturacion/ncf_form.html', ncf=ncf)

@app.route('/facturacion/ncf/<int:ncf_id>/eliminar', methods=['POST'])
@login_required
def facturacion_ncf_eliminar(ncf_id):
    """Eliminar NCF (soft delete)"""
    conn = get_db_connection()
    try:
        # No permitir eliminar si ya está asignado a alguna factura
        uso = conn.execute(
            'SELECT COUNT(*) as total FROM facturas WHERE ncf_id = %s',
            (ncf_id,)
        ).fetchone()
        total_uso = uso['total'] if uso and 'total' in uso else 0
        if total_uso and int(total_uso) > 0:
            flash(f'❌ No se puede eliminar este NCF porque ya está asignado a {total_uso} factura(s).', 'error')
            return redirect(url_for('facturacion_ncf'))

        conn.execute('UPDATE ncf SET activo = 0 WHERE id = %s', (ncf_id,))
        conn.commit()
        flash('NCF eliminado exitosamente', 'success')
        return redirect(url_for('facturacion_ncf'))
    finally:
        conn.close()

# ========== PACIENTES A FACTURAR ==========
@app.route('/facturacion/pacientes-pendientes')
@login_required
def facturacion_pacientes_pendientes():
    """Lista de pacientes pendientes/facturados - Estado de Facturación"""
    # Obtener filtros opcionales
    medico_id = request.args.get('medico_id', type=int)
    ars_id = request.args.get('ars_id', type=int)
    estado = request.args.get('estado', default='pendiente')  # Por defecto: pendiente
    
    # IDs recién agregados (para resaltar 1 sola vez)
    highlight_ids = session.pop('highlight_pendientes_ids', [])
    if highlight_ids is None:
        highlight_ids = []
    # Link disponible si hay registros recientes (PDF constancia)
    tiene_recientes = bool(session.get('ultimos_pacientes_agregados', []))

    conn = get_db_connection()
    
    # Construir query con filtros opcionales
    # NOTA: Para pendientes usamos medico_consulta, para facturados usamos medico_id
    if estado == 'pendiente':
        query = '''
            SELECT fd.*, m.nombre as medico_nombre, a.nombre_ars, 
                   COALESCE(p.nombre, fd.nombre_paciente) as paciente_nombre_completo
            FROM facturas_detalle fd
            LEFT JOIN medicos m ON fd.medico_consulta = m.id
            JOIN ars a ON fd.ars_id = a.id
            LEFT JOIN pacientes p ON fd.paciente_id = p.id
            WHERE fd.estado = %s
        '''
    else:
        query = '''
            SELECT fd.*, m.nombre as medico_nombre, a.nombre_ars, 
                   COALESCE(p.nombre, fd.nombre_paciente) as paciente_nombre_completo
            FROM facturas_detalle fd
            LEFT JOIN medicos m ON fd.medico_id = m.id
            JOIN ars a ON fd.ars_id = a.id
            LEFT JOIN pacientes p ON fd.paciente_id = p.id
            WHERE fd.estado = %s
        '''
    params = [estado]
    
    # Si el usuario tiene rol "Registro de Facturas", filtrar solo sus pacientes
    if current_user.perfil == 'Registro de Facturas':
        query += ' AND m.email = %s'
        params.append(current_user.email)
    
    if medico_id:
        # Filtrar por el campo correspondiente según el estado
        if estado == 'pendiente':
            query += ' AND fd.medico_consulta = %s'
        else:
            query += ' AND fd.medico_id = %s'
        params.append(medico_id)
    
    if ars_id:
        query += ' AND fd.ars_id = %s'
        params.append(ars_id)
    
    query += ' ORDER BY fd.created_at DESC'
    
    pendientes = conn.execute(query, params).fetchall()
    
    # Obtener listas para los filtros
    # Filtrar médicos según el perfil del usuario
    if current_user.perfil == 'Registro de Facturas':
        # Solo mostrar el médico con el mismo email del usuario
        medicos = conn.execute('SELECT * FROM medicos WHERE activo = 1 AND email = %s ORDER BY nombre', (current_user.email,)).fetchall()
    else:
        # Administrador: mostrar todos los médicos
        medicos = conn.execute('SELECT * FROM medicos WHERE activo = 1 ORDER BY nombre').fetchall()
    
    ars_list = conn.execute('SELECT * FROM ars WHERE activo = 1 ORDER BY nombre_ars').fetchall()
    
    # Obtener nombres seleccionados para mostrar en el template
    medico_seleccionado = None
    ars_seleccionada = None
    
    if medico_id:
        medico = conn.execute('SELECT nombre FROM medicos WHERE id = %s', (medico_id,)).fetchone()
        if medico:
            medico_seleccionado = medico['nombre']
    
    if ars_id:
        ars = conn.execute('SELECT nombre_ars FROM ars WHERE id = %s', (ars_id,)).fetchone()
        if ars:
            ars_seleccionada = ars['nombre_ars']
    
    conn.close()
    
    return render_template('facturacion/pacientes_pendientes.html', 
                         pendientes=pendientes,
                         medicos=medicos,
                         ars_list=ars_list,
                         medico_id_filtro=medico_id,
                         ars_id_filtro=ars_id,
                         estado_filtro=estado,
                         medico_seleccionado=medico_seleccionado,
                         ars_seleccionada=ars_seleccionada,
                         highlight_ids=highlight_ids,
                         tiene_recientes=tiene_recientes)

@app.route('/facturacion/pacientes-pendientes/pdf')
@login_required
def facturacion_pacientes_pendientes_pdf():
    """Generar PDF de pacientes pendientes/facturados con filtros opcionales"""
    if not REPORTLAB_AVAILABLE:
        flash('ReportLab no está instalado. Instalar con: pip install reportlab', 'error')
        return redirect(url_for('facturacion_pacientes_pendientes'))
    
    # Obtener filtros opcionales
    medico_id = request.args.get('medico_id', type=int)
    ars_id = request.args.get('ars_id', type=int)
    estado = request.args.get('estado', default='pendiente')
    
    conn = get_db_connection()
    
    # Construir query con filtros opcionales (mismo que la vista)
    # NOTA: Para pendientes usamos medico_consulta, para facturados usamos medico_id
    if estado == 'pendiente':
        query = '''
            SELECT fd.*, m.nombre as medico_nombre, a.nombre_ars, 
                   COALESCE(p.nombre, fd.nombre_paciente) as paciente_nombre_completo
            FROM facturas_detalle fd
            LEFT JOIN medicos m ON fd.medico_consulta = m.id
            JOIN ars a ON fd.ars_id = a.id
            LEFT JOIN pacientes p ON fd.paciente_id = p.id
            WHERE fd.estado = %s
        '''
    else:
        query = '''
            SELECT fd.*, m.nombre as medico_nombre, a.nombre_ars, 
                   COALESCE(p.nombre, fd.nombre_paciente) as paciente_nombre_completo
            FROM facturas_detalle fd
            LEFT JOIN medicos m ON fd.medico_id = m.id
            JOIN ars a ON fd.ars_id = a.id
            LEFT JOIN pacientes p ON fd.paciente_id = p.id
            WHERE fd.estado = %s
        '''
    params = [estado]
    
    # Si el usuario tiene rol "Registro de Facturas", filtrar solo sus pacientes
    if current_user.perfil == 'Registro de Facturas':
        query += ' AND m.email = %s'
        params.append(current_user.email)
    
    if medico_id:
        # Filtrar por el campo correspondiente según el estado
        if estado == 'pendiente':
            query += ' AND fd.medico_consulta = %s'
        else:
            query += ' AND fd.medico_id = %s'
        params.append(medico_id)
    
    if ars_id:
        query += ' AND fd.ars_id = %s'
        params.append(ars_id)
    
    query += ' ORDER BY fd.created_at DESC'
    
    pendientes = conn.execute(query, params).fetchall()
    
    # Obtener nombres para el encabezado si hay filtros
    medico_nombre = None
    ars_nombre = None
    
    if medico_id:
        medico = conn.execute('SELECT nombre FROM medicos WHERE id = %s', (medico_id,)).fetchone()
        if medico:
            medico_nombre = medico['nombre']
    
    if ars_id:
        ars = conn.execute('SELECT nombre_ars FROM ars WHERE id = %s', (ars_id,)).fetchone()
        if ars:
            ars_nombre = ars['nombre_ars']
    
    conn.close()
    
    # Crear PDF en memoria
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    
    # Contenedor de elementos
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#CEB0B7'),
        spaceAfter=10,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#ACACAD'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    # Agregar encabezado
    elements.append(Paragraph("Dra. Shirley Ramírez", title_style))
    elements.append(Paragraph("Ginecóloga y Obstetra", subtitle_style))
    elements.append(Spacer(1, 12))
    
    # Título del documento
    doc_title = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#282828'),
        spaceAfter=10,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    # Título dinámico según el estado
    titulo_documento = "PACIENTES PENDIENTES DE FACTURACIÓN" if estado == 'pendiente' else "PACIENTES FACTURADOS"
    elements.append(Paragraph(titulo_documento, doc_title))
    elements.append(Paragraph(f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}", subtitle_style))
    
    # Mostrar filtros en el encabezado si están activos
    if medico_nombre or ars_nombre:
        filter_style = ParagraphStyle(
            'FilterInfo',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#CEB0B7'),
            spaceAfter=10,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        if medico_nombre:
            elements.append(Paragraph(f"<b>Médico:</b> {medico_nombre}", filter_style))
        if ars_nombre:
            elements.append(Paragraph(f"<b>ARS:</b> {ars_nombre}", filter_style))
    
    elements.append(Spacer(1, 20))
    
    if pendientes:
        # Determinar columnas según filtros
        # Sin filtros: Incluir Médico y ARS en cada línea
        # Con filtros: Excluir las columnas filtradas
        
        if not medico_nombre and not ars_nombre:
            # SIN FILTROS: Mostrar Médico y ARS en cada línea
            data = [['No.', 'NSS', 'Nombre', 'Fecha', 'Servicio', 'Monto', 'Médico', 'ARS']]
            col_widths = [0.5*inch, 1.1*inch, 1.5*inch, 0.8*inch, 1.6*inch, 0.9*inch, 1.2*inch, 1*inch]
            
            total = 0
            for idx, p in enumerate(pendientes, 1):
                data.append([
                    str(idx),
                    str(p['nss']),
                    str(p['nombre_paciente'])[:22],
                    formato_fecha_pdf(p['fecha_servicio']),
                    str(p['descripcion_servicio'])[:25],
                    f"{p['monto']:,.2f}",
                    str(p['medico_nombre'])[:18],
                    str(p['nombre_ars'])[:15]
                ])
                total += p['monto']
            
            # Agregar fila de total
            data.append(['', '', '', '', 'TOTAL:', f"{total:,.2f}", '', ''])
            total_col = 5
            
        elif medico_nombre and not ars_nombre:
            # FILTRO DE MÉDICO: No mostrar Médico en líneas, sí mostrar ARS
            data = [['No.', 'NSS', 'Nombre', 'Fecha', 'Servicio', 'Monto', 'ARS']]
            col_widths = [0.5*inch, 1.2*inch, 1.8*inch, 0.9*inch, 2*inch, 1*inch, 1.2*inch]
            
            total = 0
            for idx, p in enumerate(pendientes, 1):
                data.append([
                    str(idx),
                    str(p['nss']),
                    str(p['nombre_paciente'])[:25],
                    formato_fecha_pdf(p['fecha_servicio']),
                    str(p['descripcion_servicio'])[:30],
                    f"{p['monto']:,.2f}",
                    str(p['nombre_ars'])[:18]
                ])
                total += p['monto']
            
            data.append(['', '', '', '', 'TOTAL:', f"{total:,.2f}", ''])
            total_col = 5
            
        elif not medico_nombre and ars_nombre:
            # FILTRO DE ARS: No mostrar ARS en líneas, sí mostrar Médico
            data = [['No.', 'NSS', 'Nombre', 'Fecha', 'Servicio', 'Monto', 'Médico']]
            col_widths = [0.5*inch, 1.2*inch, 1.8*inch, 0.9*inch, 2*inch, 1*inch, 1.5*inch]
            
            total = 0
            for idx, p in enumerate(pendientes, 1):
                data.append([
                    str(idx),
                    str(p['nss']),
                    str(p['nombre_paciente'])[:25],
                    formato_fecha_pdf(p['fecha_servicio']),
                    str(p['descripcion_servicio'])[:30],
                    f"{p['monto']:,.2f}",
                    str(p['medico_nombre'])[:22]
                ])
                total += p['monto']
            
            data.append(['', '', '', '', 'TOTAL:', f"{total:,.2f}", ''])
            total_col = 5
            
        else:
            # AMBOS FILTROS: No mostrar ni Médico ni ARS en líneas
            data = [['No.', 'NSS', 'Nombre', 'Fecha', 'Servicio', 'Monto']]
            col_widths = [0.6*inch, 1.5*inch, 2.2*inch, 1*inch, 2.5*inch, 1.2*inch]
            
            total = 0
            for idx, p in enumerate(pendientes, 1):
                data.append([
                    str(idx),
                    str(p['nss']),
                    str(p['nombre_paciente'])[:30],
                    formato_fecha_pdf(p['fecha_servicio']),
                    str(p['descripcion_servicio'])[:35],
                    f"{p['monto']:,.2f}"
                ])
                total += p['monto']
            
            data.append(['', '', '', '', 'TOTAL:', f"{total:,.2f}"])
            total_col = 5
        
        # Crear tabla con estilos
        t = Table(data, colWidths=col_widths)
        
        t.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#CEB0B7')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            
            # Datos
            ('BACKGROUND', (0, 1), (-1, -2), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -2), colors.HexColor('#282828')),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # No. centrado
            ('ALIGN', (5, 1), (5, -1), 'RIGHT'),    # Monto alineado a derecha
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F2E2E6')]),
            
            # Total
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ACACAD')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('ALIGN', (4, -1), (5, -1), 'RIGHT'),
            
            # Bordes
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#ACACAD')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        
        elements.append(t)
        elements.append(Spacer(1, 20))
        
        # Información adicional
        info_style = ParagraphStyle(
            'Info',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#666666'),
            alignment=TA_LEFT
        )
        elements.append(Paragraph(f"<b>Total de pacientes:</b> {len(pendientes)}", info_style))
        elements.append(Paragraph(f"<b>Monto total pendiente:</b> {total:,.2f}", info_style))
    else:
        elements.append(Paragraph("No hay pacientes pendientes de facturación.", subtitle_style))
    
    # Pie de página
    elements.append(Spacer(1, 30))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#999999'),
        alignment=TA_CENTER
    )
    elements.append(Paragraph("Este documento es una constancia de los pacientes pendientes de facturación.", footer_style))
    elements.append(Paragraph("Dra. Shirley Ramírez - Ginecóloga y Obstetra", footer_style))
    
    # Construir PDF
    doc.build(elements)
    
    # Obtener PDF del buffer
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'pacientes_pendientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    )

@app.route('/facturacion/pacientes-agregados/pdf')
@login_required
def facturacion_pacientes_agregados_pdf():
    """Generar PDF de los pacientes recién agregados"""
    if not REPORTLAB_AVAILABLE:
        flash('ReportLab no está instalado. Instalar con: pip install reportlab', 'error')
        return redirect(url_for('facturacion_pacientes_pendientes'))
    
    # Obtener los IDs de la sesión
    ids_agregados = session.get('ultimos_pacientes_agregados', [])
    
    if not ids_agregados:
        flash('No hay registros recientes para generar PDF', 'warning')
        return redirect(url_for('facturacion_pacientes_pendientes'))
    
    # Copiar los IDs para poder limpiar la sesión pero mantener la referencia
    ids_temp = ids_agregados.copy()
    # Limpiar la sesión
    session.pop('ultimos_pacientes_agregados', None)
    
    # Obtener los registros de la base de datos
    conn = get_db_connection()
    placeholders = ','.join(['%s' for _ in ids_agregados])
    pendientes = conn.execute(f'''
        SELECT fd.*, m.nombre as medico_nombre, m.email as medico_email, m.especialidad as medico_especialidad, 
               a.nombre_ars, COALESCE(p.nombre, fd.nombre_paciente) as paciente_nombre_completo
        FROM facturas_detalle fd
        LEFT JOIN medicos m ON fd.medico_consulta = m.id
        JOIN ars a ON fd.ars_id = a.id
        LEFT JOIN pacientes p ON fd.paciente_id = p.id
        WHERE fd.id IN ({placeholders})
        ORDER BY fd.id DESC
    ''', ids_agregados).fetchall()
    
    if not pendientes:
        conn.close()
        flash('No se encontraron los registros', 'error')
        return redirect(url_for('facturacion_pacientes_pendientes'))
    
    # Obtener datos del médico ANTES de cerrar la conexión
    medico_consulta_id = pendientes[0]['medico_consulta']
    medico_nombre = pendientes[0]['medico_nombre']
    medico_email = pendientes[0]['medico_email']
    medico_especialidad = pendientes[0]['medico_especialidad']
    
    # Cerrar conexión aquí, ya tenemos todos los datos
    conn.close()
    
    # Crear PDF en memoria
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    
    # Contenedor de elementos
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#CEB0B7'),
        spaceAfter=10,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#ACACAD'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    # Agregar encabezado con el nombre del médico seleccionado
    elements.append(Paragraph(medico_nombre, title_style))
    elements.append(Paragraph(medico_especialidad, subtitle_style))
    elements.append(Spacer(1, 12))
    
    # Título del documento
    doc_title = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#282828'),
        spaceAfter=10,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    elements.append(Paragraph("CONSTANCIA - PACIENTES AGREGADOS", doc_title))
    elements.append(Paragraph(f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 20))
    
    # Banner de confirmación
    confirmation_style = ParagraphStyle(
        'Confirmation',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#28a745'),
        spaceAfter=15,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    elements.append(Paragraph(f"✓ {len(pendientes)} PACIENTE(S) AGREGADO(S) EXITOSAMENTE", confirmation_style))
    elements.append(Paragraph("Estado: PENDIENTE DE FACTURACIÓN", subtitle_style))
    elements.append(Spacer(1, 15))
    
    # Crear tabla
    data = [['No.', 'NSS', 'Nombre', 'Fecha', 'Servicio', 'Monto', 'ARS']]
    
    total = 0
    for idx, p in enumerate(pendientes, 1):
        data.append([
            str(idx),
            str(p['nss']),
            str(p['nombre_paciente'])[:25],  # Limitar nombre
            formato_fecha_pdf(p['fecha_servicio']),
            str(p['descripcion_servicio'])[:30],  # Limitar servicio
            f"{p['monto']:,.2f}",
            str(p['nombre_ars'])[:15]  # Limitar ARS
        ])
        total += p['monto']
    
    # Agregar fila de total
    data.append(['', '', '', '', 'TOTAL:', f"{total:,.2f}", ''])
    
    # Crear tabla con estilos - ajustar anchos de columna
    t = Table(data, colWidths=[0.5*inch, 1.1*inch, 1.9*inch, 0.9*inch, 2*inch, 1*inch, 1.1*inch])
    
    t.setStyle(TableStyle([
        # Encabezado - Color oscuro #B89BA3
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#B89BA3')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        
        # Datos
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -2), colors.HexColor('#282828')),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # No. centrado
        ('ALIGN', (5, 1), (5, -1), 'RIGHT'),    # Monto alineado a derecha
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8F4F5')]),
        
        # Total - Color gris más oscuro
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#9A9A9B')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 11),
        ('ALIGN', (4, -1), (5, -1), 'RIGHT'),
        
        # Bordes
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#B89BA3')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 7),
    ]))
    
    elements.append(t)
    elements.append(Spacer(1, 20))
    
    # Información adicional
    info_style = ParagraphStyle(
        'Info',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#666666'),
        alignment=TA_LEFT
    )
    elements.append(Paragraph(f"<b>Total de pacientes agregados:</b> {len(pendientes)}", info_style))
    elements.append(Paragraph(f"<b>Monto total:</b> {total:,.2f}", info_style))
    elements.append(Paragraph(f"<b>Médico:</b> {pendientes[0]['medico_nombre']}", info_style))
    elements.append(Paragraph(f"<b>ARS:</b> {pendientes[0]['nombre_ars']}", info_style))
    
    # Nota importante
    elements.append(Spacer(1, 20))
    note_style = ParagraphStyle(
        'Note',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#dc3545'),
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    elements.append(Paragraph("IMPORTANTE: Estos pacientes están PENDIENTES DE FACTURACIÓN", note_style))
    
    # Pie de página
    elements.append(Spacer(1, 20))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#999999'),
        alignment=TA_CENTER
    )
    elements.append(Paragraph("Este documento es una constancia de los pacientes agregados al sistema.", footer_style))
    elements.append(Paragraph(f"{medico_nombre} - {medico_especialidad}", footer_style))
    
    # Construir PDF
    doc.build(elements)
    
    # Obtener PDF del buffer
    buffer.seek(0)
    
    # Enviar el PDF por email si el médico tiene email registrado
    if medico_email:
        # Crear una copia del buffer para el email
        email_buffer = BytesIO(buffer.getvalue())
        buffer.seek(0)  # Resetear el buffer original
        
        # Enviar email con PDF adjunto de manera asíncrona (no bloquea la descarga)
        threading.Thread(target=enviar_email_pdf_pacientes, args=(
            medico_email,
            medico_nombre,
            email_buffer,
            len(pendientes),
            total
        )).start()
        
        flash(f'✅ PDF enviado al correo: {medico_email}', 'success')
    else:
        flash('⚠️ El médico no tiene email registrado. Solo se descargará el PDF.', 'warning')
    
    # Enviar el PDF directamente para descargar
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'constancia_pacientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    )

@app.route('/facturacion/pacientes-pendientes/agregar', methods=['GET', 'POST'])
@login_required
def facturacion_facturas_nueva():
    """Agregar pacientes pendientes de facturación"""
    if request.method == 'POST':
        conn = None
        try:
            conn = get_db_connection()
            
            # Datos del encabezado
            medico_id = request.form.get('medico_id')
            ars_id = request.form.get('ars_id')
            
            # Datos de las líneas (vienen como JSON)
            import json
            lineas_json = request.form.get('lineas_json')
            lineas = json.loads(lineas_json) if lineas_json else []
            
            if not medico_id or not ars_id:
                flash('Médico y ARS son obligatorios', 'error')
                return redirect(url_for('facturacion_facturas_nueva'))
            
            if not lineas:
                flash('Debe agregar al menos un paciente', 'error')
                return redirect(url_for('facturacion_facturas_nueva'))
            
            cursor = conn.cursor()
            
            # Lista para guardar los IDs de los registros creados
            ids_creados = []
            errores = []
            errores_validacion = []
            
            # Procesar cada línea como PENDIENTE (sin factura_id)
            for idx, linea in enumerate(lineas, start=1):
                nss_raw = linea.get('nss', '')
                nombre_raw = linea.get('nombre', '')
                fecha_servicio = linea.get('fecha', '')
                autorizacion_raw = linea.get('autorizacion', '')
                servicio_raw = linea.get('servicio', '')
                monto_raw = linea.get('monto', 0)

                # Validaciones de caracteres (integridad)
                try:
                    nss = _validate_or_raise(nss_raw, _RE_NSS, 'NSS', idx=idx, max_len=20)
                    nombre = _validate_or_raise(nombre_raw, _RE_NOMBRE, 'Nombre del paciente', idx=idx, max_len=200)
                    servicio_desc = _validate_or_raise(servicio_raw, _RE_TEXTO_GENERAL, 'Servicio', idx=idx, max_len=200)
                    autorizacion = _validate_or_raise(autorizacion_raw, _RE_AUTORIZACION, 'Autorización', idx=idx, max_len=50, required=False)
                except ValueError as ve:
                    errores_validacion.append(str(ve))
                    continue

                # Monto numérico
                try:
                    monto = float(monto_raw or 0)
                except Exception:
                    errores_validacion.append(f'Línea {idx}: Monto inválido')
                    continue
                
                # VALIDACIÓN: Verificar si ya existe el mismo registro (NSS + FECHA + AUTORIZACIÓN + ARS)
                duplicado = conn.execute('''
                    SELECT * FROM facturas_detalle 
                    WHERE nss = %s AND fecha_servicio = %s AND autorizacion = %s AND ars_id = %s
                ''', (nss, fecha_servicio, autorizacion, ars_id)).fetchone()
                
                if duplicado:
                    errores.append(f'Línea {idx}: Registro duplicado - NSS {nss}, Fecha {fecha_servicio}, Autorización {autorizacion} ya existe')
                    continue
                
                # Buscar o crear paciente (llave única: NSS + ARS)
                paciente = conn.execute('SELECT * FROM pacientes WHERE nss = %s AND ars_id = %s', 
                                       (nss, ars_id)).fetchone()
                if paciente:
                    paciente_id = paciente['id']
                    # Actualizar nombre si cambió
                    conn.execute('UPDATE pacientes SET nombre = %s WHERE id = %s',
                               (nombre, paciente_id))
                else:
                    # Crear nuevo paciente con esta combinación NSS + ARS
                    cursor.execute('INSERT INTO pacientes (nss, nombre, ars_id) VALUES (%s, %s, %s)',
                                 (nss, nombre, ars_id))
                    paciente_id = cursor.lastrowid
                
                # Buscar o crear servicio
                servicio = conn.execute('SELECT * FROM tipos_servicios WHERE descripcion = %s AND activo = 1',
                                      (servicio_desc,)).fetchone()
                if servicio:
                    servicio_id = servicio['id']
                else:
                    # Crear servicio automáticamente
                    cursor.execute('INSERT INTO tipos_servicios (descripcion, precio_base) VALUES (%s, %s)',
                                 (servicio_desc, monto))
                    servicio_id = cursor.lastrowid
                
                # Insertar registro PENDIENTE (sin factura_id)
                # medico_id: Médico que FACTURA (NULL - se asignará al generar la factura)
                # medico_consulta: Médico que ATENDIÓ al paciente (el seleccionado en el formulario)
                cursor.execute('''
                    INSERT INTO facturas_detalle 
                    (paciente_id, nss, nombre_paciente, fecha_servicio, autorizacion, 
                     servicio_id, descripcion_servicio, monto, medico_id, medico_consulta, ars_id, estado)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NULL, %s, %s, 'pendiente')
                ''', (paciente_id, nss, nombre, fecha_servicio, autorizacion, 
                     servicio_id, servicio_desc, monto, medico_id, ars_id))
                
                # Guardar el ID del registro creado
                ids_creados.append(cursor.lastrowid)

            # Si hubo errores de validación, NO guardar nada (integridad)
            if errores_validacion:
                conn.rollback()
                flash('❌ No se guardó nada. Corrige los caracteres inválidos en los campos.', 'error')
                for err in errores_validacion[:8]:
                    flash(err, 'error')
                if len(errores_validacion) > 8:
                    flash(f'... y {len(errores_validacion) - 8} más.', 'warning')
                return redirect(url_for('facturacion_facturas_nueva'))
            
            # Si no se creó ningún registro, no confirmar cambios
            if not ids_creados:
                conn.rollback()
                if errores:
                    # Mostrar un resumen + primeros errores (evita spam)
                    flash(f'❌ No se agregó ningún paciente. Se detectaron {len(errores)} registro(s) duplicado(s).', 'error')
                    for error in errores[:5]:
                        flash(error, 'error')
                    if len(errores) > 5:
                        flash(f'... y {len(errores) - 5} más. Revise su Excel/tabla antes de reintentar.', 'warning')
                else:
                    flash('❌ No se agregó ningún paciente. Verifique los datos e intente nuevamente.', 'error')
                return redirect(url_for('facturacion_facturas_nueva'))

            # Confirmar registros creados (aunque existan duplicados omitidos)
            conn.commit()
            
            # Guardar los IDs en la sesión para generar el PDF
            session['ultimos_pacientes_agregados'] = ids_creados
            # IDs para resaltar en el listado (solo 1 vez)
            session['highlight_pendientes_ids'] = ids_creados
            session['descargar_pdf_pacientes'] = True  # Flag para disparar descarga automática
            
            total_enviado = len(lineas)
            total_creado = len(ids_creados)
            flash(f'✅ {total_creado} de {total_enviado} paciente(s) agregado(s) como PENDIENTES DE FACTURACIÓN', 'success')

            if errores:
                flash(f'⚠️ Se omitieron {len(errores)} registro(s) por duplicados. Puede revisar los detalles en los mensajes de error.', 'warning')
                for error in errores[:5]:
                    flash(error, 'warning')
                if len(errores) > 5:
                    flash(f'... y {len(errores) - 5} más.', 'warning')
            
            # Redirigir al estado de facturación filtrado para que el usuario vea lo guardado
            return redirect(url_for('facturacion_pacientes_pendientes',
                                    estado='pendiente',
                                    medico_id=medico_id,
                                    ars_id=ars_id))
            
        except Exception as e:
            if conn:
                conn.rollback()
            flash(f'Error al agregar pacientes: {str(e)}', 'error')
            return redirect(url_for('facturacion_facturas_nueva'))
        finally:
            if conn:
                conn.close()
    
    # GET: Mostrar formulario
    conn = get_db_connection()
    try:
        # Filtrar médicos según el perfil del usuario
        if current_user.perfil == 'Registro de Facturas':
            # Solo mostrar el médico con el mismo email del usuario
            medicos = conn.execute('SELECT * FROM medicos WHERE activo = 1 AND email = %s ORDER BY nombre', (current_user.email,)).fetchall()
        else:
            # Administrador: mostrar todos los médicos
            medicos = conn.execute('SELECT * FROM medicos WHERE activo = 1 ORDER BY nombre').fetchall()
        
        ars_list = conn.execute('SELECT * FROM ars WHERE activo = 1 ORDER BY nombre_ars').fetchall()
        servicios_list = conn.execute('SELECT * FROM tipos_servicios WHERE activo = 1 ORDER BY descripcion').fetchall()
        
        # Obtener centros médicos de todos los médicos
        centros_medicos = conn.execute('''
            SELECT mc.medico_id, mc.centro_id, mc.es_defecto, c.nombre as centro_nombre
            FROM medico_centro mc
            JOIN centros_medicos c ON mc.centro_id = c.id
            WHERE mc.activo = 1 AND c.activo = 1
            ORDER BY mc.medico_id, mc.es_defecto DESC, c.nombre
        ''').fetchall()
        
        # Verificar si hay que descargar el PDF automáticamente
        descargar_pdf = session.pop('descargar_pdf_pacientes', False)
        
        return render_template('facturacion/facturas_form.html', 
                             medicos=medicos, 
                             ars_list=ars_list, 
                             servicios_list=servicios_list,
                             centros_medicos=centros_medicos,
                             descargar_pdf=descargar_pdf)
    finally:
        conn.close()

# ========== DESCARGAR PLANTILLA EXCEL ==========
@app.route('/facturacion/descargar-plantilla-excel')
@login_required
def descargar_plantilla_excel():
    """Generar y descargar plantilla Excel para carga masiva"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        from io import BytesIO
        
        # Crear un nuevo workbook
        wb = Workbook()
        
        # Hoja 1: Plantilla de Pacientes
        ws_pacientes = wb.active
        ws_pacientes.title = "Pacientes"
        
        # Encabezados
        headers = ['NSS', 'NOMBRE', 'FECHA', 'AUTORIZACIÓN', 'SERVICIO', 'MONTO']
        ws_pacientes.append(headers)
        
        # Estilo para encabezados
        header_fill = PatternFill(start_color="CEB0B7", end_color="CEB0B7", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Aplicar estilo y protección a los encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws_pacientes.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.protection = Protection(locked=True)  # Proteger encabezado
        
        # Ajustar ancho de columnas
        column_widths = {'A': 15, 'B': 30, 'C': 12, 'D': 15, 'E': 30, 'F': 12}
        for col, width in column_widths.items():
            ws_pacientes.column_dimensions[col].width = width
        
        # Hoja 2: Lista de Servicios
        ws_servicios = wb.create_sheet("Servicios Disponibles")
        ws_servicios.append(['SERVICIOS DISPONIBLES', 'PRECIO BASE'])
        
        # Encabezados hoja servicios
        for col_num in range(1, 3):
            cell = ws_servicios.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Obtener servicios de la base de datos
        conn = get_db_connection()
        servicios = conn.execute('SELECT descripcion, precio_base FROM tipos_servicios WHERE activo = 1 ORDER BY descripcion').fetchall()
        conn.close()
        
        # Validar que servicios no sea None
        if servicios is None:
            servicios = []
        
        for servicio in servicios:
            ws_servicios.append([servicio['descripcion'], f"{servicio['precio_base']:,.2f}" if servicio['precio_base'] else ''])
        
        # Ajustar ancho de columnas en hoja servicios
        ws_servicios.column_dimensions['A'].width = 40
        ws_servicios.column_dimensions['B'].width = 15
        
        # Crear lista desplegable para la columna SERVICIO (columna E)
        # Referencia: desde la fila 2 hasta la última fila con servicios en la hoja "Servicios Disponibles"
        num_servicios = len(servicios) if servicios else 0
        if num_servicios > 0:
            # Crear la validación de datos
            dv = DataValidation(
                type="list",
                formula1=f"'Servicios Disponibles'!$A$2:$A${num_servicios + 1}",
                allow_blank=False
            )
            dv.error = 'Seleccione un servicio de la lista'
            dv.errorTitle = 'Servicio Inválido'
            dv.prompt = 'Por favor seleccione un servicio de la lista desplegable'
            dv.promptTitle = 'Seleccionar Servicio'
            
            # Aplicar la validación a la columna E (SERVICIO) desde la fila 2 hasta la 1000
            ws_pacientes.add_data_validation(dv)
            dv.add('E2:E1000')
        
        # Agregar filas vacías para poder aplicar protección
        # Agregamos al menos 100 filas vacías para que el usuario pueda trabajar
        for i in range(2, 102):  # Filas 2 a 101
            ws_pacientes.append(['', '', '', '', '', ''])
        
        # Desbloquear las celdas de datos (filas 2 en adelante) y alinear a la izquierda
        alineacion_izquierda = Alignment(horizontal='left', vertical='center')
        for row in range(2, 102):
            for col in range(1, 7):  # Columnas A-F (1-6)
                cell = ws_pacientes.cell(row=row, column=col)
                cell.protection = Protection(locked=False)
                cell.alignment = alineacion_izquierda
        
        # Proteger la hoja de Pacientes (solo el encabezado quedará bloqueado)
        ws_pacientes.protection.sheet = True
        ws_pacientes.protection.password = ''  # Sin contraseña para facilitar uso
        ws_pacientes.protection.enable()
        
        # Hoja 3: Instrucciones
        ws_instrucciones = wb.create_sheet("Instrucciones")
        instrucciones = [
            ['INSTRUCCIONES PARA CARGAR PACIENTES'],
            [''],
            ['1. Complete la hoja "Pacientes" con los datos de los pacientes'],
            ['2. NSS: Solo números y guiones (ej: 001-234-5678)'],
            ['3. NOMBRE: Nombre completo del paciente'],
            ['4. FECHA: Formato AAAA-MM-DD (ej: 2025-10-16)'],
            ['5. AUTORIZACIÓN: Solo números, debe ser única para cada paciente'],
            ['6. SERVICIO: Seleccione de la lista desplegable (se alimenta de la hoja "Servicios Disponibles")'],
            ['7. MONTO: Cantidad en pesos (solo números)'],
            [''],
            ['IMPORTANTE:'],
            ['- Los encabezados están protegidos y NO se pueden modificar'],
            ['- La columna SERVICIO tiene lista desplegable - haga clic en la flecha para seleccionar'],
            ['- Cada autorización debe ser única'],
            ['- Complete directamente desde la fila 2'],
            [''],
            ['NOTA: Antes de cargar, debe seleccionar el Médico y ARS en la página web']
        ]
        
        for row in instrucciones:
            ws_instrucciones.append(row)
        
        # Estilo para título de instrucciones
        ws_instrucciones.cell(row=1, column=1).font = Font(bold=True, size=14, color="CEB0B7")
        ws_instrucciones.cell(row=11, column=1).font = Font(bold=True, size=12, color="B89BA3")
        
        ws_instrucciones.column_dimensions['A'].width = 70
        
        # Guardar en memoria
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'plantilla_pacientes_{datetime.now().strftime("%Y%m%d")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        import traceback
        error_detallado = traceback.format_exc()
        print(f"❌ ERROR AL GENERAR PLANTILLA EXCEL:\n{error_detallado}")
        flash(f'Error al generar plantilla: {str(e)}', 'error')
        return redirect(url_for('facturacion_facturas_nueva'))

# ========== PROCESAR CARGA DESDE EXCEL ==========
@app.route('/facturacion/procesar-excel', methods=['POST'])
@login_required
def procesar_excel():
    """Procesar archivo Excel cargado y retornar JSON con los datos"""
    try:
        from openpyxl import load_workbook
        
        # Verificar que se subió un archivo
        if 'archivo_excel' not in request.files:
            return jsonify({'error': 'No se subió ningún archivo'}), 400
        
        file = request.files['archivo_excel']
        
        if file.filename == '':
            return jsonify({'error': 'No se seleccionó ningún archivo'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'El archivo debe ser formato Excel (.xlsx o .xls)'}), 400
        
        # Leer el archivo Excel
        wb = load_workbook(file, data_only=True)
        ws = wb['Pacientes']
        
        pacientes = []
        errores = []
        autorizaciones_usadas = set()
        total_filas_con_datos = 0
        
        # Leer filas (saltar encabezado)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Saltar filas vacías
            if not any(row):
                continue

            total_filas_con_datos += 1
            
            nss_raw, nombre_raw, fecha_raw, autorizacion_raw, servicio_raw, monto_raw = row[0], row[1], row[2], row[3], row[4], row[5]
            
            # ========== VALIDACIÓN 1: Campos Obligatorios ==========
            if not nss_raw or not nombre_raw:
                errores.append(f'❌ Fila {row_num}: NSS y NOMBRE son obligatorios')
                continue
            
            # ========== VALIDACIÓN 2: NSS (Solo números y guiones) ==========
            import re
            # Limpiar NSS: si Excel lo convirtió a float (ej: 136605324.0), remover decimales
            try:
                # Si es un número, convertir a int para eliminar decimales espurios
                nss_temp = float(nss_raw)
                if nss_temp.is_integer():
                    nss = str(int(nss_temp)).strip()
                else:
                    nss = str(nss_raw).strip()
            except (ValueError, TypeError):
                # Si no es número, usar como string (puede tener guiones)
                nss = str(nss_raw).strip()
            
            if not re.match(r'^[0-9\-]+$', nss):
                errores.append(f'❌ Fila {row_num}: NSS "{nss}" solo debe contener números y guiones')
                continue
            
            # ========== VALIDACIÓN 3: NOMBRE (Sin números) ==========
            nombre = str(nombre_raw).strip().upper()
            if not nombre:
                errores.append(f'❌ Fila {row_num}: NOMBRE no puede estar vacío')
                continue
            
            # ========== VALIDACIÓN 4: FECHA (Formato válido) ==========
            fecha = ''
            if fecha_raw:
                try:
                    if hasattr(fecha_raw, 'strftime'):
                        fecha = fecha_raw.strftime('%Y-%m-%d')
                    else:
                        fecha_str = str(fecha_raw).strip()
                        # Intentar parsear varios formatos
                        from datetime import datetime
                        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                            try:
                                fecha_obj = datetime.strptime(fecha_str, fmt)
                                fecha = fecha_obj.strftime('%Y-%m-%d')
                                break
                            except:
                                continue
                        if not fecha:
                            errores.append(f'⚠️ Fila {row_num}: FECHA "{fecha_str}" formato inválido (use AAAA-MM-DD)')
                            fecha = datetime.now().strftime('%Y-%m-%d')  # Usar fecha actual como fallback
                except:
                    errores.append(f'⚠️ Fila {row_num}: Error al procesar FECHA, usando fecha actual')
                    from datetime import datetime
                    fecha = datetime.now().strftime('%Y-%m-%d')
            else:
                from datetime import datetime
                fecha = datetime.now().strftime('%Y-%m-%d')
            
            # ========== VALIDACIÓN 5: AUTORIZACIÓN (Alfanumérico y única) ==========
            autorizacion = ''
            if autorizacion_raw:
                # Convertir a string y limpiar (eliminar espacios en blanco)
                autorizacion = str(autorizacion_raw).strip().upper()
                
                # Si viene como número decimal (ej: 123.0), convertir a entero primero
                try:
                    if '.' in autorizacion and autorizacion.replace('.', '').replace('-', '').isdigit():
                        autorizacion = str(int(float(autorizacion_raw)))
                except:
                    pass
                
                # Validar que no esté vacío después de limpiar
                if not autorizacion:
                    errores.append(f'❌ Fila {row_num}: AUTORIZACIÓN no puede estar vacía')
                    continue
                
                # Validar longitud máxima
                if len(autorizacion) > 50:
                    errores.append(f'❌ Fila {row_num}: AUTORIZACIÓN "{autorizacion}" supera los 50 caracteres')
                    continue
                
                # Validar que sea única
                if autorizacion in autorizaciones_usadas:
                    errores.append(f'❌ Fila {row_num}: AUTORIZACIÓN "{autorizacion}" está duplicada en el Excel')
                    continue
                
                autorizaciones_usadas.add(autorizacion)
            
            # ========== VALIDACIÓN 6: SERVICIO (Sin números) ==========
            servicio = str(servicio_raw).strip().upper() if servicio_raw else 'CONSULTA'
            
            # Verificar que no contenga números
            if any(char.isdigit() for char in servicio):
                errores.append(f'❌ Fila {row_num}: SERVICIO "{servicio}" no debe contener números')
                continue
            
            if not servicio:
                servicio = 'CONSULTA'
            
            # ========== VALIDACIÓN 7: MONTO (Debe ser numérico positivo) ==========
            monto = 0
            if monto_raw:
                try:
                    monto = float(monto_raw)
                    if monto < 0:
                        errores.append(f'⚠️ Fila {row_num}: MONTO no puede ser negativo, usando 0')
                        monto = 0
                except:
                    errores.append(f'❌ Fila {row_num}: MONTO "{monto_raw}" no es un número válido')
                    continue
            
            # Si llegó hasta aquí, la fila es válida
            pacientes.append({
                'nss': nss,
                'nombre': nombre,
                'fecha': fecha,
                'autorizacion': autorizacion,
                'servicio': servicio,
                'monto': monto,
                'fila': row_num
            })

        if not pacientes:
            # Si no hay ninguna fila válida, devolver error (no tiene sentido continuar)
            return jsonify({
                'error': 'El archivo no contiene datos válidos',
                'errores': errores,
                'total_errores': len(errores),
                'total_excel': total_filas_con_datos
            }), 400
        
        return jsonify({
            'success': True,
            'pacientes': pacientes,
            # Compatibilidad (existente)
            'total': len(pacientes),
            # Nuevos campos para UX/alertas
            'total_validos': len(pacientes),
            'total_excel': total_filas_con_datos,
            'total_omitidos': max(total_filas_con_datos - len(pacientes), 0),
            'errores': errores,
            'total_errores': len(errores)
        })
        
    except Exception as e:
        return jsonify({'error': f'Error al procesar Excel: {str(e)}'}), 500

# ========== GENERAR FACTURAS ==========
# PÁGINA DESACTIVADA - Se usa /facturacion/historico en su lugar
# @app.route('/facturacion/facturas')
# @login_required
# def facturacion_facturas():
#     """Lista de Facturas Generadas"""
#     conn = get_db_connection()
#     facturas = conn.execute('''
#         SELECT f.*, m.nombre as medico_nombre, a.nombre_ars,
#                (SELECT COUNT(*) FROM facturas_detalle WHERE factura_id = f.id) as num_pacientes
#         FROM facturas f
#         JOIN medicos m ON f.medico_id = m.id
#         JOIN ars a ON f.ars_id = a.id
#         WHERE f.activo = 1
#         ORDER BY f.created_at DESC
#     ''').fetchall()
#     conn.close()
#     return render_template('facturacion/facturas.html', facturas=facturas)

@app.route('/facturacion/generar-factura', methods=['GET', 'POST'])
@login_required
def facturacion_generar():
    """Generar factura desde pacientes pendientes - PASO A PASO"""
    conn = get_db_connection()
    
    if request.method == 'POST':
        # Verificar si es el paso de configuración o el paso final
        step = request.form.get('step', '1')
        
        if step == '1':
            # PASO 1: Configurar parámetros iniciales
            ars_id = request.form.get('ars_id', type=int)
            ncf_id = request.form.get('ncf_id', type=int)
            medico_factura_id = request.form.get('medico_factura_id', type=int)
            fecha_factura = request.form.get('fecha_factura')
            
            if not ars_id or not ncf_id or not medico_factura_id or not fecha_factura:
                flash('Todos los campos son obligatorios', 'error')
                return redirect(url_for('facturacion_generar'))
            
            # Verificar que el médico esté habilitado para facturar
            medico_habilitado = conn.execute('SELECT * FROM medicos WHERE id = %s AND activo = 1 AND factura = 1', 
                                             (medico_factura_id,)).fetchone()
            if not medico_habilitado:
                flash('El médico seleccionado no está habilitado para facturar', 'error')
                return redirect(url_for('facturacion_generar'))
            
            # Obtener TODOS los pacientes pendientes de esta ARS
            # Usar LEFT JOIN para medicos en caso de que medico_consulta sea NULL o inválido
            pendientes = conn.execute('''
                SELECT fd.*, 
                       COALESCE(m.nombre, 'Sin médico asignado') as medico_nombre, 
                       a.nombre_ars, 
                       COALESCE(p.nombre, fd.nombre_paciente) as paciente_nombre_completo
                FROM facturas_detalle fd
                LEFT JOIN medicos m ON fd.medico_consulta = m.id
                JOIN ars a ON fd.ars_id = a.id
                LEFT JOIN pacientes p ON fd.paciente_id = p.id
                WHERE fd.estado = 'pendiente' AND fd.ars_id = %s
                ORDER BY fd.fecha_servicio DESC
            ''', (ars_id,)).fetchall()
            
            # DEBUG: Registrar cuántos pacientes se encontraron
            print(f"\n{'='*60}")
            print(f"DEBUG - PACIENTES PENDIENTES ENCONTRADOS")
            print(f"{'='*60}")
            print(f"ARS ID: {ars_id}")
            print(f"Total pacientes: {len(pendientes)}")
            print(f"IDs de pacientes: {[p['id'] for p in pendientes]}")
            if pendientes:
                print(f"Nombres: {[p['nombre_paciente'] for p in pendientes]}")
            print(f"{'='*60}\n")
            
            # Obtener info de ARS, NCF y Médico para facturar
            ars = conn.execute('SELECT * FROM ars WHERE id = %s', (ars_id,)).fetchone()
            ncf = conn.execute('SELECT * FROM ncf WHERE id = %s', (ncf_id,)).fetchone()
            
            # Obtener lista de médicos únicos en los pendientes (para filtro visual)
            # Usar LEFT JOIN para incluir registros sin médico asignado
            medicos = conn.execute('''
                SELECT DISTINCT m.id, m.nombre 
                FROM facturas_detalle fd
                LEFT JOIN medicos m ON fd.medico_consulta = m.id
                WHERE fd.estado = 'pendiente' AND fd.ars_id = %s AND m.id IS NOT NULL
                ORDER BY m.nombre
            ''', (ars_id,)).fetchall()
            
            conn.close()
            
            return render_template('facturacion/generar_factura_step2.html',
                                 pendientes=pendientes,
                                 ars=ars,
                                 ncf=ncf,
                                 fecha_factura=fecha_factura,
                                 medicos=medicos,
                                 medico_factura_id=medico_factura_id,
                                 medico_factura_nombre=medico_habilitado['nombre'])
        
        # PASO 2: Mostrar vista previa de la factura
        if step == '2':
            import json
            pacientes_ids = request.form.get('pacientes_ids')
            ids_list = json.loads(pacientes_ids) if pacientes_ids else []
            
            # Obtener parámetros del formulario
            ars_id = request.form.get('ars_id', type=int)
            ncf_id = request.form.get('ncf_id', type=int)
            medico_factura_id = request.form.get('medico_factura_id', type=int)
            fecha_factura = request.form.get('fecha_factura')
            
            if not ids_list:
                flash('Debe seleccionar al menos un paciente', 'error')
                return redirect(url_for('facturacion_generar'))
            
            if not ars_id or not ncf_id or not medico_factura_id or not fecha_factura:
                flash('Faltan parámetros de configuración', 'error')
                return redirect(url_for('facturacion_generar'))
            
            # Obtener datos de los pacientes seleccionados (SIN datos del médico)
            placeholders = ','.join(['%s' for _ in ids_list])
            pacientes = conn.execute(f'''
                SELECT fd.*
                FROM facturas_detalle fd
                WHERE fd.id IN ({placeholders}) AND fd.estado = 'pendiente' AND fd.ars_id = %s
            ''', ids_list + [ars_id]).fetchall()
            
            if not pacientes:
                flash('No se encontraron pacientes pendientes válidos', 'error')
                return redirect(url_for('facturacion_generar'))

            # VALIDACIÓN: La fecha de la factura debe ser >= fecha de consulta más reciente
            try:
                from datetime import date, datetime

                fecha_factura_date = date.fromisoformat(str(fecha_factura)[:10])
                fechas_pacientes = []
                for p in pacientes:
                    fs = p.get('fecha_servicio')
                    if not fs:
                        continue
                    if hasattr(fs, 'date'):
                        # datetime -> date, date -> date
                        fs_date = fs.date() if hasattr(fs, 'hour') else fs
                        fechas_pacientes.append(fs_date)
                    elif isinstance(fs, str):
                        fs_str = fs.strip()
                        # intentar varios formatos comunes
                        for fmt in ['%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y']:
                            try:
                                fechas_pacientes.append(datetime.strptime(fs_str, fmt).date())
                                break
                            except Exception:
                                continue

                if fechas_pacientes:
                    max_fecha_paciente = max(fechas_pacientes)
                    if fecha_factura_date < max_fecha_paciente:
                        flash(
                            f'❌ La fecha de la factura ({fecha_factura_date.strftime("%d/%m/%Y")}) debe ser mayor o igual a '
                            f'la fecha de consulta más reciente ({max_fecha_paciente.strftime("%d/%m/%Y")}).',
                            'error'
                        )

                        # Re-render PASO 2 manteniendo selección
                        ars = conn.execute('SELECT * FROM ars WHERE id = %s', (ars_id,)).fetchone()
                        ncf = conn.execute('SELECT * FROM ncf WHERE id = %s', (ncf_id,)).fetchone()
                        medico = conn.execute('SELECT * FROM medicos WHERE id = %s AND activo = 1 AND factura = 1', (medico_factura_id,)).fetchone()

                        pendientes = conn.execute('''
                            SELECT fd.*, 
                                   COALESCE(m.nombre, 'Sin médico asignado') as medico_nombre, 
                                   a.nombre_ars, 
                                   COALESCE(p.nombre, fd.nombre_paciente) as paciente_nombre_completo
                            FROM facturas_detalle fd
                            LEFT JOIN medicos m ON fd.medico_consulta = m.id
                            JOIN ars a ON fd.ars_id = a.id
                            LEFT JOIN pacientes p ON fd.paciente_id = p.id
                            WHERE fd.estado = 'pendiente' AND fd.ars_id = %s
                            ORDER BY fd.fecha_servicio DESC
                        ''', (ars_id,)).fetchall()

                        medicos = conn.execute('''
                            SELECT DISTINCT m.id, m.nombre 
                            FROM facturas_detalle fd
                            LEFT JOIN medicos m ON fd.medico_consulta = m.id
                            WHERE fd.estado = 'pendiente' AND fd.ars_id = %s AND m.id IS NOT NULL
                            ORDER BY m.nombre
                        ''', (ars_id,)).fetchall()

                        conn.close()

                        return render_template(
                            'facturacion/generar_factura_step2.html',
                            pendientes=pendientes,
                            ars=ars,
                            ncf=ncf,
                            fecha_factura=fecha_factura,
                            medicos=medicos,
                            medico_factura_id=medico_factura_id,
                            medico_factura_nombre=(medico['nombre'] if medico else ''),
                            selected_ids=ids_list
                        )
            except Exception:
                # Si falla el parseo, no bloquear (se valida más adelante)
                pass
            
            # Obtener info del MÉDICO QUE FACTURA
            medico = conn.execute('''
                SELECT * FROM medicos WHERE id = %s AND activo = 1 AND factura = 1
            ''', (medico_factura_id,)).fetchone()
            
            if not medico:
                flash('El médico seleccionado no está habilitado para facturar', 'error')
                return redirect(url_for('facturacion_generar'))
            
            # Obtener info de ARS y NCF
            ars = conn.execute('SELECT * FROM ars WHERE id = %s', (ars_id,)).fetchone()
            ncf = conn.execute('SELECT * FROM ncf WHERE id = %s', (ncf_id,)).fetchone()
            
            # Calcular el próximo NCF
            proximo_numero = ncf['ultimo_numero'] + 1
            ncf_completo = f"{ncf['prefijo']}{str(proximo_numero).zfill(ncf['tamaño'])}"
            
            # Calcular total
            total = sum(p['monto'] for p in pacientes)
            
            # ITBIS = 0 (sin impuesto)
            itbis = 0
            total_final = total
            
            # Obtener centro médico por defecto del médico
            centro_medico = conn.execute('''
                SELECT c.nombre, c.direccion
                FROM medico_centro mc
                JOIN centros_medicos c ON mc.centro_id = c.id
                WHERE mc.medico_id = %s AND mc.es_defecto = 1 AND mc.activo = 1 AND c.activo = 1
                LIMIT 1
            ''', (medico_factura_id,)).fetchone()
            
            # Si no hay centro por defecto, buscar el primero disponible
            if not centro_medico:
                centro_medico = conn.execute('''
                    SELECT c.nombre, c.direccion
                    FROM medico_centro mc
                    JOIN centros_medicos c ON mc.centro_id = c.id
                    WHERE mc.medico_id = %s AND mc.activo = 1 AND c.activo = 1
                    LIMIT 1
                ''', (medico_factura_id,)).fetchone()
            
            # Si aún no hay centro, usar valores por defecto
            if not centro_medico:
                centro_medico = {
                    'nombre': 'Centro Oriental de Ginecología y Obstetricia',
                    'direccion': 'Zona Oriental, República Dominicana'
                }
            
            conn.close()
            
            # Mostrar vista previa con los datos del médico que factura
            return render_template('facturacion/vista_previa_factura.html',
                                 pacientes=pacientes,
                                 medico=medico,
                                 ars=ars,
                                 ncf=ncf,
                                 ncf_completo=ncf_completo,
                                 fecha_factura=fecha_factura,
                                 subtotal=total,
                                 itbis=itbis,
                                 total=total_final,
                                 pacientes_ids=json.dumps(ids_list),
                                 ars_id=ars_id,
                                 ncf_id=ncf_id,
                                 medico_factura_id=medico_factura_id,
                                 centro_medico=centro_medico)
    
    # GET: PASO 1 - Mostrar formulario para seleccionar ARS, NCF, Médico y fecha
    from datetime import date
    
    # Detectar si viene de dashboard (auto-fill)
    ars_id_prefill = request.args.get('ars_id', type=int)
    auto_mode = request.args.get('auto') == '1'
    
    # Variables para pre-llenar
    medico_id_prefill = None
    ncf_id_prefill = None
    mensaje_auto = None
    
    # Si viene de dashboard, pre-llenar datos
    if auto_mode and ars_id_prefill:
        # 1. Buscar médico asociado al usuario logueado
        medico_usuario = conn.execute(
            'SELECT id FROM medicos WHERE email = %s AND activo = 1 AND factura = 1',
            (current_user.email,)
        ).fetchone()
        
        if medico_usuario:
            medico_id_prefill = medico_usuario['id']
        else:
            # Si no hay médico asociado, usar el primero disponible (Opción B)
            primer_medico = conn.execute(
                'SELECT id FROM medicos WHERE activo = 1 AND factura = 1 ORDER BY nombre LIMIT 1'
            ).fetchone()
            if primer_medico:
                medico_id_prefill = primer_medico['id']
        
        # 2. Detectar tipo de NCF según ARS
        ars_info = conn.execute('SELECT nombre_ars FROM ars WHERE id = %s', (ars_id_prefill,)).fetchone()
        
        if ars_info:
            # Si es SENASA → Crédito Gubernamental
            if 'SENASA' in ars_info['nombre_ars'].upper():
                ncf_gubernamental = conn.execute(
                    "SELECT id FROM ncf WHERE tipo LIKE '%Gubernamental%' AND activo = 1 ORDER BY id LIMIT 1"
                ).fetchone()
                if ncf_gubernamental:
                    ncf_id_prefill = ncf_gubernamental['id']
            else:
                # Otras ARS → Crédito Fiscal
                ncf_fiscal = conn.execute(
                    "SELECT id FROM ncf WHERE tipo LIKE '%Fiscal%' AND activo = 1 ORDER BY id LIMIT 1"
                ).fetchone()
                if ncf_fiscal:
                    ncf_id_prefill = ncf_fiscal['id']
            
            mensaje_auto = f"✅ Formulario pre-llenado para facturar: {ars_info['nombre_ars']}"
    
    # Obtener todas las ARS activas
    ars_list = conn.execute('''
        SELECT * FROM ars WHERE activo = 1 ORDER BY nombre_ars
    ''').fetchall()
    
    # Obtener todos los NCF activos
    ncf_list = conn.execute('''
        SELECT * FROM ncf WHERE activo = 1 ORDER BY tipo, prefijo
    ''').fetchall()
    
    # Obtener médicos habilitados para facturar
    medicos_habilitados = conn.execute('''
        SELECT * FROM medicos WHERE activo = 1 AND factura = 1 ORDER BY nombre
    ''').fetchall()
    
    conn.close()
    
    # Fecha actual por defecto (ajustada a zona horaria RD)
    fecha_hoy_rd = obtener_fecha_rd()
    fecha_actual_sql = fecha_hoy_rd.strftime('%Y-%m-%d')  # Para backend
    fecha_actual_visual = fecha_hoy_rd.strftime('%d/%m/%Y')  # Para mostrar dd/mm/yyyy
    
    return render_template('facturacion/generar_factura.html', 
                         ars_list=ars_list, 
                         ncf_list=ncf_list,
                         medicos_habilitados=medicos_habilitados,
                         fecha_actual=fecha_actual_sql,
                         fecha_actual_visual=fecha_actual_visual,
                         ars_id_prefill=ars_id_prefill,
                         medico_id_prefill=medico_id_prefill,
                         ncf_id_prefill=ncf_id_prefill,
                         auto_mode=auto_mode,
                         mensaje_auto=mensaje_auto)

@app.route('/facturacion/generar-factura-final', methods=['POST'])
@login_required
def facturacion_generar_final():
    """PASO 3: Generar factura definitiva, guardar, descargar y enviar"""
    print("🚀 INICIANDO GENERACIÓN DE FACTURA FINAL")
    conn = get_db_connection()
    
    try:
        import json
        from io import BytesIO
        
        pacientes_ids = request.form.get('pacientes_ids')
        print(f"📋 pacientes_ids recibido: {pacientes_ids}")
        ids_list = json.loads(pacientes_ids) if pacientes_ids else []
        print(f"📋 ids_list parseado: {ids_list}")
        
        # Obtener parámetros del formulario
        ars_id = request.form.get('ars_id', type=int)
        ncf_id = request.form.get('ncf_id', type=int)
        medico_factura_id = request.form.get('medico_factura_id', type=int)
        fecha_factura = request.form.get('fecha_factura')
        
        if not ids_list or not ars_id or not ncf_id or not medico_factura_id or not fecha_factura:
            error_msg = f'Faltan parámetros: ids_list={bool(ids_list)}, ars_id={ars_id}, ncf_id={ncf_id}, medico={medico_factura_id}, fecha={fecha_factura}'
            flash(error_msg, 'error')
            print(f"❌ ERROR: {error_msg}")  # Para debug
            return redirect(url_for('facturacion_generar'))
        
        # Obtener datos del MÉDICO QUE FACTURA
        medico = conn.execute('''
            SELECT * FROM medicos WHERE id = %s AND activo = 1 AND factura = 1
        ''', (medico_factura_id,)).fetchone()
        
        if not medico:
            flash('El médico seleccionado no está habilitado para facturar', 'error')
            return redirect(url_for('facturacion_generar'))
        
        medico_id = medico['id']
        medico_email = medico['email']
        
        # Obtener datos de los pacientes seleccionados
        placeholders = ','.join(['%s' for _ in ids_list])
        pacientes = conn.execute(f'''
            SELECT fd.*
            FROM facturas_detalle fd
            WHERE fd.id IN ({placeholders}) AND fd.estado = 'pendiente' AND fd.ars_id = %s
        ''', ids_list + [ars_id]).fetchall()
        
        if not pacientes:
            flash('No se encontraron pacientes pendientes válidos', 'error')
            return redirect(url_for('facturacion_generar'))

        # VALIDACIÓN (seguridad): Fecha factura >= fecha paciente más reciente
        try:
            from datetime import date, datetime
            fecha_factura_date = date.fromisoformat(str(fecha_factura)[:10])
            fechas_pacientes = []
            for p in pacientes:
                fs = p.get('fecha_servicio')
                if not fs:
                    continue
                if hasattr(fs, 'date'):
                    fs_date = fs.date() if hasattr(fs, 'hour') else fs
                    fechas_pacientes.append(fs_date)
                elif isinstance(fs, str):
                    fs_str = fs.strip()
                    for fmt in ['%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y']:
                        try:
                            fechas_pacientes.append(datetime.strptime(fs_str, fmt).date())
                            break
                        except Exception:
                            continue
            if fechas_pacientes:
                max_fecha_paciente = max(fechas_pacientes)
                if fecha_factura_date < max_fecha_paciente:
                    flash(
                        f'❌ No se puede generar la factura. La fecha de la factura ({fecha_factura_date.strftime("%d/%m/%Y")}) '
                        f'debe ser mayor o igual a la fecha de consulta más reciente ({max_fecha_paciente.strftime("%d/%m/%Y")}).',
                        'error'
                    )
                    return redirect(url_for('facturacion_generar', ars_id=ars_id))
        except Exception:
            pass
        
        # Calcular total
        total = sum(p['monto'] for p in pacientes)
        
        # Obtener y actualizar NCF
        ncf = conn.execute('SELECT * FROM ncf WHERE id = %s', (ncf_id,)).fetchone()
        proximo_numero = ncf['ultimo_numero'] + 1
        ncf_completo = f"{ncf['prefijo']}{str(proximo_numero).zfill(ncf['tamaño'])}"
        
        # Crear factura
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO facturas (fecha_factura, medico_id, ars_id, ncf_id, ncf_numero, total)
            VALUES (%s, %s, %s, %s, %s, %s)
        ''', (fecha_factura, medico_id, ars_id, ncf_id, ncf_completo, total))
        
        factura_id = cursor.lastrowid
        
        # Actualizar el último número del NCF
        cursor.execute('''
            UPDATE ncf SET ultimo_numero = %s WHERE id = %s
        ''', (proximo_numero, ncf_id))
        
        # Actualizar pacientes a facturados y asignar el médico que factura
        cursor.execute(f'''
            UPDATE facturas_detalle 
            SET factura_id = %s, estado = 'facturado', medico_id = %s
            WHERE id IN ({placeholders})
        ''', [factura_id, medico_id] + ids_list)
        
        conn.commit()
        
        # Generar PDF si ReportLab está disponible
        if REPORTLAB_AVAILABLE:
            # Obtener datos completos para el PDF (usando el médico que factura)
            pacientes_pdf = conn.execute(f'''
                SELECT fd.*, 
                       m.nombre as medico_nombre, 
                       m.especialidad as medico_especialidad,
                       m.cedula as medico_cedula,
                       m.exequatur as medico_exequatur,
                       a.nombre_ars, a.rnc as ars_rnc,
                       ca.codigo_ars as codigo_ars
                FROM facturas_detalle fd
                JOIN medicos m ON m.id = %s
                JOIN ars a ON fd.ars_id = a.id
                LEFT JOIN codigo_ars ca ON ca.medico_id = m.id AND ca.ars_id = a.id
                WHERE fd.id IN ({placeholders})
            ''', [medico_id] + ids_list).fetchall()
            
            # Obtener datos del NCF
            ncf_data = conn.execute('SELECT fecha_fin, tipo FROM ncf WHERE id = %s', (ncf_id,)).fetchone()
            
            # Obtener centro médico por defecto del médico
            centro_defecto = conn.execute('''
                SELECT c.nombre, c.direccion
                FROM medico_centro mc
                JOIN centros_medicos c ON mc.centro_id = c.id
                WHERE mc.medico_id = %s AND mc.es_defecto = 1 AND mc.activo = 1 AND c.activo = 1
                LIMIT 1
            ''', (medico_factura_id,)).fetchone()
            
            # Si no hay centro por defecto, buscar el primero disponible
            if not centro_defecto:
                centro_defecto = conn.execute('''
                    SELECT c.nombre, c.direccion
                    FROM medico_centro mc
                    JOIN centros_medicos c ON mc.centro_id = c.id
                    WHERE mc.medico_id = %s AND mc.activo = 1 AND c.activo = 1
                    LIMIT 1
                ''', (medico_factura_id,)).fetchone()
            
            # Generar PDF
            pdf_buffer = generar_pdf_factura(factura_id, ncf_completo, fecha_factura, pacientes_pdf, total, ncf_data, centro_defecto)
            
            # Enviar por email si hay email del médico (de manera asíncrona)
            if medico_email:
                # Crear copia del buffer para el thread
                email_pdf_buffer = BytesIO(pdf_buffer.getvalue())
                threading.Thread(target=enviar_email_factura, args=(medico_email, factura_id, ncf_completo, email_pdf_buffer, total)).start()
            
            conn.close()
            
            # En lugar de descargar directamente, redirigir a ver factura con parámetro para mostrar modal de impresión
            flash(f'✅ Factura #{factura_id} generada exitosamente con NCF {ncf_completo} | {len(ids_list)} paciente(s) facturado(s)', 'success')
            return redirect(url_for('facturacion_ver_factura', factura_id=factura_id, mostrar_modal_imprimir=1))
        else:
            conn.close()
            flash(f'✅ Factura #{factura_id} generada exitosamente con NCF {ncf_completo} | {len(ids_list)} paciente(s) facturado(s)', 'success')
            return redirect(url_for('facturacion_historico'))
        
    except Exception as e:
        import traceback
        error_detallado = traceback.format_exc()
        print(f"❌ ERROR AL GENERAR FACTURA:\n{error_detallado}")
        if conn:
            conn.rollback()
            conn.close()
        flash(f'Error al generar factura: {str(e)}', 'error')
        return redirect(url_for('facturacion_generar'))

def generar_pdf_factura(factura_id, ncf, fecha, pacientes, total, ncf_data=None, centro_medico=None):
    """Generar PDF de la factura con formato actualizado"""
    from io import BytesIO
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=25, bottomMargin=70)
    elements = []
    styles = getSampleStyleSheet()
    
    # Definir constante de paginación
    REGISTROS_POR_PAGINA = 20
    total_pacientes = len(pacientes)
    total_paginas_estimadas = (total_pacientes + REGISTROS_POR_PAGINA - 1) // REGISTROS_POR_PAGINA if total_pacientes > 0 else 1
    
    # Preparar datos del footer para el callback
    footer_data = {}
    if pacientes and len(pacientes) > 0:
        footer_data['medico_nombre'] = pacientes[0]['medico_nombre'] if 'medico_nombre' in pacientes[0].keys() else 'N/A'
        footer_data['medico_especialidad'] = pacientes[0]['medico_especialidad'] if 'medico_especialidad' in pacientes[0].keys() and pacientes[0]['medico_especialidad'] else ''
        footer_data['medico_cedula'] = pacientes[0]['medico_cedula'] if 'medico_cedula' in pacientes[0].keys() and pacientes[0]['medico_cedula'] else ''
        footer_data['medico_exequatur'] = pacientes[0]['medico_exequatur'] if 'medico_exequatur' in pacientes[0].keys() and pacientes[0]['medico_exequatur'] else ''
    
    # Agregar información del centro médico
    if centro_medico:
        footer_data['centro_nombre'] = centro_medico['nombre']
        footer_data['centro_direccion'] = centro_medico['direccion']
    else:
        # Fallback al texto original
        footer_data['centro_nombre'] = 'Centro Oriental de Ginecología y Obstetricia'
        footer_data['centro_direccion'] = 'Zona Oriental, República Dominicana'
    
    # Función para dibujar el footer solo en la última página
    def agregar_footer(canvas, doc):
        # Solo dibujar en la última página
        if doc.page != total_paginas_estimadas:
            return
        
        canvas.saveState()
        
        # Línea de firma a la izquierda (arriba del footer)
        firma_y = 110  # Posición desde el fondo para la firma
        firma_x_start = 60  # Margen izquierdo
        firma_ancho = 200  # Ancho de la línea de firma
        
        # Dibujar línea para firma
        canvas.setStrokeColor(colors.black)
        canvas.setLineWidth(0.5)
        canvas.line(firma_x_start, firma_y, firma_x_start + firma_ancho, firma_y)
        
        # Nombre de la doctora debajo de la línea
        canvas.setFont('Helvetica', 9)
        canvas.setFillColor(colors.black)
        canvas.drawString(firma_x_start, firma_y - 15, "Dra. Shirley Ramírez")
        
        # Footer centrado siempre al final de la página
        footer_y = 50  # 50 puntos desde el fondo
        
        if footer_data:
            # Línea 1: Nombre del médico (negrita)
            canvas.setFont('Helvetica-Bold', 8)
            canvas.setFillColor(colors.grey)
            canvas.drawCentredString(letter[0]/2, footer_y + 20, footer_data.get('medico_nombre', ''))
            
            # Línea 2: Especialidad, Cédula y Exequátur
            canvas.setFont('Helvetica', 8)
            linea2 = footer_data.get('medico_especialidad', '')
            if footer_data.get('medico_cedula'):
                linea2 += f" | Cédula: {footer_data['medico_cedula']}"
            if footer_data.get('medico_exequatur'):
                linea2 += f" | EXEQUATUR: {footer_data['medico_exequatur']}"
            canvas.drawCentredString(letter[0]/2, footer_y + 10, linea2)
            
            # Línea 3: Centro Médico (dinámico)
            centro_texto = f"{footer_data.get('centro_nombre', '')}, {footer_data.get('centro_direccion', '')}"
            canvas.drawCentredString(letter[0]/2, footer_y, centro_texto)
        
        canvas.restoreState()
    
    # Función helper para formatear moneda
    def formato_moneda(valor):
        return "{:,.2f}".format(float(valor))
    
    # Header compacto: Logo a la izquierda, FACTURA en el centro
    logo_path = os.path.join('static', 'logos', 'logo-dra-shirley.png')
    if os.path.exists(logo_path):
        header_data = [[
            Image(logo_path, width=1*inch, height=1*inch),
            Paragraph("<para align='center'><b><font size='20' color='black'>FACTURA</font></b><br/><font size='8' color='black'>CRÉDITO FISCAL</font></para>", styles['Normal']),
            ''
        ]]
        header_table = Table(header_data, colWidths=[1.2*inch, 5.6*inch, 0.5*inch])
        header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 0.15*inch))
    else:
        # Si no hay logo, solo título centrado
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18,
                                     textColor=colors.HexColor('#CEB0B7'), alignment=TA_CENTER, spaceAfter=8)
        elements.append(Paragraph("FACTURA", title_style))
        elements.append(Spacer(1, 0.2*inch))
    
    # Información en 3 columnas (como el HTML)
    if pacientes and len(pacientes) > 0:
        # Extraer datos dinámicos
        ars_nombre = pacientes[0].get('nombre_ars', 'N/A')
        ars_rnc = pacientes[0].get('ars_rnc', 'N/A')
        es_senasa = 'SENASA' in (str(ars_nombre or '')).upper()
        medico_nombre = pacientes[0].get('medico_nombre', 'N/A')
        medico_especialidad = pacientes[0].get('medico_especialidad', 'N/A')
        codigo_ars = pacientes[0].get('codigo_ars', 'N/A')
        medico_cedula = pacientes[0].get('medico_cedula', '')
        medico_exequatur = pacientes[0].get('medico_exequatur', '')
        ncf_tipo = ncf_data.get('tipo', 'CRÉDITO FISCAL') if ncf_data else 'CRÉDITO FISCAL'
        ncf_fecha_fin = ncf_data.get('fecha_fin', '') if ncf_data else ''
        
        # Construir las 3 columnas (sin etiquetas de título, letras más grandes)
        regimen_line = "<br/><b>REGIMEN CONTRIBUTIVO</b>" if es_senasa else ""
        col1_text = f"<font size='10'>Fecha: {formato_fecha_pdf(fecha)}<br/>Cliente: {ars_nombre}{regimen_line}<br/>RNC: {ars_rnc}</font>"
        
        col2_text = f"<b>NCF</b><br/><font size='11' color='#CEB0B7'><b>{ncf}</b></font><br/><font size='10'>Tipo: {ncf_tipo}"
        if ncf_fecha_fin:
            col2_text += f"<br/>Válido hasta: {formato_fecha_pdf(ncf_fecha_fin)}"
        col2_text += "</font>"
        
        col3_text = f"<font size='10'><b>{medico_nombre}</b><br/>{medico_especialidad}<br/>Código: {codigo_ars}"
        if medico_exequatur:
            col3_text += f"<br/>Exequátur: {medico_exequatur}"
        if medico_cedula:
            col3_text += f"<br/>Cédula: {medico_cedula}"
        col3_text += "</font>"
        
        info_data = [[
            Paragraph(col1_text, styles['Normal']),
            Paragraph(col2_text, styles['Normal']),
            Paragraph(col3_text, styles['Normal'])
        ]]
        
        info_table = Table(info_data, colWidths=[2.4*inch, 2.3*inch, 2.6*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CEB0B7')),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.2*inch))
    
    # Tabla de pacientes con paginación (20 registros por página)
    REGISTROS_POR_PAGINA = 20
    total_pacientes = len(pacientes)
    total_paginas = (total_pacientes + REGISTROS_POR_PAGINA - 1) // REGISTROS_POR_PAGINA  # Redondear hacia arriba
    
    for pagina in range(total_paginas):
        inicio = pagina * REGISTROS_POR_PAGINA
        fin = min(inicio + REGISTROS_POR_PAGINA, total_pacientes)
        pacientes_pagina = pacientes[inicio:fin]
        
        # Encabezado de la tabla
        data = [['No.', 'NOMBRES PACIENTE', 'NSS/CONTRATO', 'FECHA', 'AUTORIZACIÓN', 'SERVICIO', 'V/UNITARIO']]
        
        # Agregar pacientes de esta página
        for idx, p in enumerate(pacientes_pagina, inicio + 1):
            data.append([
                str(idx),
                p['nombre_paciente'],
                p['nss'],
                formato_fecha_pdf(p['fecha_servicio']),
                p['autorizacion'],
                p['descripcion_servicio'],
                formato_moneda(p['monto'])
            ])
        
        # Solo agregar totales en la última página
        if pagina == total_paginas - 1:
            subtotal = total
            total_final = subtotal
            
            data.append(['', '', '', '', '', 'SUB-TOTAL:', formato_moneda(subtotal)])
            data.append(['', '', '', '', '', 'ITBIS:', '*E'])
            data.append(['', '', '', '', '', 'TOTAL:', formato_moneda(total_final)])
            
            num_filas_datos = len(data) - 4  # Todas las filas excepto encabezado y 3 filas de totales
        else:
            num_filas_datos = len(data) - 1  # Todas las filas excepto encabezado
        
        table = Table(data, colWidths=[0.5*inch, 2*inch, 1.3*inch, 0.8*inch, 1*inch, 1.5*inch, 1*inch])
        
        # Estilos base de la tabla
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#CEB0B7')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            # Alineaciones específicas por columna
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # No. centrado
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),    # NOMBRES PACIENTE a la izquierda
            ('ALIGN', (2, 1), (2, -1), 'LEFT'),    # NSS/CONTRATO a la izquierda
            ('ALIGN', (3, 1), (3, -1), 'CENTER'),  # FECHA centrado
            ('ALIGN', (4, 1), (4, -1), 'CENTER'),  # AUTORIZACIÓN centrado
            ('ALIGN', (5, 1), (5, -1), 'LEFT'),    # SERVICIO a la izquierda
            ('ALIGN', (6, 1), (6, -1), 'RIGHT'),   # V/UNITARIO a la derecha
        ]
        
        # Agregar grid solo para las filas de datos
        if pagina == total_paginas - 1:
            # Última página: grid para datos (sin incluir totales)
            table_style.append(('GRID', (0, 0), (-1, num_filas_datos), 0.5, colors.grey))
            # Estilos especiales para filas de totales
            table_style.extend([
                ('BACKGROUND', (0, -3), (-1, -1), colors.white),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 10),
            ])
        else:
            # Páginas intermedias: grid para todas las filas
            table_style.append(('GRID', (0, 0), (-1, -1), 0.5, colors.grey))
        
        table.setStyle(TableStyle(table_style))
        elements.append(table)
        
        # Agregar página nueva si no es la última página
        if pagina < total_paginas - 1:
            from reportlab.platypus import PageBreak
            elements.append(PageBreak())
            # Repetir el encabezado de la factura en la siguiente página
            elements.append(Spacer(1, 0.3*inch))
            continue_style = ParagraphStyle('Continue', parent=styles['Normal'], fontSize=10,
                                          textColor=colors.HexColor('#CEB0B7'), alignment=TA_CENTER, fontName='Helvetica-Bold')
            elements.append(Paragraph(f"FACTURA {ncf} - Página {pagina + 2} de {total_paginas}", continue_style))
            elements.append(Spacer(1, 0.2*inch))
    
    # El footer ahora se dibuja automáticamente al final de cada página mediante el callback
    doc.build(elements, onFirstPage=agregar_footer, onLaterPages=agregar_footer)
    buffer.seek(0)
    return buffer

def generar_excel_factura(factura_id, ncf, fecha, pacientes, total, ncf_data=None, centro_medico=None, factura=None):
    """Generar Excel de la factura con el mismo formato que el PDF"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Factura {factura_id}"
    
    # Definir estilos
    header_font = Font(name='Arial', size=20, bold=True, color='000000')
    subtitle_font = Font(name='Arial', size=10, color='666666')
    section_font = Font(name='Arial', size=11, bold=True, color='B89BA3')
    bold_font = Font(name='Arial', size=10, bold=True)
    normal_font = Font(name='Arial', size=10)
    
    # Colores
    header_fill = PatternFill(start_color='CEB0B7', end_color='CEB0B7', fill_type='solid')
    info_fill = PatternFill(start_color='F2E2E6', end_color='F2E2E6', fill_type='solid')
    
    # Bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Alineaciones
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # HEADER: FACTURA
    row = 1
    ws.merge_cells(f'A{row}:G{row}')
    cell = ws[f'A{row}']
    cell.value = 'FACTURA'
    cell.font = header_font
    cell.alignment = center_align
    
    row += 1
    ws.merge_cells(f'A{row}:G{row}')
    cell = ws[f'A{row}']
    cell.value = f'CRÉDITO FISCAL #{factura_id}'
    cell.font = subtitle_font
    cell.alignment = center_align
    
    row += 2  # Espacio
    
    # INFORMACIÓN EN 3 COLUMNAS
    if pacientes and len(pacientes) > 0:
        # Extraer datos
        ars_nombre = pacientes[0].get('nombre_ars', 'N/A')
        ars_rnc = pacientes[0].get('ars_rnc', 'N/A')
        es_senasa = 'SENASA' in (str(ars_nombre or '')).upper()
        medico_nombre = pacientes[0].get('medico_nombre', 'N/A')
        medico_especialidad = pacientes[0].get('medico_especialidad', 'N/A')
        medico_cedula = pacientes[0].get('medico_cedula', 'N/A')
        medico_exequatur = pacientes[0].get('medico_exequatur', '')
        ncf_tipo = ncf_data.get('tipo', 'CRÉDITO FISCAL') if ncf_data else 'CRÉDITO FISCAL'
        ncf_fecha_fin = ncf_data.get('fecha_fin', '') if ncf_data else ''
        
        # Columna 1: Información de Factura
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = 'Información de Factura'
        cell.font = section_font
        cell.fill = info_fill
        cell.border = thin_border
        
        # Columna 2: NCF
        ws.merge_cells(f'C{row}:D{row}')
        cell = ws[f'C{row}']
        cell.value = 'NCF'
        cell.font = section_font
        cell.fill = info_fill
        cell.border = thin_border
        
        # Columna 3: Información del Médico
        ws.merge_cells(f'E{row}:G{row}')
        cell = ws[f'E{row}']
        cell.value = 'Información del Médico'
        cell.font = section_font
        cell.fill = info_fill
        cell.border = thin_border
        
        row += 1
        
        # Datos de columna 1
        col1_row_start = row
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = f'Fecha: {fecha}'
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].border = thin_border
        row += 1
        
        ws.merge_cells(f'A{row}:B{row}')
        cliente_text = f'Cliente (ARS): {ars_nombre}'
        if es_senasa:
            cliente_text += '\nREGIMEN CONTRIBUTIVO'
        ws[f'A{row}'] = cliente_text
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].border = thin_border
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
        row += 1
        
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = f'RNC: {ars_rnc}'
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].border = thin_border
        
        # Datos de columna 2 (NCF)
        row_ncf = col1_row_start
        ws.merge_cells(f'C{row_ncf}:D{row_ncf}')
        ws[f'C{row_ncf}'] = ncf
        ws[f'C{row_ncf}'].font = Font(name='Arial', size=12, bold=True, color='B89BA3')
        ws[f'C{row_ncf}'].border = thin_border
        row_ncf += 1
        
        ws.merge_cells(f'C{row_ncf}:D{row_ncf}')
        ws[f'C{row_ncf}'] = f'Tipo: {ncf_tipo}'
        ws[f'C{row_ncf}'].font = normal_font
        ws[f'C{row_ncf}'].border = thin_border
        row_ncf += 1
        
        if ncf_fecha_fin:
            ws.merge_cells(f'C{row_ncf}:D{row_ncf}')
            ws[f'C{row_ncf}'] = f'Válido hasta: {ncf_fecha_fin}'
            ws[f'C{row_ncf}'].font = normal_font
            ws[f'C{row_ncf}'].border = thin_border
        
        # Datos de columna 3 (Médico)
        row_med = col1_row_start
        ws.merge_cells(f'E{row_med}:G{row_med}')
        ws[f'E{row_med}'] = f'Médico: {medico_nombre}'
        ws[f'E{row_med}'].font = bold_font
        ws[f'E{row_med}'].border = thin_border
        row_med += 1
        
        ws.merge_cells(f'E{row_med}:G{row_med}')
        ws[f'E{row_med}'] = medico_especialidad
        ws[f'E{row_med}'].font = normal_font
        ws[f'E{row_med}'].border = thin_border
        row_med += 1
        
        ws.merge_cells(f'E{row_med}:G{row_med}')
        ws[f'E{row_med}'] = f'Cédula: {medico_cedula}'
        ws[f'E{row_med}'].font = normal_font
        ws[f'E{row_med}'].border = thin_border
        
        if medico_exequatur:
            row_med += 1
            ws.merge_cells(f'E{row_med}:G{row_med}')
            ws[f'E{row_med}'] = f'Exequátur: {medico_exequatur}'
            ws[f'E{row_med}'].font = normal_font
            ws[f'E{row_med}'].border = thin_border
        
        row += 3  # Espacio después de la info
    
    # TABLA DE PACIENTES
    headers = ['No.', 'NOMBRES PACIENTE', 'NSS/CONTRATO', 'FECHA', 'AUTORIZACIÓN', 'SERVICIO', 'V/UNITARIO']
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    row += 1
    
    # Datos de pacientes
    for idx, paciente in enumerate(pacientes, 1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=paciente['nombre_paciente'])
        ws.cell(row=row, column=3, value=paciente['nss'])
        ws.cell(row=row, column=4, value=paciente['fecha_servicio'])
        ws.cell(row=row, column=5, value=paciente['autorizacion'])
        ws.cell(row=row, column=6, value=paciente['descripcion_servicio'])
        ws.cell(row=row, column=7, value=float(paciente['monto']))
        
        # Aplicar estilos
        for col_idx in range(1, 8):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = normal_font
            cell.border = thin_border
            
            if col_idx == 1:
                cell.alignment = center_align
            elif col_idx == 7:
                cell.alignment = right_align
                cell.number_format = '"RD$"#,##0.00'
            else:
                cell.alignment = left_align
        
        row += 1
    
    row += 1  # Espacio antes de totales
    
    # TOTALES
    subtotal = total
    
    # SUB-TOTAL
    ws.merge_cells(f'F{row}:F{row}')
    cell = ws[f'F{row}']
    cell.value = 'SUB-TOTAL:'
    cell.font = bold_font
    cell.alignment = right_align
    
    cell = ws[f'G{row}']
    cell.value = float(subtotal)
    cell.font = bold_font
    cell.alignment = right_align
    cell.number_format = '"RD$"#,##0.00'
    
    row += 1
    
    # ITBIS
    ws.merge_cells(f'F{row}:F{row}')
    cell = ws[f'F{row}']
    cell.value = 'ITBIS:'
    cell.font = normal_font
    cell.alignment = right_align
    
    cell = ws[f'G{row}']
    cell.value = '*E'
    cell.font = normal_font
    cell.alignment = right_align
    
    row += 1
    
    # TOTAL
    ws.merge_cells(f'F{row}:F{row}')
    cell = ws[f'F{row}']
    cell.value = 'TOTAL:'
    cell.font = Font(name='Arial', size=12, bold=True)
    cell.alignment = right_align
    
    cell = ws[f'G{row}']
    cell.value = float(total)
    cell.font = Font(name='Arial', size=12, bold=True, color='B89BA3')
    cell.alignment = right_align
    cell.number_format = '"RD$"#,##0.00'
    
    row += 3  # Espacio
    
    # FOOTER
    if pacientes and len(pacientes) > 0:
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws[f'A{row}']
        cell.value = medico_nombre
        cell.font = bold_font
        cell.alignment = center_align
        row += 1
        
        footer_text = medico_especialidad
        if medico_cedula:
            footer_text += f' | Cédula: {medico_cedula}'
        if medico_exequatur:
            footer_text += f' | EXEQUATUR: {medico_exequatur}'
        
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws[f'A{row}']
        cell.value = footer_text
        cell.font = normal_font
        cell.alignment = center_align
        row += 1
        
        if centro_medico:
            centro_texto = f"{centro_medico['nombre']}, {centro_medico['direccion']}"
        else:
            centro_texto = 'Centro Oriental de Ginecología y Obstetricia, Zona Oriental, República Dominicana'
        
        ws.merge_cells(f'A{row}:G{row}')
        cell = ws[f'A{row}']
        cell.value = centro_texto
        cell.font = normal_font
        cell.alignment = center_align
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 15
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def enviar_email_factura(destinatario, factura_id, ncf, pdf_buffer, monto_total=0.00):
    """Enviar factura por email con template estandarizado usando SendGrid API"""
    try:
        if not EMAIL_CONFIGURED:
            print("\n⚠️  Email no configurado. La factura no se envió por correo.")
            return False
        
        # Usar template estandarizado
        html = template_factura(factura_id, ncf, monto_total)
        
        # Nombre del archivo
        filename = f'Factura_{ncf}_{factura_id}.pdf'
        
        # Enviar con SendGrid API
        success = send_email_sendgrid(
            to_email=destinatario,
            subject=f'💰 Factura #{factura_id} - NCF: {ncf}',
            html_content=html,
            attachment_data=pdf_buffer,
            attachment_filename=filename
        )
        
        if success:
            print("\n" + "=" * 60)
            print("✅ EMAIL CON FACTURA ENVIADO EXITOSAMENTE")
            print("=" * 60)
            print(f"📧 Destinatario: {destinatario}")
            print(f"📄 Factura: #{factura_id}")
            print(f"🔢 NCF: {ncf}")
            print(f"💰 Monto: ${monto_total:,.2f}")
            print("=" * 60 + "\n")
        
        return success
        
    except Exception as e:
        print("\n" + "=" * 60)
        print("❌ ERROR AL ENVIAR EMAIL CON FACTURA")
        print("=" * 60)
        print(f"Error: {e}")
        print("\nLa factura se generó correctamente y se descargó el PDF.")
        print("=" * 60 + "\n")
        return False

@app.route('/facturacion/historico')
@login_required
def facturacion_historico():
    """Histórico de facturas generadas con filtros"""
    try:
        conn = get_db_connection()
        
        # Obtener filtros
        ars_id = request.args.get('ars_id', type=int)
        medico_id = request.args.get('medico_id', type=int)
        ncf = request.args.get('ncf', '').strip()
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        
        # Construir query con filtros
        query = '''
            SELECT f.*, a.nombre_ars, m.nombre as medico_nombre,
                   (SELECT COUNT(*) FROM facturas_detalle WHERE factura_id = f.id) as num_pacientes
            FROM facturas f
            JOIN ars a ON f.ars_id = a.id
            JOIN medicos m ON f.medico_id = m.id
            WHERE f.activo = 1
        '''
        
        params = []
        
        if ars_id:
            query += ' AND f.ars_id = %s'
            params.append(ars_id)
        
        if medico_id:
            query += ' AND f.medico_id = %s'
            params.append(medico_id)
        
        if ncf:
            query += ' AND (f.ncf LIKE %s OR f.ncf_numero LIKE %s)'
            params.append(f'%{ncf}%')
            params.append(f'%{ncf}%')
        
        if fecha_desde:
            query += ' AND f.fecha_factura >= %s'
            params.append(fecha_desde)
        
        if fecha_hasta:
            query += ' AND f.fecha_factura <= %s'
            params.append(fecha_hasta)
        
        query += ' ORDER BY f.fecha_factura DESC, f.id DESC'
        
        facturas = conn.execute(query, params).fetchall()
        
        # Obtener lista de ARS para el filtro
        ars_list = conn.execute('SELECT * FROM ars WHERE activo = 1 ORDER BY nombre_ars').fetchall()
        
        # Obtener lista de Médicos habilitados para facturar
        medicos_list = conn.execute('SELECT * FROM medicos WHERE activo = 1 AND factura = 1 ORDER BY nombre').fetchall()
        
        conn.close()
        
        return render_template('facturacion/historico.html',
                             facturas=facturas,
                             ars_list=ars_list,
                             medicos_list=medicos_list,
                             ars_id_seleccionado=ars_id,
                             medico_id_seleccionado=medico_id,
                             ncf=ncf,
                             fecha_desde=fecha_desde,
                             fecha_hasta=fecha_hasta)
    except Exception as e:
        import traceback
        print(f"❌ ERROR EN HISTÓRICO DE FACTURAS:")
        print(traceback.format_exc())
        if 'conn' in locals():
            conn.close()
        flash(f'Error al cargar el histórico de facturas: {str(e)}', 'error')
        return redirect(url_for('facturacion_menu'))

@app.route('/facturacion/editar-factura/<int:factura_id>', methods=['GET', 'POST'])
@login_required
def facturacion_editar_factura(factura_id):
    """Editar factura existente - Solo si tiene menos de 30 días"""
    from datetime import datetime, timedelta
    
    conn = get_db_connection()
    
    # Obtener la factura
    factura = conn.execute('''
        SELECT f.*, a.nombre_ars, m.nombre as medico_nombre
        FROM facturas f
        JOIN ars a ON f.ars_id = a.id
        JOIN medicos m ON f.medico_id = m.id
        WHERE f.id = %s AND f.activo = 1
    ''', (factura_id,)).fetchone()
    
    if not factura:
        flash('Factura no encontrada', 'error')
        conn.close()
        return redirect(url_for('facturacion_historico'))
    
    # Validar que la factura tenga menos de 30 días
    fecha_factura = datetime.strptime(str(factura['fecha_factura']), '%Y-%m-%d')
    fecha_limite = fecha_factura + timedelta(days=30)
    fecha_actual = datetime.now()
    
    if fecha_actual > fecha_limite:
        dias_transcurridos = (fecha_actual - fecha_factura).days
        flash(f'⚠️ No se puede editar esta factura. Han transcurrido {dias_transcurridos} días desde su creación. El límite es 30 días.', 'error')
        conn.close()
        return redirect(url_for('facturacion_historico'))
    
    if request.method == 'POST':
        try:
            # Obtener los IDs de pacientes a agregar
            import json
            pacientes_agregar_ids = request.form.get('pacientes_agregar_ids', '[]')
            pacientes_agregar_list = json.loads(pacientes_agregar_ids)
            
            # Obtener los IDs de pacientes a eliminar
            pacientes_eliminar_ids = request.form.get('pacientes_eliminar_ids', '[]')
            pacientes_eliminar_list = json.loads(pacientes_eliminar_ids)
            
            print(f"📝 EDITAR FACTURA #{factura_id}")
            print(f"   Pacientes a agregar: {pacientes_agregar_list}")
            print(f"   Pacientes a eliminar: {pacientes_eliminar_list}")
            
            # AGREGAR PACIENTES NUEVOS
            if pacientes_agregar_list:
                placeholders = ','.join(['%s' for _ in pacientes_agregar_list])
                # Actualizar pacientes pendientes a facturados
                conn.execute(f'''
                    UPDATE facturas_detalle 
                    SET factura_id = %s, estado = 'facturado', medico_id = %s
                    WHERE id IN ({placeholders}) AND estado = 'pendiente'
                ''', [factura_id, factura['medico_id']] + pacientes_agregar_list)
                print(f"   ✅ {len(pacientes_agregar_list)} paciente(s) agregado(s)")
            
            # ELIMINAR PACIENTES DE LA FACTURA (regresar a pendiente)
            if pacientes_eliminar_list:
                placeholders = ','.join(['%s' for _ in pacientes_eliminar_list])
                conn.execute(f'''
                    UPDATE facturas_detalle 
                    SET factura_id = NULL, estado = 'pendiente', medico_id = NULL
                    WHERE id IN ({placeholders}) AND factura_id = %s
                ''', pacientes_eliminar_list + [factura_id])
                print(f"   ✅ {len(pacientes_eliminar_list)} paciente(s) eliminado(s)")
            
            # RECALCULAR TOTAL DE LA FACTURA
            total_result = conn.execute('''
                SELECT COALESCE(SUM(monto), 0) as total
                FROM facturas_detalle
                WHERE factura_id = %s AND estado = 'facturado'
            ''', (factura_id,)).fetchone()
            
            # Convertir a float para asegurar que se guarde correctamente
            nuevo_total = float(total_result['total']) if total_result['total'] else 0.0
            
            print(f"📊 RECALCULAR TOTAL - Factura #{factura_id}")
            print(f"   Total en DB (raw): {total_result['total']}")
            print(f"   Nuevo total convertido: RD${nuevo_total:,.2f}")
            
            # Actualizar el total en la factura
            conn.execute('''
                UPDATE facturas SET total = %s WHERE id = %s
            ''', (nuevo_total, factura_id))
            
            # Verificar que se actualizó correctamente
            verificacion = conn.execute('''
                SELECT total FROM facturas WHERE id = %s
            ''', (factura_id,)).fetchone()
            print(f"   Total en facturas después del UPDATE: {verificacion['total']}")
            print(f"   ✅ Total actualizado en base de datos")
            
            # REGISTRO DE AUDITORÍA
            observacion = f"Factura editada por {current_user.nombre}. "
            if pacientes_agregar_list:
                observacion += f"Agregados: {len(pacientes_agregar_list)} paciente(s). "
            if pacientes_eliminar_list:
                observacion += f"Eliminados: {len(pacientes_eliminar_list)} paciente(s). "
            observacion += f"Nuevo total: RD${nuevo_total:,.2f}"
            
            conn.execute('''
                UPDATE facturas 
                SET observaciones = CONCAT(COALESCE(observaciones, ''), '\n[', NOW(), '] ', %s)
                WHERE id = %s
            ''', (observacion, factura_id))
            
            # Hacer commit ANTES de cerrar la conexión
            conn.commit()
            print(f"   ✅ Commit realizado exitosamente")
            
            conn.close()
            
            flash(f'✅ Factura #{factura_id} actualizada exitosamente. Nuevo total: RD${nuevo_total:,.2f}', 'success')
            return redirect(url_for('facturacion_ver_factura', factura_id=factura_id))
            
        except Exception as e:
            print(f"❌ ERROR al editar factura: {str(e)}")
            import traceback
            traceback.print_exc()
            conn.rollback()
            conn.close()
            flash(f'Error al actualizar factura: {str(e)}', 'error')
            return redirect(url_for('facturacion_editar_factura', factura_id=factura_id))
    
    # GET: Mostrar formulario de edición
    
    # Verificar si la tabla servicios existe
    try:
        conn.execute("SELECT 1 FROM servicios LIMIT 1")
        servicios_existe = True
    except:
        servicios_existe = False
    
    # Obtener pacientes actuales de la factura
    if servicios_existe:
        pacientes_factura = conn.execute('''
            SELECT fd.*, 
                   COALESCE(m.nombre, 'Sin médico') as medico_nombre,
                   COALESCE(s.nombre, 'N/A') as servicio_nombre
            FROM facturas_detalle fd
            LEFT JOIN medicos m ON fd.medico_consulta = m.id
            LEFT JOIN servicios s ON fd.servicio_id = s.id
            WHERE fd.factura_id = %s AND fd.estado = 'facturado'
            ORDER BY fd.fecha_servicio DESC
        ''', (factura_id,)).fetchall()
    else:
        pacientes_factura = conn.execute('''
            SELECT fd.*, 
                   COALESCE(m.nombre, 'Sin médico') as medico_nombre,
                   'N/A' as servicio_nombre
            FROM facturas_detalle fd
            LEFT JOIN medicos m ON fd.medico_consulta = m.id
            WHERE fd.factura_id = %s AND fd.estado = 'facturado'
            ORDER BY fd.fecha_servicio DESC
        ''', (factura_id,)).fetchall()
    
    # Obtener pacientes pendientes del MISMO ARS (para agregar)
    if servicios_existe:
        pacientes_pendientes = conn.execute('''
            SELECT fd.*, 
                   COALESCE(m.nombre, 'Sin médico') as medico_nombre,
                   COALESCE(s.nombre, 'N/A') as servicio_nombre,
                   COALESCE(p.nombre, fd.nombre_paciente) as paciente_nombre_completo
            FROM facturas_detalle fd
            LEFT JOIN medicos m ON fd.medico_consulta = m.id
            LEFT JOIN servicios s ON fd.servicio_id = s.id
            LEFT JOIN pacientes p ON fd.paciente_id = p.id
            WHERE fd.estado = 'pendiente' AND fd.ars_id = %s
            ORDER BY fd.fecha_servicio DESC
        ''', (factura['ars_id'],)).fetchall()
    else:
        pacientes_pendientes = conn.execute('''
            SELECT fd.*, 
                   COALESCE(m.nombre, 'Sin médico') as medico_nombre,
                   'N/A' as servicio_nombre,
                   COALESCE(p.nombre, fd.nombre_paciente) as paciente_nombre_completo
            FROM facturas_detalle fd
            LEFT JOIN medicos m ON fd.medico_consulta = m.id
            LEFT JOIN pacientes p ON fd.paciente_id = p.id
            WHERE fd.estado = 'pendiente' AND fd.ars_id = %s
            ORDER BY fd.fecha_servicio DESC
        ''', (factura['ars_id'],)).fetchall()
    
    conn.close()
    
    # Calcular días restantes
    dias_transcurridos = (fecha_actual - fecha_factura).days
    dias_restantes = 30 - dias_transcurridos
    
    return render_template('facturacion/editar_factura.html',
                         factura=factura,
                         pacientes_factura=pacientes_factura,
                         pacientes_pendientes=pacientes_pendientes,
                         dias_restantes=dias_restantes,
                         dias_transcurridos=dias_transcurridos)

@app.route('/facturacion/dashboard')
@login_required
def facturacion_dashboard():
    """Dashboard de indicadores de facturación - Administradores, Registro de Facturas y Nivel 2"""
    # Perfiles permitidos
    if current_user.perfil not in ['Administrador', 'Registro de Facturas', 'Nivel 2']:
        flash('No tienes permisos para acceder a esta sección', 'error')
        return redirect(url_for('facturacion_menu'))
    
    from datetime import datetime, timedelta, timezone
    
    # Calcular fechas por defecto: últimos 12 meses (zona horaria RD)
    tz_rd = timezone(timedelta(hours=-4))
    fecha_actual = datetime.now(tz_rd)
    fecha_hace_12_meses = fecha_actual - timedelta(days=365)
    
    # Obtener filtros de fecha (por defecto: últimos 12 meses)
    fecha_desde = request.args.get('fecha_desde', default=fecha_hace_12_meses.strftime('%Y-%m-%d'))
    fecha_hasta = request.args.get('fecha_hasta', default=fecha_actual.strftime('%Y-%m-%d'))
    
    # Obtener filtros de ARS y Médico (pueden ser múltiples)
    ars_ids = request.args.getlist('ars_id', type=int)  # Lista de ARS
    medico_factura_ids = request.args.getlist('medico_factura_id', type=int)  # Médicos que facturan
    medico_consulta_ids = request.args.getlist('medico_consulta_id', type=int)  # Médicos que atienden
    
    # Filtrar valores vacíos o None
    ars_ids = [id for id in ars_ids if id]
    medico_factura_ids = [id for id in medico_factura_ids if id]
    medico_consulta_ids = [id for id in medico_consulta_ids if id]
    
    # Si medico_consulta_ids está presente, ignorar medico_factura_ids (mutuamente excluyentes)
    if medico_consulta_ids:
        medico_factura_ids = []
    
    conn = get_db_connection()
    
    # Obtener listas para los filtros
    ars_list = conn.execute('SELECT * FROM ars WHERE activo = 1 ORDER BY nombre_ars').fetchall()
    medicos_factura_list = conn.execute('SELECT * FROM medicos WHERE activo = 1 AND factura = 1 ORDER BY nombre').fetchall()
    medicos_consulta_list = conn.execute('SELECT * FROM medicos WHERE activo = 1 ORDER BY nombre').fetchall()
    
    # Variables de control para el template
    es_administrador = current_user.perfil == 'Administrador'
    medico_usuario_id = None
    
    # Si es perfil "Registro de Facturas" o "Nivel 2", filtrar por médico con mismo email
    if current_user.perfil in ['Registro de Facturas', 'Nivel 2']:
        # Buscar médico con el mismo email del usuario
        medico_usuario = conn.execute(
            'SELECT * FROM medicos WHERE email = %s AND activo = 1',
            (current_user.email,)
        ).fetchone()
        
        if medico_usuario:
            medico_usuario_id = medico_usuario['id']
            # Filtrar lista de médicos consulta para solo mostrar el suyo
            medicos_consulta_list = [medico_usuario]
            # Forzar el filtro automáticamente si no hay selección
            if not medico_consulta_ids:
                medico_consulta_ids = [medico_usuario_id]
        else:
            # Si no encuentra su médico, mostrar mensaje y redirigir
            flash('No se encontró un médico asociado a tu cuenta. Contacta al administrador.', 'error')
            conn.close()
            return redirect(url_for('facturacion_menu'))
    
    # ==================== INDICADORES GENERALES ====================
    # Crear placeholders para reutilizar en todas las queries
    placeholders_ars = ','.join(['%s'] * len(ars_ids)) if ars_ids else ''
    placeholders_medicos_f = ','.join(['%s'] * len(medico_factura_ids)) if medico_factura_ids else ''
    placeholders_medicos_c = ','.join(['%s'] * len(medico_consulta_ids)) if medico_consulta_ids else ''
    
    # Total de facturas generadas (en el rango y filtros)
    # Si se filtra por medico_consulta, contar facturas distintas con esos médicos
    if medico_consulta_ids:
        # Contar facturas distintas que tienen esos médicos consulta
        query_facturas = '''
            SELECT COUNT(DISTINCT fd.factura_id) as total 
            FROM facturas_detalle fd
            JOIN facturas f ON fd.factura_id = f.id
            WHERE fd.activo = 1 
            AND f.fecha_factura BETWEEN %s AND %s
            AND fd.medico_consulta IN ({})
        '''.format(placeholders_medicos_c)
        params_facturas = [fecha_desde, fecha_hasta] + medico_consulta_ids
        
        if ars_ids:
            query_facturas += ' AND f.ars_id IN (' + placeholders_ars + ')'
            params_facturas.extend(ars_ids)
    else:
        # Contar facturas normalmente
        query_facturas = '''
            SELECT COUNT(*) as total FROM facturas 
            WHERE activo = 1 
            AND fecha_factura BETWEEN %s AND %s
        '''
        params_facturas = [fecha_desde, fecha_hasta]
        
        if ars_ids:
            query_facturas += ' AND ars_id IN (' + placeholders_ars + ')'
            params_facturas.extend(ars_ids)
        
        if medico_factura_ids:
            query_facturas += ' AND medico_id IN (' + placeholders_medicos_f + ')'
            params_facturas.extend(medico_factura_ids)
    
    total_facturas = conn.execute(query_facturas, params_facturas).fetchone()['total']
    
    # Total facturado (monto en el rango y filtros)
    if medico_consulta_ids:
        # Si filtramos por médicos consulta, sumar desde facturas_detalle
        query_monto = '''
            SELECT COALESCE(SUM(fd.monto), 0) as total 
            FROM facturas_detalle fd
            JOIN facturas f ON fd.factura_id = f.id
            WHERE fd.activo = 1 
            AND f.fecha_factura BETWEEN %s AND %s
            AND fd.medico_consulta IN (''' + ','.join(['%s'] * len(medico_consulta_ids)) + ''')
        '''
        params_monto = [fecha_desde, fecha_hasta] + medico_consulta_ids
        
        if ars_ids:
            query_monto += ' AND f.ars_id IN (' + placeholders_ars + ')'
            params_monto.extend(ars_ids)
    else:
        # Si no hay filtro de médico consulta, usar facturas normal
        query_monto = '''
            SELECT COALESCE(SUM(total), 0) as total FROM facturas 
            WHERE activo = 1 
            AND fecha_factura BETWEEN %s AND %s
        '''
        params_monto = [fecha_desde, fecha_hasta]
        
        if ars_ids:
            query_monto += ' AND ars_id IN (' + placeholders_ars + ')'
            params_monto.extend(ars_ids)
        
        if medico_factura_ids:
            query_monto += ' AND medico_id IN (' + placeholders_medicos_f + ')'
            params_monto.extend(medico_factura_ids)
    
    total_facturado = conn.execute(query_monto, params_monto).fetchone()['total']
    
    # ARS Pendientes por Facturar (nombres de ARS con pacientes pendientes y montos)
    query_ars_pendientes = '''
        SELECT a.id as ars_id, a.nombre_ars, COALESCE(SUM(fd.monto), 0) as monto_pendiente
        FROM facturas_detalle fd
        JOIN ars a ON fd.ars_id = a.id
        WHERE fd.estado = 'pendiente'
    '''
    params_ars_pendientes = []
    
    if ars_ids:
        query_ars_pendientes += ' AND fd.ars_id IN (' + placeholders_ars + ')'
        params_ars_pendientes.extend(ars_ids)
    
    if medico_factura_ids:
        query_ars_pendientes += ' AND fd.medico_id IN (' + placeholders_medicos_f + ')'
        params_ars_pendientes.extend(medico_factura_ids)
    
    if medico_consulta_ids:
        query_ars_pendientes += ' AND fd.medico_consulta IN (' + placeholders_medicos_c + ')'
        params_ars_pendientes.extend(medico_consulta_ids)
    
    query_ars_pendientes += ' GROUP BY a.id, a.nombre_ars ORDER BY monto_pendiente DESC, a.nombre_ars'
    
    try:
        ars_pendientes = conn.execute(query_ars_pendientes, params_ars_pendientes).fetchall()
        # Convertir a lista de diccionarios con id, nombre y monto
        ars_pendientes_detalle = []
        for row in ars_pendientes:
            try:
                ars_detalle = {
                    'id': row.get('ars_id') or row.get('id', 0),
                    'nombre': row.get('nombre_ars', 'N/A'),
                    'monto': row.get('monto_pendiente', 0)
                }
                ars_pendientes_detalle.append(ars_detalle)
            except Exception as e:
                print(f"⚠️ Error procesando ARS: {e}")
                continue
    except Exception as e:
        print(f"❌ Error en query ARS pendientes: {e}")
        import traceback
        traceback.print_exc()
        # Fallback: lista vacía
        ars_pendientes_detalle = []
    
    # Monto total pendiente por facturar (suma de todos los costos de servicios pendientes)
    query_monto_pendiente = '''
        SELECT COALESCE(SUM(fd.monto), 0) as total 
        FROM facturas_detalle fd
        WHERE fd.estado = 'pendiente'
    '''
    params_monto_pendiente = []
    
    if ars_ids:
        query_monto_pendiente += ' AND fd.ars_id IN (' + placeholders_ars + ')'
        params_monto_pendiente.extend(ars_ids)
    
    if medico_factura_ids:
        query_monto_pendiente += ' AND fd.medico_id IN (' + placeholders_medicos_f + ')'
        params_monto_pendiente.extend(medico_factura_ids)
    
    if medico_consulta_ids:
        query_monto_pendiente += ' AND fd.medico_consulta IN (' + placeholders_medicos_c + ')'
        params_monto_pendiente.extend(medico_consulta_ids)
    
    monto_pendiente = conn.execute(query_monto_pendiente, params_monto_pendiente).fetchone()['total']
    
    # Pacientes facturados (en el rango y filtros)
    query_facturados = '''
        SELECT COUNT(*) as total FROM facturas_detalle fd
        JOIN facturas f ON fd.factura_id = f.id
        WHERE fd.estado = 'facturado'
        AND f.fecha_factura BETWEEN %s AND %s
    '''
    params_facturados = [fecha_desde, fecha_hasta]
    
    if ars_ids:
        query_facturados += ' AND f.ars_id IN (' + placeholders_ars + ')'
        params_facturados.extend(ars_ids)
    
    if medico_consulta_ids:
        query_facturados += ' AND fd.medico_consulta IN (' + placeholders_medicos_c + ')'
        params_facturados.extend(medico_consulta_ids)
    elif medico_factura_ids:
        query_facturados += ' AND f.medico_id IN (' + placeholders_medicos_f + ')'
        params_facturados.extend(medico_factura_ids)
    
    pacientes_facturados = conn.execute(query_facturados, params_facturados).fetchone()['total']
    
    # ==================== FACTURACIÓN POR MES ====================
    if medico_consulta_ids:
        # Si filtramos por médicos consulta, sumar desde facturas_detalle
        query_mes = '''
            SELECT 
                DATE_FORMAT(f.fecha_factura, '%%Y-%%m') as mes,
                COALESCE(SUM(fd.monto), 0) as total_monto
            FROM facturas_detalle fd
            JOIN facturas f ON fd.factura_id = f.id
            WHERE fd.activo = 1 
            AND f.fecha_factura BETWEEN %s AND %s
            AND fd.medico_consulta IN (''' + ','.join(['%s'] * len(medico_consulta_ids)) + ''')
        '''
        params_mes = [fecha_desde, fecha_hasta] + medico_consulta_ids
        
        if ars_ids:
            query_mes += ' AND f.ars_id IN (' + placeholders_ars + ')'
            params_mes.extend(ars_ids)
        
        query_mes += ' GROUP BY DATE_FORMAT(f.fecha_factura, \'%%Y-%%m\') ORDER BY mes ASC'
    else:
        # Si no hay filtro de médico consulta, usar facturas normal
        query_mes = '''
            SELECT 
                DATE_FORMAT(fecha_factura, '%%Y-%%m') as mes,
                COALESCE(SUM(total), 0) as total_monto
            FROM facturas
            WHERE activo = 1 
            AND fecha_factura BETWEEN %s AND %s
        '''
        params_mes = [fecha_desde, fecha_hasta]
        
        if ars_ids:
            query_mes += ' AND ars_id IN (' + placeholders_ars + ')'
            params_mes.extend(ars_ids)
        
        if medico_factura_ids:
            query_mes += ' AND medico_id IN (' + placeholders_medicos_f + ')'
            params_mes.extend(medico_factura_ids)
        
        query_mes += ' GROUP BY DATE_FORMAT(fecha_factura, \'%%Y-%%m\') ORDER BY mes ASC'
    
    facturacion_por_mes = conn.execute(query_mes, params_mes).fetchall()
    
    # ==================== FACTURACIÓN POR ARS ====================
    if medico_consulta_ids:
        # Si filtramos por médicos consulta, sumar desde facturas_detalle
        query_ars = '''
            SELECT a.nombre_ars, COALESCE(SUM(fd.monto), 0) as total_monto
            FROM facturas_detalle fd
            JOIN facturas f ON fd.factura_id = f.id
            JOIN ars a ON f.ars_id = a.id
            WHERE fd.activo = 1
            AND f.fecha_factura BETWEEN %s AND %s
            AND fd.medico_consulta IN (''' + ','.join(['%s'] * len(medico_consulta_ids)) + ''')
        '''
        params_ars = [fecha_desde, fecha_hasta] + medico_consulta_ids
        
        if ars_ids:
            query_ars += ' AND f.ars_id IN (' + placeholders_ars + ')'
            params_ars.extend(ars_ids)
        
        query_ars += ' GROUP BY a.id, a.nombre_ars ORDER BY total_monto DESC'
    else:
        # Si no hay filtro de médico consulta, usar facturas normal
        query_ars = '''
            SELECT a.nombre_ars, COALESCE(SUM(f.total), 0) as total_monto
            FROM facturas f
            JOIN ars a ON f.ars_id = a.id
            WHERE f.activo = 1
            AND f.fecha_factura BETWEEN %s AND %s
        '''
        params_ars = [fecha_desde, fecha_hasta]
        
        if ars_ids:
            query_ars += ' AND f.ars_id IN (' + placeholders_ars + ')'
            params_ars.extend(ars_ids)
        
        if medico_factura_ids:
            query_ars += ' AND f.medico_id IN (' + placeholders_medicos_f + ')'
            params_ars.extend(medico_factura_ids)
        
        query_ars += ' GROUP BY a.id, a.nombre_ars ORDER BY total_monto DESC'
    
    facturacion_por_ars = conn.execute(query_ars, params_ars).fetchall()
    
    # ==================== FACTURACIÓN POR ARS Y MES (para gráfico) ====================
    if medico_consulta_ids:
        # Si filtramos por médicos consulta, sumar desde facturas_detalle
        query_ars_mes = '''
            SELECT 
                a.nombre_ars,
                DATE_FORMAT(f.fecha_factura, '%%Y-%%m') as mes,
                COALESCE(SUM(fd.monto), 0) as total_monto
            FROM facturas_detalle fd
            JOIN facturas f ON fd.factura_id = f.id
            JOIN ars a ON f.ars_id = a.id
            WHERE fd.activo = 1
            AND f.fecha_factura BETWEEN %s AND %s
            AND fd.medico_consulta IN (''' + ','.join(['%s'] * len(medico_consulta_ids)) + ''')
        '''
        params_ars_mes = [fecha_desde, fecha_hasta] + medico_consulta_ids
        
        if ars_ids:
            query_ars_mes += ' AND f.ars_id IN (' + placeholders_ars + ')'
            params_ars_mes.extend(ars_ids)
        
        query_ars_mes += ' GROUP BY a.id, a.nombre_ars, DATE_FORMAT(f.fecha_factura, \'%%Y-%%m\') ORDER BY a.nombre_ars, mes ASC'
    else:
        # Si no hay filtro de médico consulta, usar facturas normal
        query_ars_mes = '''
            SELECT 
                a.nombre_ars,
                DATE_FORMAT(f.fecha_factura, '%%Y-%%m') as mes,
                COALESCE(SUM(f.total), 0) as total_monto
            FROM facturas f
            JOIN ars a ON f.ars_id = a.id
            WHERE f.activo = 1
            AND f.fecha_factura BETWEEN %s AND %s
        '''
        params_ars_mes = [fecha_desde, fecha_hasta]
        
        if ars_ids:
            query_ars_mes += ' AND f.ars_id IN (' + placeholders_ars + ')'
            params_ars_mes.extend(ars_ids)
        
        if medico_factura_ids:
            query_ars_mes += ' AND f.medico_id IN (' + placeholders_medicos_f + ')'
            params_ars_mes.extend(medico_factura_ids)
        
        query_ars_mes += ' GROUP BY a.id, a.nombre_ars, DATE_FORMAT(f.fecha_factura, \'%%Y-%%m\') ORDER BY a.nombre_ars, mes ASC'
    
    facturacion_ars_mes = conn.execute(query_ars_mes, params_ars_mes).fetchall()
    
    conn.close()
    
    return render_template('facturacion/dashboard.html',
                         total_facturas=total_facturas,
                         total_facturado=total_facturado,
                         ars_pendientes_detalle=ars_pendientes_detalle,
                         monto_pendiente=monto_pendiente,
                         pacientes_facturados=pacientes_facturados,
                         facturacion_por_mes=facturacion_por_mes,
                         facturacion_por_ars=facturacion_por_ars,
                         facturacion_ars_mes=facturacion_ars_mes,
                         fecha_desde=fecha_desde,
                         fecha_hasta=fecha_hasta,
                         ars_list=ars_list,
                         medicos_factura_list=medicos_factura_list,
                         medicos_consulta_list=medicos_consulta_list,
                         ars_ids_seleccionados=ars_ids,
                         medico_factura_ids_seleccionados=medico_factura_ids,
                         medico_consulta_ids_seleccionados=medico_consulta_ids,
                         es_administrador=es_administrador,
                         medico_usuario_id=medico_usuario_id)

@app.route('/facturacion/ver-factura/<int:factura_id>')
@login_required
def facturacion_ver_factura(factura_id):
    """Ver factura completa generada"""
    conn = get_db_connection()
    
    # Obtener datos de la factura (incluyendo email del médico)
    factura = conn.execute('''
        SELECT f.*, a.nombre_ars, a.rnc as ars_rnc, 
               m.nombre as medico_nombre, m.especialidad as medico_especialidad, 
               m.cedula as medico_cedula, m.exequatur as medico_exequatur,
               m.email as medico_email,
               n.tipo as ncf_tipo, n.prefijo as ncf_prefijo, n.fecha_fin as ncf_fecha_fin
        FROM facturas f
        JOIN ars a ON f.ars_id = a.id
        JOIN medicos m ON f.medico_id = m.id
        LEFT JOIN ncf n ON f.ncf_id = n.id
        WHERE f.id = %s AND f.activo = 1
    ''', (factura_id,)).fetchone()
    
    if not factura:
        flash('Factura no encontrada', 'error')
        conn.close()
        return redirect(url_for('facturacion_historico'))
    
    # Obtener pacientes de la factura
    pacientes = conn.execute('''
        SELECT fd.*, m.nombre as medico_nombre
        FROM facturas_detalle fd
        JOIN medicos m ON fd.medico_id = m.id
        WHERE fd.factura_id = %s
        ORDER BY fd.id
    ''', (factura_id,)).fetchall()
    
    # Obtener centro médico por defecto del médico
    centro_medico = conn.execute('''
        SELECT c.nombre, c.direccion
        FROM medico_centro mc
        JOIN centros_medicos c ON mc.centro_id = c.id
        WHERE mc.medico_id = %s AND mc.es_defecto = 1 AND mc.activo = 1 AND c.activo = 1
        LIMIT 1
    ''', (factura['medico_id'],)).fetchone()
    
    # Si no hay centro por defecto, buscar el primero disponible
    if not centro_medico:
        centro_medico = conn.execute('''
            SELECT c.nombre, c.direccion
            FROM medico_centro mc
            JOIN centros_medicos c ON mc.centro_id = c.id
            WHERE mc.medico_id = %s AND mc.activo = 1 AND c.activo = 1
            LIMIT 1
        ''', (factura['medico_id'],)).fetchone()
    
    # Si aún no hay centro, usar valores por defecto
    if not centro_medico:
        centro_medico = {
            'nombre': 'Centro Oriental de Ginecología y Obstetricia',
            'direccion': 'Zona Oriental, República Dominicana'
        }
    
    conn.close()
    
    # Calcular totales (sin ITBIS)
    subtotal = factura['total']
    itbis = 0
    total_final = subtotal
    
    return render_template('facturacion/ver_factura.html',
                         factura=factura,
                         pacientes=pacientes,
                         subtotal=subtotal,
                         itbis=itbis,
                         total=total_final,
                         centro_medico=centro_medico)

@app.route('/facturacion/enviar-email/<int:factura_id>', methods=['POST'])
@login_required
def facturacion_enviar_email(factura_id):
    """Enviar factura por email con PDF adjunto"""
    conn = get_db_connection()
    
    try:
        # Obtener destinatario del formulario
        destinatario = request.form.get('destinatario', '').strip()
        
        if not destinatario:
            flash('Debes proporcionar un email de destinatario', 'error')
            return redirect(url_for('facturacion_ver_factura', factura_id=factura_id))
        
        # Obtener datos de la factura
        factura = conn.execute('''
            SELECT f.*, a.nombre_ars, a.rnc as ars_rnc, 
                   m.nombre as medico_nombre, m.especialidad as medico_especialidad, 
                   m.cedula as medico_cedula, m.exequatur as medico_exequatur,
                   m.email as medico_email,
                   n.tipo as ncf_tipo, n.prefijo as ncf_prefijo, n.fecha_fin as ncf_fecha_fin
            FROM facturas f
            JOIN ars a ON f.ars_id = a.id
            JOIN medicos m ON f.medico_id = m.id
            LEFT JOIN ncf n ON f.ncf_id = n.id
            WHERE f.id = %s AND f.activo = 1
        ''', (factura_id,)).fetchone()
        
        if not factura:
            flash('Factura no encontrada', 'error')
            conn.close()
            return redirect(url_for('facturacion_historico'))
        
        # Obtener pacientes de la factura
        pacientes = conn.execute('''
            SELECT fd.*, m.nombre as medico_nombre, m.especialidad as medico_especialidad,
                   m.cedula as medico_cedula, m.exequatur as medico_exequatur,
                   a.nombre_ars, a.rnc as ars_rnc, ca.codigo_ars
            FROM facturas_detalle fd
            JOIN medicos m ON fd.medico_id = m.id
            JOIN ars a ON fd.ars_id = a.id
            LEFT JOIN codigo_ars ca ON ca.medico_id = m.id AND ca.ars_id = a.id
            WHERE fd.factura_id = %s
            ORDER BY fd.id
        ''', (factura_id,)).fetchall()
        
        if not pacientes:
            flash('No se encontraron pacientes en esta factura', 'error')
            conn.close()
            return redirect(url_for('facturacion_ver_factura', factura_id=factura_id))
        
        # Generar PDF
        if REPORTLAB_AVAILABLE:
            # Obtener datos del NCF
            ncf_data = conn.execute('SELECT fecha_fin, tipo FROM ncf WHERE id = %s', (factura['ncf_id'],)).fetchone()
            
            # Generar PDF
            ncf_numero = factura['ncf_numero'] if factura['ncf_numero'] else factura.get('ncf', 'N/A')
            pdf_buffer = generar_pdf_factura(factura_id, ncf_numero, factura['fecha_factura'], pacientes, factura['total'], ncf_data)
            
            # Enviar email de manera asíncrona (no bloquea la respuesta)
            # Crear copia del buffer para el thread
            email_pdf_buffer = BytesIO(pdf_buffer.getvalue())
            
            def enviar_email_async():
                if enviar_email_factura(destinatario, factura_id, ncf_numero, email_pdf_buffer, factura['total']):
                    print(f'✅ Factura enviada exitosamente a {destinatario}')
                else:
                    print('⚠️ Hubo un problema al enviar el email')
            
            threading.Thread(target=enviar_email_async).start()
            flash(f'✅ Factura enviándose a {destinatario}. Recibirás el email en breve.', 'success')
        else:
            flash('No se puede generar el PDF. ReportLab no está disponible.', 'error')
        
        conn.close()
        return redirect(url_for('facturacion_ver_factura', factura_id=factura_id))
        
    except Exception as e:
        print(f"❌ Error al enviar email de factura: {e}")
        if conn:
            conn.close()
        flash(f'Error al enviar la factura por email: {str(e)}', 'error')
        return redirect(url_for('facturacion_ver_factura', factura_id=factura_id))

@app.route('/facturacion/descargar-pdf/<int:factura_id>')
@login_required
def facturacion_descargar_pdf(factura_id):
    """Descargar PDF de factura generada"""
    conn = get_db_connection()
    
    try:
        # Obtener datos de la factura
        factura = conn.execute('''
            SELECT f.*, a.nombre_ars, a.rnc as ars_rnc, 
                   m.nombre as medico_nombre, m.especialidad as medico_especialidad, 
                   m.cedula as medico_cedula, m.exequatur as medico_exequatur,
                   m.email as medico_email,
                   n.tipo as ncf_tipo, n.prefijo as ncf_prefijo, n.fecha_fin as ncf_fecha_fin
            FROM facturas f
            JOIN ars a ON f.ars_id = a.id
            JOIN medicos m ON f.medico_id = m.id
            LEFT JOIN ncf n ON f.ncf_id = n.id
            WHERE f.id = %s AND f.activo = 1
        ''', (factura_id,)).fetchone()
        
        if not factura:
            flash('Factura no encontrada', 'error')
            conn.close()
            return redirect(url_for('facturacion_historico'))
        
        # Obtener pacientes de la factura
        pacientes = conn.execute('''
            SELECT fd.*, m.nombre as medico_nombre, m.especialidad as medico_especialidad,
                   m.cedula as medico_cedula, m.exequatur as medico_exequatur,
                   a.nombre_ars, a.rnc as ars_rnc, ca.codigo_ars
            FROM facturas_detalle fd
            JOIN medicos m ON fd.medico_id = m.id
            JOIN ars a ON fd.ars_id = a.id
            LEFT JOIN codigo_ars ca ON ca.medico_id = m.id AND ca.ars_id = a.id
            WHERE fd.factura_id = %s
            ORDER BY fd.id
        ''', (factura_id,)).fetchall()
        
        if not pacientes:
            flash('No se encontraron pacientes en esta factura', 'error')
            conn.close()
            return redirect(url_for('facturacion_ver_factura', factura_id=factura_id))
        
        # Generar PDF
        if REPORTLAB_AVAILABLE:
            # Obtener datos del NCF
            ncf_data = conn.execute('SELECT fecha_fin, tipo FROM ncf WHERE id = %s', (factura['ncf_id'],)).fetchone()
            
            # Generar PDF
            ncf_numero = factura['ncf_numero'] if factura['ncf_numero'] else factura.get('ncf', 'N/A')
            pdf_buffer = generar_pdf_factura(factura_id, ncf_numero, factura['fecha_factura'], pacientes, factura['total'], ncf_data)
            
            conn.close()
            
            # Nombre del archivo
            nombre_archivo = f"Factura_{factura_id}_NCF_{ncf_numero.replace(' ', '_')}.pdf"
            
            return send_file(
                pdf_buffer,
                mimetype='application/pdf',
                as_attachment=True,
                download_name=nombre_archivo
            )
        else:
            flash('No se puede generar el PDF. ReportLab no está disponible.', 'error')
            conn.close()
            return redirect(url_for('facturacion_ver_factura', factura_id=factura_id))
        
    except Exception as e:
        print(f"❌ Error al descargar PDF de factura: {e}")
        if conn:
            conn.close()
        flash(f'Error al descargar el PDF: {str(e)}', 'error')
        return redirect(url_for('facturacion_ver_factura', factura_id=factura_id))

@app.route('/facturacion/descargar-excel/<int:factura_id>')
@login_required
def facturacion_descargar_excel(factura_id):
    """Descargar Excel de factura generada"""
    conn = get_db_connection()
    
    try:
        # Obtener datos de la factura
        factura = conn.execute('''
            SELECT f.*, a.nombre_ars, a.rnc as ars_rnc, 
                   m.nombre as medico_nombre, m.especialidad as medico_especialidad, 
                   m.cedula as medico_cedula, m.exequatur as medico_exequatur,
                   m.email as medico_email,
                   n.tipo as ncf_tipo, n.prefijo as ncf_prefijo, n.fecha_fin as ncf_fecha_fin
            FROM facturas f
            JOIN ars a ON f.ars_id = a.id
            JOIN medicos m ON f.medico_id = m.id
            LEFT JOIN ncf n ON f.ncf_id = n.id
            WHERE f.id = %s AND f.activo = 1
        ''', (factura_id,)).fetchone()
        
        if not factura:
            flash('Factura no encontrada', 'error')
            conn.close()
            return redirect(url_for('facturacion_historico'))
        
        # Obtener pacientes de la factura
        pacientes = conn.execute('''
            SELECT fd.*, m.nombre as medico_nombre, m.especialidad as medico_especialidad,
                   m.cedula as medico_cedula, m.exequatur as medico_exequatur,
                   a.nombre_ars, a.rnc as ars_rnc, ca.codigo_ars
            FROM facturas_detalle fd
            JOIN medicos m ON fd.medico_id = m.id
            JOIN ars a ON fd.ars_id = a.id
            LEFT JOIN codigo_ars ca ON ca.medico_id = m.id AND ca.ars_id = a.id
            WHERE fd.factura_id = %s
            ORDER BY fd.id
        ''', (factura_id,)).fetchall()
        
        if not pacientes:
            flash('No se encontraron pacientes en esta factura', 'error')
            conn.close()
            return redirect(url_for('facturacion_ver_factura', factura_id=factura_id))
        
        # Obtener datos del NCF
        ncf_data = conn.execute('SELECT fecha_fin, tipo FROM ncf WHERE id = %s', (factura['ncf_id'],)).fetchone()
        
        # Obtener centro médico (con manejo de error si no existe centro_medico_id)
        centro_medico = None
        if 'centro_medico_id' in factura.keys() and factura['centro_medico_id']:
            centro_medico = conn.execute('SELECT * FROM centros_medicos WHERE id = %s', (factura['centro_medico_id'],)).fetchone()
        
        # Si no hay centro médico, obtener el predeterminado
        if not centro_medico:
            centro_medico = conn.execute('SELECT * FROM centros_medicos WHERE id = 1').fetchone()
        
        # Generar Excel
        ncf_numero = factura['ncf_numero'] if factura['ncf_numero'] else factura.get('ncf', 'N/A')
        excel_buffer = generar_excel_factura(factura_id, ncf_numero, factura['fecha_factura'], pacientes, factura['total'], ncf_data, centro_medico, factura)
        
        conn.close()
        
        # Nombre del archivo
        nombre_archivo = f"Factura_{factura_id}_NCF_{ncf_numero.replace(' ', '_')}.xlsx"
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre_archivo
        )
        
    except Exception as e:
        print(f"❌ Error al descargar Excel de factura: {e}")
        import traceback
        traceback.print_exc()
        if conn:
            conn.close()
        flash(f'Error al descargar el Excel: {str(e)}', 'error')
        return redirect(url_for('facturacion_ver_factura', factura_id=factura_id))

@app.route('/facturacion/paciente/<int:paciente_id>/editar', methods=['GET', 'POST'])
@login_required
def facturacion_paciente_editar(paciente_id):
    """Editar un paciente pendiente de facturación"""
    conn = get_db_connection()
    
    # Detectar si viene de facturación (doble click)
    from_factura = request.args.get('from_factura') == '1' or request.form.get('from_factura') == '1'
    ars_id_return = request.args.get('ars_id') or request.form.get('ars_id')
    ncf_id_return = request.args.get('ncf_id') or request.form.get('ncf_id')
    medico_id_return = request.args.get('medico_id') or request.form.get('medico_id')
    fecha_return = request.args.get('fecha') or request.form.get('fecha')
    
    # Verificar que el paciente existe y está pendiente
    paciente = conn.execute('''
        SELECT * FROM facturas_detalle WHERE id = %s
    ''', (paciente_id,)).fetchone()
    
    if not paciente:
        flash('Paciente no encontrado', 'error')
        return redirect(url_for('facturacion_pacientes_pendientes'))
    
    if paciente['estado'] != 'pendiente':
        flash('No se puede editar un registro facturado', 'error')
        return redirect(url_for('facturacion_pacientes_pendientes'))
    
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            nss_raw = request.form.get('nss', '')
            nombre_raw = request.form.get('nombre', '')
            fecha = request.form.get('fecha', '').strip()
            autorizacion_raw = request.form.get('autorizacion', '')
            servicio_raw = request.form.get('servicio', '')
            monto = request.form.get('monto', type=float, default=0)

            # Validar caracteres (integridad)
            try:
                nss = _validate_or_raise(nss_raw, _RE_NSS, 'NSS', max_len=20)
                nombre = _validate_or_raise(nombre_raw, _RE_NOMBRE, 'Nombre del paciente', max_len=200)
                servicio = _validate_or_raise(servicio_raw, _RE_TEXTO_GENERAL, 'Servicio', max_len=200)
                autorizacion = _validate_or_raise(autorizacion_raw, _RE_AUTORIZACION, 'Autorización', max_len=50)
            except ValueError as ve:
                flash(f'❌ {str(ve)}', 'error')
                conn.close()
                return redirect(url_for('facturacion_paciente_editar', paciente_id=paciente_id))
            
            # Validaciones
            if not nss or not nombre:
                flash('NSS y Nombre son obligatorios', 'error')
                conn.close()
                return redirect(url_for('facturacion_paciente_editar', paciente_id=paciente_id))
            
            # Validar que la fecha no sea futura
            try:
                from datetime import date
                fecha_obj = date.fromisoformat(fecha) if isinstance(fecha, str) else fecha
                fecha_hoy_rd = obtener_fecha_rd()
                if fecha_obj > fecha_hoy_rd:
                    flash('❌ La fecha de consulta no puede ser futura', 'error')
                    conn.close()
                    return redirect(url_for('facturacion_paciente_editar', paciente_id=paciente_id))
            except:
                pass  # Si hay error en el formato, dejar que continúe (se validará después)
            
            # Verificar autorización única (excepto el mismo registro)
            existe_autorizacion = conn.execute('''
                SELECT id FROM facturas_detalle 
                WHERE autorizacion = %s AND id != %s
            ''', (autorizacion, paciente_id)).fetchone()
            
            if existe_autorizacion:
                flash('Esta autorización ya existe en otro registro', 'error')
                conn.close()
                return redirect(url_for('facturacion_paciente_editar', paciente_id=paciente_id))
            
            # Actualizar registro
            conn.execute('''
                UPDATE facturas_detalle
                SET nss = %s, nombre_paciente = %s, fecha_servicio = %s,
                    autorizacion = %s, descripcion_servicio = %s, monto = %s
                WHERE id = %s
            ''', (nss, nombre, fecha, autorizacion, servicio, monto, paciente_id))
            
            conn.commit()
            conn.close()
            
            flash('✅ Registro actualizado exitosamente', 'success')
            
            # Si viene de facturación, volver ahí con formulario auto-submit
            if from_factura and ars_id_return and ncf_id_return and medico_id_return and fecha_return:
                # Renderizar página intermedia que hace POST automático
                return render_template('facturacion/volver_facturacion.html',
                                     ars_id=ars_id_return,
                                     ncf_id=ncf_id_return,
                                     medico_id=medico_id_return,
                                     fecha=fecha_return)
            
            return redirect(url_for('facturacion_pacientes_pendientes'))
            
        except Exception as e:
            conn.close()
            flash(f'Error al actualizar: {str(e)}', 'error')
            return redirect(url_for('facturacion_paciente_editar', paciente_id=paciente_id))
    
    # GET: Mostrar formulario de edición
    # Obtener listas para los selects
    medicos = conn.execute('SELECT * FROM medicos WHERE activo = 1 ORDER BY nombre').fetchall()
    ars_list = conn.execute('SELECT * FROM ars WHERE activo = 1 ORDER BY nombre_ars').fetchall()
    conn.close()
    
    # Fecha actual para validación (no permitir fechas futuras) - Zona horaria RD
    fecha_actual = obtener_fecha_rd().strftime('%Y-%m-%d')

    # Normalizar fecha_servicio para input type="date" (debe ser YYYY-MM-DD)
    fecha_servicio_iso = ''
    try:
        from datetime import datetime
        fs = paciente.get('fecha_servicio')
        if fs:
            if hasattr(fs, 'strftime'):
                fecha_servicio_iso = fs.strftime('%Y-%m-%d')
            elif isinstance(fs, str):
                fs = fs.strip()
                for fmt in ['%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y']:
                    try:
                        fecha_servicio_iso = datetime.strptime(fs, fmt).strftime('%Y-%m-%d')
                        break
                    except Exception:
                        continue
                if not fecha_servicio_iso and len(fs) >= 10:
                    # Último intento: tomar primeros 10 chars si parece ISO
                    candidato = fs[:10]
                    if candidato.count('-') == 2:
                        fecha_servicio_iso = candidato
    except Exception:
        fecha_servicio_iso = ''
    
    return render_template('facturacion/paciente_editar.html', 
                         paciente=paciente,
                         medicos=medicos,
                         ars_list=ars_list,
                         from_factura=from_factura,
                         ars_id_return=ars_id_return,
                         ncf_id_return=ncf_id_return,
                         medico_id_return=medico_id_return,
                         fecha_return=fecha_return,
                         fecha_actual=fecha_actual,
                         fecha_servicio_iso=fecha_servicio_iso)

@app.route('/facturacion/paciente/<int:paciente_id>/eliminar')
@login_required
def facturacion_paciente_eliminar(paciente_id):
    """Eliminar (DELETE real) un paciente pendiente de facturación"""
    conn = get_db_connection()
    
    # Verificar que el paciente existe
    paciente = conn.execute('''
        SELECT * FROM facturas_detalle WHERE id = %s
    ''', (paciente_id,)).fetchone()
    
    if not paciente:
        flash('❌ Paciente no encontrado', 'error')
        conn.close()
        return redirect(url_for('facturacion_pacientes_pendientes'))
    
    # Verificar que NO esté facturado (no tiene factura_id)
    if paciente['factura_id'] is not None:
        flash('❌ No se puede eliminar un registro que ya está en una factura', 'error')
        conn.close()
        return redirect(url_for('facturacion_pacientes_pendientes'))
    
    # Verificar que esté pendiente
    if paciente['estado'] != 'pendiente':
        flash('❌ Solo se pueden eliminar registros pendientes', 'error')
        conn.close()
        return redirect(url_for('facturacion_pacientes_pendientes'))
    
    try:
        # ELIMINACIÓN FÍSICA (DELETE real)
        conn.execute('''
            DELETE FROM facturas_detalle WHERE id = %s
        ''', (paciente_id,))
        
        conn.commit()
        conn.close()
        
        flash('✅ Registro eliminado exitosamente', 'success')
    except Exception as e:
        conn.close()
        flash(f'❌ Error al eliminar: {str(e)}', 'error')
    
    return redirect(url_for('facturacion_pacientes_pendientes'))

# ========== APIs PARA AUTOCOMPLETAR ==========
@app.route('/api/facturacion/buscar-paciente/<nss>')
@app.route('/api/facturacion/buscar-paciente/<nss>/<int:ars_id>')
def api_buscar_paciente(nss, ars_id=None):
    """API para buscar paciente por NSS + ARS (llave única)"""
    conn = get_db_connection()
    
    # Si se proporciona ARS, buscar por NSS + ARS (llave única)
    if ars_id:
        paciente = conn.execute('''
            SELECT p.*, a.nombre_ars 
            FROM pacientes p
            LEFT JOIN ars a ON p.ars_id = a.id
            WHERE p.nss = %s AND p.ars_id = %s AND p.activo = 1
        ''', (nss, ars_id)).fetchone()
    else:
        # Si no se proporciona ARS, buscar solo por NSS (puede haber múltiples)
        paciente = conn.execute('''
            SELECT p.*, a.nombre_ars 
            FROM pacientes p
            LEFT JOIN ars a ON p.ars_id = a.id
            WHERE p.nss = %s AND p.activo = 1
            LIMIT 1
        ''', (nss,)).fetchone()
    
    conn.close()
    
    if paciente:
        return jsonify({
            'encontrado': True,
            'nombre': paciente['nombre'],
            'ars_id': paciente['ars_id'],
            'ars_nombre': paciente['nombre_ars'] if paciente['nombre_ars'] else ''
        })
    else:
        return jsonify({'encontrado': False})

@app.route('/api/facturacion/buscar-servicio')
def api_buscar_servicio():
    """API para buscar servicios (autocompletado)"""
    query = request.args.get('q', '')
    conn = get_db_connection()
    servicios = conn.execute('''
        SELECT id, descripcion, precio_base 
        FROM tipos_servicios 
        WHERE descripcion LIKE %s AND activo = 1 
        ORDER BY descripcion 
        LIMIT 10
    ''', (f'%{query}%',)).fetchall()
    conn.close()
    
    return jsonify([{
        'id': s['id'],
        'descripcion': s['descripcion'],
        'precio_base': s['precio_base']
    } for s in servicios])

# ==================== SEO ROUTES ====================
@app.route('/sitemap.xml')
def sitemap():
    """Genera sitemap.xml dinámico para SEO"""
    from datetime import datetime
    from flask import make_response
    
    # Páginas principales (optimizadas para SEO)
    base_url = _canonical_base_url()
    urls = [
        {'path': url_for('index'), 'priority': '1.0', 'changefreq': 'daily'},
        {'path': url_for('services'), 'priority': '0.9', 'changefreq': 'weekly'},
        {'path': url_for('aesthetic_treatments'), 'priority': '0.95', 'changefreq': 'weekly'},
        {'path': url_for('about'), 'priority': '0.85', 'changefreq': 'monthly'},
        {'path': url_for('testimonials'), 'priority': '0.8', 'changefreq': 'weekly'},
        {'path': url_for('contact'), 'priority': '0.9', 'changefreq': 'monthly'},
        {'path': url_for('request_appointment'), 'priority': '1.0', 'changefreq': 'daily'},
        {'path': url_for('seo_ginecologa_santo_domingo'), 'priority': '0.85', 'changefreq': 'weekly'},
    ]
    
    today = datetime.now().strftime('%Y-%m-%d')
    
    # XML válido (esto impacta directamente a Google Search Console)
    sitemap_xml = '<?xml version="1.0" encoding="UTF-8"?>\n'
    sitemap_xml += '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9" '
    sitemap_xml += 'xmlns:image="http://www.google.com/schemas/sitemap-image/1.1">\n'
    og_image = f"{base_url}{_canonicalize_path(url_for('static', filename='images/dra-shirley-profesional.jpg'))}"
    
    for url in urls:
        loc = f"{base_url}{_canonicalize_path(url.get('path', '/'))}"
        sitemap_xml += '  <url>\n'
        sitemap_xml += f'    <loc>{loc}</loc>\n'
        sitemap_xml += f'    <lastmod>{today}</lastmod>\n'
        sitemap_xml += f'    <changefreq>{url["changefreq"]}</changefreq>\n'
        sitemap_xml += f'    <priority>{url["priority"]}</priority>\n'
        sitemap_xml += '    <image:image>\n'
        sitemap_xml += f'      <image:loc>{og_image}</image:loc>\n'
        sitemap_xml += '    </image:image>\n'
        sitemap_xml += '  </url>\n'
    
    sitemap_xml += '</urlset>\n'
    
    response = make_response(sitemap_xml)
    response.headers['Content-Type'] = 'application/xml; charset=utf-8'
    response.headers['Cache-Control'] = 'public, max-age=3600'
    return response

@app.route('/robots.txt')
def robots_txt():
    """Archivo robots.txt dinámico para control de crawlers"""
    base_url = _canonical_base_url()
    
    # IMPORTANTE: NO bloquear /static/ porque ahí vive /static/sitemap.xml y otros recursos que Google necesita leer.
    content = f"""User-agent: *
Allow: /

# Permitir acceso a recursos estáticos
Allow: /static/
Allow: /images/
Allow: /css/
Allow: /js/

# Bloquear acceso a secciones privadas
Disallow: /admin/
Disallow: /login
Disallow: /cambiar-password-obligatorio
Disallow: /solicitar-recuperacion
Disallow: /recuperar-contrasena/
Disallow: /facturacion/

# Bloquear acceso a archivos sensibles
Disallow: /*.env
Disallow: /*.log
Disallow: /database.db
Disallow: /requirements.txt
Disallow: /Procfile

# Sitemap
Sitemap: {base_url}/sitemap.xml

# Crawl-delay para ser respetuosos (algunos bots lo ignoran)
Crawl-delay: 1
"""
    response = make_response(content)
    response.headers["Content-Type"] = "text/plain"
    response.headers['Cache-Control'] = 'public, max-age=3600'
    return response

# ====================
# Compatibilidad SEO / cachés antiguos
# - Algunos bots y Google piden /favicon.ico en la raíz
# - /favicon.ico debe existir aunque haya <link rel="icon"> en HTML
# ====================
@app.route('/favicon.ico')
def favicon_ico():
    """
    Algunos bots y navegadores piden /favicon.ico.
    Este repo solo incluye favicon.svg, así que redirigimos para evitar 404.
    """
    return redirect(url_for('static', filename='favicon.svg'), code=301)

@app.route('/static/logos/logo-dra-shirley.png')
def logo_dra_shirley_png():
    """Fallback: evitar 404 si falta el logo PNG en /static/logos/."""
    return redirect(url_for('static', filename='favicon.svg'), code=302)

# Forzar mimetype correcto para WebP en URLs específicas usadas por <picture>.
# Algunos entornos/proxies sirven .webp como octet-stream; esto evita fallos de render/caché.
def _send_webp_or_jpg(webp_filename: str, jpg_filename: str):
    webp_path = os.path.join(app.static_folder, 'images', webp_filename)
    if os.path.exists(webp_path):
        return send_file(webp_path, mimetype='image/webp')
    jpg_path = os.path.join(app.static_folder, 'images', jpg_filename)
    if os.path.exists(jpg_path):
        return send_file(jpg_path, mimetype='image/jpeg')

    # Fallback: devolver un SVG simple en vez de 404/500
    from flask import make_response
    title = "Imagen no disponible"
    subtitle = f"{jpg_filename} / {webp_filename}"
    svg = f"""<svg xmlns="http://www.w3.org/2000/svg" width="1200" height="630" viewBox="0 0 1200 630">
  <defs>
    <linearGradient id="g" x1="0" x2="1" y1="0" y2="1">
      <stop offset="0" stop-color="#F2E2E6"/>
      <stop offset="1" stop-color="#DBC4CA"/>
    </linearGradient>
  </defs>
  <rect width="1200" height="630" fill="url(#g)"/>
  <rect x="40" y="40" width="1120" height="550" rx="24" fill="#fff" opacity="0.85"/>
  <text x="80" y="160" font-family="Arial, sans-serif" font-size="48" fill="#4A4A4A">{title}</text>
  <text x="80" y="230" font-family="Arial, sans-serif" font-size="28" fill="#6B6B6B">{subtitle}</text>
  <text x="80" y="320" font-family="Arial, sans-serif" font-size="24" fill="#6B6B6B">
    Sube los archivos a /static/images/ para ver la imagen real.
  </text>
</svg>"""
    resp = make_response(svg)
    resp.headers["Content-Type"] = "image/svg+xml; charset=utf-8"
    resp.headers["Cache-Control"] = "no-store"
    return resp

@app.route('/static/images/97472.webp')
def image_97472_webp():
    return _send_webp_or_jpg('97472.webp', '97472.jpg')

@app.route('/static/images/dra-shirley-profesional.webp')
def image_dra_shirley_profesional_webp():
    return _send_webp_or_jpg('dra-shirley-profesional.webp', 'dra-shirley-profesional.jpg')

# ============================================================================
# GESTIÓN DE USUARIOS
# ============================================================================

@app.route('/admin/usuarios')
@login_required
def admin_usuarios():
    """Listar usuarios del sistema"""
    # Solo administradores pueden ver la lista de usuarios
    if current_user.perfil != 'Administrador':
        flash('No tienes permisos para acceder a esta sección', 'error')
        return redirect(url_for('admin'))
    
    conn = get_db_connection()
    usuarios = conn.execute('SELECT * FROM usuarios ORDER BY created_at DESC').fetchall()
    conn.close()
    
    return render_template('admin_usuarios.html', usuarios=usuarios)

@app.route('/admin/usuarios/nuevo', methods=['GET', 'POST'])
@login_required
def admin_usuarios_nuevo():
    """Crear nuevo usuario"""
    # Solo administradores pueden crear usuarios
    if current_user.perfil != 'Administrador':
        flash('No tienes permisos para acceder a esta sección', 'error')
        return redirect(url_for('admin'))
    
    if request.method == 'POST':
        nombre = sanitize_input(request.form.get('nombre', ''), 100)
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password_nuevo', '')  # Campo diferente para nuevo usuario
        perfil = request.form.get('perfil', '')
        
        # Validaciones
        if not nombre or not email or not password or not perfil:
            flash('Todos los campos son obligatorios', 'error')
            return redirect(url_for('admin_usuarios_nuevo'))
        
        if not validate_email(email):
            flash('Email inválido', 'error')
            return redirect(url_for('admin_usuarios_nuevo'))
        
        # Validación robusta de contraseña
        password_errors = validar_password_segura(password)
        if password_errors:
            flash(f'Contraseña no cumple los requisitos: {", ".join(password_errors)}', 'error')
            return redirect(url_for('admin_usuarios_nuevo'))
        
        if perfil not in ['Administrador', 'Nivel 2', 'Registro de Facturas']:
            flash('Perfil inválido', 'error')
            return redirect(url_for('admin_usuarios_nuevo'))
        
        # Verificar que el email no exista
        conn = get_db_connection()
        existe = conn.execute('SELECT id FROM usuarios WHERE email = %s', (email,)).fetchone()
        
        if existe:
            conn.close()
            flash('Ya existe un usuario con ese email', 'error')
            return redirect(url_for('admin_usuarios_nuevo'))
        
        # Crear usuario con contraseña temporal
        password_hash = generate_password_hash(password)
        conn.execute('''
            INSERT INTO usuarios (nombre, email, password_hash, perfil, activo, password_temporal)
            VALUES (%s, %s, %s, %s, 1, 1)
        ''', (nombre, email, password_hash, perfil))
        conn.commit()
        conn.close()
        
        # Enviar email de bienvenida si es usuario de facturación
        if perfil in ['Nivel 2', 'Registro de Facturas']:
            try:
                from email_templates import template_bienvenida_facturacion
                
                # Generar link del admin (usar el dominio de Railway o el actual)
                link_admin = request.url_root.rstrip('/') + url_for('admin')
                
                # Determinar si puede generar facturas
                puede_generar_facturas = (perfil == 'Nivel 2')
                
                # Generar HTML del email
                html_body = template_bienvenida_facturacion(
                    nombre=nombre,
                    email=email,
                    password_temporal=password,
                    link_admin=link_admin,
                    puede_generar_facturas=puede_generar_facturas
                )
                
                # Enviar email
                send_email(
                    destinatario=email,
                    asunto=f'🎉 Bienvenido al Sistema de Facturación - Dra. Shirley Ramírez',
                    cuerpo=html_body
                )
                
                flash(f'Usuario {nombre} creado exitosamente. Se ha enviado un email a {email} con las credenciales e instrucciones de acceso.', 'success')
            except Exception as e:
                print(f"Error al enviar email de bienvenida: {e}")
                flash(f'Usuario {nombre} creado exitosamente. Contraseña temporal: {password} (el usuario deberá cambiarla en el primer login). No se pudo enviar el email de bienvenida.', 'warning')
        else:
            flash(f'Usuario {nombre} creado exitosamente. Contraseña temporal: {password} (el usuario deberá cambiarla en el primer login)', 'success')
        
        return redirect(url_for('admin_usuarios'))
    
    return render_template('admin_usuarios_form.html', usuario=None)

@app.route('/admin/usuarios/<int:usuario_id>/editar', methods=['GET', 'POST'])
@login_required
def admin_usuarios_editar(usuario_id):
    """Editar usuario existente"""
    # Solo administradores pueden editar usuarios
    if current_user.perfil != 'Administrador':
        flash('No tienes permisos para acceder a esta sección', 'error')
        return redirect(url_for('admin'))
    
    conn = get_db_connection()
    usuario = conn.execute('SELECT * FROM usuarios WHERE id = %s', (usuario_id,)).fetchone()
    
    if not usuario:
        conn.close()
        flash('Usuario no encontrado', 'error')
        return redirect(url_for('admin_usuarios'))
    
    if request.method == 'POST':
        nombre = sanitize_input(request.form.get('nombre', ''), 100)
        email = request.form.get('email', '').strip().lower()
        perfil = request.form.get('perfil', '')
        activo = request.form.get('activo') == '1'
        cambiar_password = request.form.get('cambiar_password') == '1'
        password = request.form.get('password', '')
        enviar_email = request.form.get('enviar_email') == '1'
        
        # Validaciones
        if not nombre or not email or not perfil:
            flash('Nombre, email y perfil son obligatorios', 'error')
            return redirect(url_for('admin_usuarios_editar', usuario_id=usuario_id))
        
        if not validate_email(email):
            flash('Email inválido', 'error')
            return redirect(url_for('admin_usuarios_editar', usuario_id=usuario_id))
        
        if perfil not in ['Administrador', 'Nivel 2', 'Registro de Facturas']:
            flash('Perfil inválido', 'error')
            return redirect(url_for('admin_usuarios_editar', usuario_id=usuario_id))
        
        # No permitir que el usuario se desactive a sí mismo
        if usuario_id == current_user.id and not activo:
            flash('No puedes desactivar tu propia cuenta', 'error')
            return redirect(url_for('admin_usuarios_editar', usuario_id=usuario_id))
        
        # Verificar que el email no exista en otro usuario
        existe = conn.execute('SELECT id FROM usuarios WHERE email = %s AND id != %s',
                             (email, usuario_id)).fetchone()
        
        if existe:
            conn.close()
            flash('Ya existe otro usuario con ese email', 'error')
            return redirect(url_for('admin_usuarios_editar', usuario_id=usuario_id))
        
        # Actualizar usuario
        if cambiar_password and password:
            if len(password) < 8:
                flash('La contraseña debe tener al menos 8 caracteres', 'error')
                return redirect(url_for('admin_usuarios_editar', usuario_id=usuario_id))
            
            password_hash = generate_password_hash(password)
            conn.execute('''
                UPDATE usuarios 
                SET nombre = %s, email = %s, password_hash = %s, perfil = %s, activo = %s, password_temporal = 1
                WHERE id = %s
            ''', (nombre, email, password_hash, perfil, activo, usuario_id))
            
            # IMPORTANTE: Si el usuario tiene sesión activa, cerrarla automáticamente
            # Esto fuerza que use la nueva contraseña temporal
            from flask import session as flask_session
            from flask_login import logout_user
            
            # Si es el mismo usuario que está editando, cerrar su sesión
            if usuario_id == current_user.id:
                conn.commit()
                conn.close()
                logout_user()
                flash('Tu contraseña ha sido cambiada. Inicia sesión con la nueva contraseña temporal.', 'warning')
                return redirect(url_for('login'))
            
            # Si es otro usuario, invalidar su sesión (si existe)
            # Nota: Flask-Login no tiene una forma directa de cerrar sesiones de otros usuarios
            # Pero al marcar password_temporal=1, el sistema le pedirá cambiar contraseña en su próximo login
            
            # Enviar email con nueva contraseña si está marcado
            if enviar_email and EMAIL_CONFIGURED:
                def enviar_email_async():
                    try:
                        html = template_nueva_contrasena(nombre, email, password)
                        send_email_sendgrid(
                            to_email=email,
                            subject='🔐 Nueva Contraseña Temporal - Panel Administrativo',
                            html_content=html
                        )
                    except Exception as e:
                        print(f"Error al enviar email de nueva contraseña: {e}")
                
                # Enviar en segundo plano
                email_thread = threading.Thread(target=enviar_email_async)
                email_thread.start()
                
                flash(f'Usuario {nombre} actualizado exitosamente. Se ha enviado un email con la nueva contraseña.', 'success')
            else:
                flash(f'Usuario {nombre} actualizado exitosamente. Nueva contraseña configurada.', 'success')
        else:
            conn.execute('''
                UPDATE usuarios 
                SET nombre = %s, email = %s, perfil = %s, activo = %s
                WHERE id = %s
            ''', (nombre, email, perfil, activo, usuario_id))
            
            flash(f'Usuario {nombre} actualizado exitosamente', 'success')
        
        conn.commit()
        conn.close()
        
        return redirect(url_for('admin_usuarios'))
    
    conn.close()
    return render_template('admin_usuarios_form.html', usuario=usuario)

@app.route('/admin/usuarios/<int:usuario_id>/eliminar', methods=['POST'])
@login_required
def admin_usuarios_eliminar(usuario_id):
    """Eliminar usuario - DESHABILITADO POR SEGURIDAD"""
    flash('La eliminación de usuarios está deshabilitada por seguridad. Puedes desactivar usuarios en su lugar.', 'warning')
    return redirect(url_for('admin_usuarios'))

if __name__ == '__main__':
    # Inicializar base de datos
    init_db()
    create_sample_data()
    
    # Crear directorio de templates si no existe
    os.makedirs('templates', exist_ok=True)
    
    # Configuración para producción
    port = int(os.getenv('PORT', 5000))
    host = os.getenv('HOST', '0.0.0.0')
    debug = os.getenv('FLASK_ENV') != 'production'
    
    print("\n" + "="*60)
    print("🚀 SERVIDOR INICIADO")
    print("="*60)
    print(f"🌐 Entorno: {'PRODUCCIÓN' if not debug else 'DESARROLLO'}")
    print(f"🌐 Host: {host}:{port}")
    print(f"📧 Email configurado: {'✅ SÍ' if EMAIL_PASSWORD and EMAIL_PASSWORD != 'tu_password_aqui' else '❌ NO'}")
    print(f"📄 PDF disponible: {'✅ SÍ' if REPORTLAB_AVAILABLE else '❌ NO'}")
    print(f"🔒 Seguridad: {'✅ ACTIVADA' if not debug else '⚠️ DESARROLLO'}")
    print("="*60 + "\n")

    
    
    
    
    app.run(debug=debug, host=host, port=port)
    
    

 
 
 
 
